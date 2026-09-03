import os
import re
import json
import sqlite3
import statistics
import unicodedata
import uuid
import zipfile
from datetime import datetime, timezone
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from collections.abc import Mapping
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from pypdf import PdfReader
import streamlit as st
from google import genai
from google.genai import types
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel, Field

try:
    import psycopg
    from psycopg.rows import dict_row
except Exception:
    psycopg = None
    dict_row = None


# =========================================================
# CONFIGURACIÓN GENERAL
# =========================================================

st.set_page_config(
    page_title="Sistema de Presupuestación Asistida",
    page_icon=None,
    layout="wide",
)


# =========================================================
# UTILIDADES
# =========================================================


def ahora_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def get_secret(nombre: str, default=None):
    try:
        if nombre in st.secrets:
            return st.secrets[nombre]
    except Exception:
        pass
    return os.getenv(nombre, default)


def normalizar_texto(texto) -> str:
    """
    Normaliza cualquier valor recibido, no solamente cadenas.

    Al leer archivos Excel antiguos pueden aparecer números, booleanos, fechas,
    resultados de fórmulas o valores vacíos en las mismas filas que se recorren
    buscando encabezados. Convertir primero a texto evita errores como:
    AttributeError: 'float' object has no attribute 'strip'
    """
    if texto is None:
        texto = ""
    elif not isinstance(texto, str):
        texto = str(texto)

    texto = texto.strip().lower()
    texto = "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )
    texto = re.sub(r"[^a-z0-9\s]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def abreviar(texto: str, longitud: int = 3) -> str:
    limpio = normalizar_texto(texto).upper()
    palabras = [p for p in limpio.split() if p not in {"DE", "DEL", "LA", "EL", "EN", "Y"}]
    if not palabras:
        return "PRY"
    base = palabras[0]
    if len(base) >= longitud:
        return base[:longitud]
    return (base + "XXX")[:longitud]


def abreviar_cliente(texto: str, max_len: int = 4) -> str:
    """
    Genera una abreviatura estable para el cliente.
    Si hay varias palabras utiliza sus iniciales; si hay una sola, usa sus
    primeras letras. Ej.: "Desarrollos de la Vega" -> "DDV".
    """
    limpio = normalizar_texto(texto).upper()
    palabras = [
        p for p in limpio.split()
        if p not in {"DE", "DEL", "LA", "LAS", "EL", "LOS", "EN", "Y", "SA", "CV"}
    ]
    if not palabras:
        return "CLI"
    if len(palabras) >= 2:
        iniciales = "".join(p[0] for p in palabras if p)
        return iniciales[:max_len] or "CLI"
    return (palabras[0] + "XXXX")[:3]


def abreviacion_tipo(tipo: str) -> str:
    mapa = {
        "Baño": "BAN",
        "Cocina": "COC",
        "Recámara": "REC",
        "Sala / comedor": "SAL",
        "Local comercial": "LOC",
        "Oficina": "OFI",
        "Remodelación interior general": "REM",
        "Caseta / acceso": "CAS",
        "Otro": "OTR",
    }
    return mapa.get(tipo, abreviar(tipo))


def formato_moneda(valor: float) -> str:
    return f"${valor:,.2f}"


def score_similitud(a: str, b: str) -> float:
    a_n = normalizar_texto(a)
    b_n = normalizar_texto(b)
    if not a_n or not b_n:
        return 0.0

    secuencia = SequenceMatcher(None, a_n, b_n).ratio()
    ta = set(a_n.split())
    tb = set(b_n.split())
    union = ta | tb
    jaccard = len(ta & tb) / len(union) if union else 0.0
    return 0.65 * secuencia + 0.35 * jaccard


def limpiar_codigo(codigo: str, fallback: str) -> str:
    codigo = re.sub(r"[^A-Za-z0-9\-]", "", (codigo or "").upper())
    return codigo[:24] if codigo else fallback


SECCIONES_COMERCIALES_PREFERENTES = [
    # Orden de presentación basado en la secuencia normal de preparación y obra.
    "PROYECTO Y TRÁMITES",
    "PRELIMINARES Y PROTECCIONES",
    "DESMONTAJES Y DEMOLICIONES",
    "ALBAÑILERÍA Y ESTRUCTURA",
    "INSTALACIONES HIDROSANITARIAS",
    "INSTALACIONES ELÉCTRICAS",
    "ACABADOS Y RECUBRIMIENTOS",
    "CARPINTERÍA",
    "CANCELERÍA Y HERRERÍA",
    "EXTERIORES Y AMENIDADES",
    "OTROS TRABAJOS",
    "LIMPIEZA Y ENTREGA",
]

SECCION_ALIASES = {
    "PROYECTO Y TRAMITES": "PROYECTO Y TRÁMITES",
    "PROYECTO": "PROYECTO Y TRÁMITES",
    "TRAMITES": "PROYECTO Y TRÁMITES",
    "PRELIMINARES": "PRELIMINARES Y PROTECCIONES",
    "PRELIMINARES Y PROTECCIONES": "PRELIMINARES Y PROTECCIONES",
    "PROTECCIONES": "PRELIMINARES Y PROTECCIONES",
    "PREPARACION": "PRELIMINARES Y PROTECCIONES",
    "PREPARACION Y DEMOLICIONES": "DESMONTAJES Y DEMOLICIONES",
    "PRELIMINARES Y DEMOLICIONES": "DESMONTAJES Y DEMOLICIONES",
    "DEMOLICIONES": "DESMONTAJES Y DEMOLICIONES",
    "DESMONTAJES": "DESMONTAJES Y DEMOLICIONES",
    "ALBANILERIA": "ALBAÑILERÍA Y ESTRUCTURA",
    "ESTRUCTURA": "ALBAÑILERÍA Y ESTRUCTURA",
    "ALBANILERIA Y ESTRUCTURA": "ALBAÑILERÍA Y ESTRUCTURA",
    "ELECTRICA": "INSTALACIONES ELÉCTRICAS",
    "INSTALACION ELECTRICA": "INSTALACIONES ELÉCTRICAS",
    "INSTALACIONES ELECTRICAS": "INSTALACIONES ELÉCTRICAS",
    "HIDROSANITARIA": "INSTALACIONES HIDROSANITARIAS",
    "INSTALACIONES HIDROSANITARIAS": "INSTALACIONES HIDROSANITARIAS",
    "PLOMERIA": "INSTALACIONES HIDROSANITARIAS",
    "ACABADOS": "ACABADOS Y RECUBRIMIENTOS",
    "ACABADOS Y RECUBRIMIENTOS": "ACABADOS Y RECUBRIMIENTOS",
    "CARPINTERIA": "CARPINTERÍA",
    "HERRERIA Y CANCELERIA": "CANCELERÍA Y HERRERÍA",
    "CANCELERIA Y HERRERIA": "CANCELERÍA Y HERRERÍA",
    "EXTERIORES": "EXTERIORES Y AMENIDADES",
    "EXTERIORES Y AMENIDADES": "EXTERIORES Y AMENIDADES",
    "LIMPIEZA": "LIMPIEZA Y ENTREGA",
    "LIMPIEZA FINAL": "LIMPIEZA Y ENTREGA",
    "LIMPIEZA Y ENTREGA": "LIMPIEZA Y ENTREGA",
}


def normalizar_seccion_comercial(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return "OTROS TRABAJOS"
    key = normalizar_texto(raw).upper()
    return SECCION_ALIASES.get(key, raw.upper())


def titulo_comercial_item(item: dict) -> str:
    value = str(item.get("commercial_title") or "").strip()
    if value:
        return value
    fallback = str(item.get("subcategory") or "").strip()
    if fallback:
        return fallback
    description = str(item.get("description") or "").strip()
    if not description:
        return "Concepto"
    first = re.split(r"[.;:]", description, maxsplit=1)[0].strip()
    return first[:80] or description[:80]


def seccion_ejecucion_item(item: dict) -> str:
    """
    Clasificación final de seguridad.

    La clasificación principal la revisa Gemini viendo el presupuesto completo.
    Python solo corrige dos casos inequívocos de secuencia:
    - protección TEMPORAL de áreas de obra -> preliminares;
    - limpieza FINA/FINAL de entrega -> cierre.

    No se usan palabras aisladas como "protección", porque pueden describir una
    propiedad técnica de un elemento permanente y provocar falsos positivos.
    """
    category = normalizar_seccion_comercial(item.get("category"))

    title_context = normalizar_texto(
        " ".join(
            str(item.get(k) or "")
            for k in ("subcategory", "commercial_title")
        )
    ).upper()

    preliminary_phrases = (
        "PROTECCION DE AREAS",
        "PROTECCION DE AREA",
        "PROTECCION DE PISOS",
        "PROTECCION DE PISO",
        "PROTECCION DE ACCESOS",
        "PROTECCION DE MOBILIARIO",
        "PROTECCION TEMPORAL",
        "PROTECCIONES TEMPORALES",
        "CUBRIR AREAS",
        "CUBRIR PISOS",
        "TAPIAL DE OBRA",
        "TAPIALES DE OBRA",
        "TRAZO Y REPLANTEO",
        "REPLANTEO",
    )
    if any(phrase in title_context for phrase in preliminary_phrases):
        return "PRELIMINARES Y PROTECCIONES"

    closing_phrases = (
        "LIMPIEZA FINA",
        "LIMPIEZA FINAL",
        "LIMPIEZA DE ENTREGA",
        "LIMPIEZA Y ENTREGA",
        "ASEO FINAL",
        "ENTREGA FINAL",
        "CIERRE DE OBRA",
    )
    if any(phrase in title_context for phrase in closing_phrases):
        return "LIMPIEZA Y ENTREGA"

    return category


def ordenar_items_comercialmente(items: list[dict]) -> list[dict]:
    """
    Ordena por fase macro de obra y después por orden_ejecucion auditado.
    """
    preferred = {
        name: index for index, name in enumerate(SECCIONES_COMERCIALES_PREFERENTES)
    }

    def key(pair):
        idx, item = pair
        section = seccion_ejecucion_item(item)
        try:
            execution_order = int(item.get("execution_order") or 500)
        except (TypeError, ValueError):
            execution_order = 500
        phase = preferred.get(
            section,
            preferred.get("OTROS TRABAJOS", 10),
        )
        return (phase, execution_order, idx)

    ordered = [dict(item) for _, item in sorted(enumerate(items), key=key)]
    for item in ordered:
        item["category"] = seccion_ejecucion_item(item)
    return ordered


def nombre_partida_excel(section: str) -> str:
    """Nombre breve de partida para el Excel, en orden de ejecución."""
    section = normalizar_seccion_comercial(section)
    preferred_names = {
        "PROYECTO Y TRÁMITES": "Trámites",
        "PRELIMINARES Y PROTECCIONES": "Preliminares y Protecciones",
        "DESMONTAJES Y DEMOLICIONES": "Desmontajes y Demoliciones",
        "ALBAÑILERÍA Y ESTRUCTURA": "Albañilería y Estructura",
        "INSTALACIONES HIDROSANITARIAS": "Instalaciones Hidrosanitarias",
        "INSTALACIONES ELÉCTRICAS": "Instalaciones Eléctricas",
        "ACABADOS Y RECUBRIMIENTOS": "Acabados",
        "CARPINTERÍA": "Carpintería",
        "CANCELERÍA Y HERRERÍA": "Cancelería y Herrería",
        "EXTERIORES Y AMENIDADES": "Exteriores y Amenidades",
        "LIMPIEZA Y ENTREGA": "Limpieza y Entrega",
        "OTROS TRABAJOS": "Otros Trabajos",
    }
    if section in preferred_names:
        return preferred_names[section]
    return section.title()


def nombre_subpartida_excel(item: dict) -> str:
    """
    La subpartida del Excel debe ser corta: Pisos, Muros, Frentes, Barra, etc.
    Para datos históricos se usa título comercial como respaldo.
    """
    value = str(item.get("subcategory") or "").strip()
    if value:
        return value
    return titulo_comercial_item(item)


def estructura_partidas_excel(items: list[dict]) -> list[dict]:
    """
    Asigna numeración jerárquica estable según el orden comercial:
      1. Trámites       / 1.1 Licencias
      2. Acabados       / 2.1 Pisos
      3. Carpintería    / 3.1 Frentes
    La numeración se genera en Python y no se deja a criterio de Gemini.
    """
    ordered = ordenar_items_comercialmente(items)
    section_numbers = {}
    section_counts = {}
    output = []

    for item in ordered:
        section = seccion_ejecucion_item(item)

        if section not in section_numbers:
            section_numbers[section] = len(section_numbers) + 1
            section_counts[section] = 0

        part_num = section_numbers[section]
        section_counts[section] += 1
        sub_num = section_counts[section]

        enriched = dict(item)
        enriched["part_number"] = part_num
        enriched["subpart_number"] = sub_num
        enriched["partida_excel"] = f"{part_num}. {nombre_partida_excel(section)}"
        enriched["subpartida_excel"] = (
            f"{part_num}.{sub_num} {nombre_subpartida_excel(item)}"
        )
        output.append(enriched)

    return output


def descripcion_excel_item(item: dict) -> str:
    """
    El ejemplo de la empresa coloca una descripción técnica amplia en una sola
    celda. Anteponemos el título comercial cuando aporta contexto y no está ya
    incluido al inicio de la descripción.
    """
    title = titulo_comercial_item(item).strip()
    description = str(item.get("description") or "").strip()

    if not title:
        return description
    if not description:
        return title

    norm_title = normalizar_texto(title).upper()
    norm_desc = normalizar_texto(description).upper()
    if norm_desc.startswith(norm_title):
        return description
    return f"{title}. {description}"


AREA_GENERAL = "General"
BRAND_MARKUP_PCT = 30.0


PATRONES_AREAS_EXPLICITAS = [
    # Espacios con calificadores: primero los patrones más específicos.
    r"\bmedio\s+bañ[oó](?:\s+(?:de|en)\s+planta\s+baja)?\b",
    r"\bbañ[oó]\s+de\s+visita(?:s)?\b",
    r"\bbañ[oó]\s+principal\b",
    r"\bbañ[oó]\s+secundari[oa]\b",
    r"\bbañ[oó]\s+social\b",
    r"\bbañ[oó]\s+[1-9]\b",
    r"\bbañ[oó]s?\b",
    r"\brec[aá]mara\s+principal\b",
    r"\brec[aá]mara\s+secundaria\b",
    r"\brec[aá]mara\s+[1-9]\b",
    r"\brec[aá]mara\b",
    r"\bhabitaci[oó]n\s+principal\b",
    r"\bhabitaci[oó]n\s+secundaria\b",
    r"\bhabitaci[oó]n\s+[1-9]\b",
    r"\bhabitaci[oó]n\b",
    r"\broof\s*(?:top|garden)\b",
    r"\b(?:cuarto|área|area)\s+de\s+lavado\b",
    r"\b(?:walk[\s-]*in\s+closet|walking\s+closet|vestidor)\b",
    r"\bcl[oó]set\s+de\s+blancos\b",
    r"\bdespacho\b",
    r"\boficina\b",
    r"\bcocina\b",
    r"\bsala(?:\s*(?:/|y)\s*comedor)?\b",
    r"\bcomedor\b",
    r"\bestudio(?:\s+[1-9])?\b",
    r"\bpatio(?:\s+(?:trasero|posterior|frontal|delantero))?\b",
    r"\bazotea\b",
    r"\bterraza\b",
    r"\bbalc[oó]n\b",
    r"\bfachada\b",
    r"\bescalera\b",
    r"\blavander[ií]a\b",
    r"\bbodega\b",
    r"\bestacionamiento(?:\s+[1-9])?\b",
    r"\bcochera(?:\s+[1-9])?\b",
    r"\bjard[ií]n\b",
    r"\bpasillo\b",
]


def normalizar_nombre_area(value: str) -> str:
    raw = re.sub(r"\s+", " ", str(value or "").strip())
    if not raw:
        return AREA_GENERAL

    key = normalizar_texto(raw).upper()
    if key in {
        "GENERAL", "GENERALES", "AREA GENERAL", "AREAS GENERALES",
        "TODA LA OBRA", "TODO EL PROYECTO",
    }:
        return AREA_GENERAL

    display = raw.replace("area ", "Área ").replace("Area ", "Área ")
    words = []
    for word in display.split():
        if any(ch.isdigit() for ch in word):
            words.append(word)
        elif word.upper() in {"UV"}:
            words.append(word.upper())
        else:
            words.append(word[:1].upper() + word[1:].lower())
    return " ".join(words)


def tipo_base_area(area: str) -> str:
    key = normalizar_texto(area).upper()
    equivalencias = [
        ("BANO", "BANO"), ("COCINA", "COCINA"), ("SALA", "SALA"),
        ("COMEDOR", "COMEDOR"), ("RECAMARA", "DORMITORIO"),
        ("HABITACION", "DORMITORIO"), ("ESTUDIO", "ESTUDIO"),
        ("PATIO", "PATIO"), ("AZOTEA", "AZOTEA"),
        ("ROOF GARDEN", "ROOF GARDEN"), ("TERRAZA", "TERRAZA"),
        ("BALCON", "BALCON"), ("FACHADA", "FACHADA"),
        ("ESCALERA", "ESCALERA"), ("CUARTO DE LAVADO", "LAVADO"),
        ("AREA DE LAVADO", "LAVADO"), ("LAVANDERIA", "LAVADO"),
        ("BODEGA", "BODEGA"), ("WALK IN CLOSET", "VESTIDOR"),
        ("WALKING CLOSET", "VESTIDOR"), ("VESTIDOR", "VESTIDOR"),
        ("ESTACIONAMIENTO", "ESTACIONAMIENTO"), ("COCHERA", "ESTACIONAMIENTO"),
        ("JARDIN", "JARDIN"), ("PASILLO", "PASILLO"),
        ("MEDIO BANO", "BANO"), ("DESPACHO", "DESPACHO"),
        ("OFICINA", "DESPACHO"), ("CLOSET DE BLANCOS", "CLOSET"),
        ("ROOF TOP", "ROOF TOP"),
    ]
    for prefix, base in equivalencias:
        if key.startswith(prefix):
            return base
    return key


def detectar_areas_explicitas(project_data: dict) -> list[str]:
    """Detecta áreas solo en la descripción original escrita por el usuario."""
    source = str(project_data.get("description") or "")
    if not source.strip():
        return []

    matches = []
    occupied = []
    for pattern in PATRONES_AREAS_EXPLICITAS:
        for match in re.finditer(pattern, source, flags=re.I):
            span = match.span()
            if any(not (span[1] <= a or span[0] >= b) for a, b in occupied):
                continue
            occupied.append(span)
            matches.append((span[0], normalizar_nombre_area(match.group(0))))

    matches.sort(key=lambda x: x[0])
    output, seen = [], set()
    for _, area in matches:
        key = normalizar_texto(area).upper()
        if key not in seen:
            seen.add(key)
            output.append(area)
    return output


def _texto_item_para_area(item: dict) -> str:
    return " ".join(
        str(item.get(key) or "")
        for key in (
            "subcategory", "commercial_title", "description",
            "quantity_criterion", "inclusion_basis",
        )
    )


def _areas_mencionadas_en_item(project_data: dict, item: dict) -> list[str]:
    areas = detectar_areas_explicitas(project_data)
    if not areas:
        return []

    item_key = normalizar_texto(_texto_item_para_area(item)).upper()
    exact = []
    for area in areas:
        area_key = normalizar_texto(area).upper()
        if area_key and area_key in item_key:
            exact.append(area)
    if exact:
        return exact

    by_type = {}
    for area in areas:
        by_type.setdefault(tipo_base_area(area), []).append(area)

    generic = []
    tokens_by_type = {
        "BANO": ("BANO",), "COCINA": ("COCINA",), "SALA": ("SALA",),
        "COMEDOR": ("COMEDOR",), "DORMITORIO": ("RECAMARA", "HABITACION"),
        "ESTUDIO": ("ESTUDIO",), "PATIO": ("PATIO",), "AZOTEA": ("AZOTEA",),
        "ROOF GARDEN": ("ROOF GARDEN",), "TERRAZA": ("TERRAZA",),
        "BALCON": ("BALCON",), "FACHADA": ("FACHADA",),
        "ESCALERA": ("ESCALERA",), "LAVADO": ("LAVADO", "LAVANDERIA"),
        "BODEGA": ("BODEGA",), "VESTIDOR": ("VESTIDOR", "CLOSET"),
        "ESTACIONAMIENTO": ("ESTACIONAMIENTO", "COCHERA"),
        "JARDIN": ("JARDIN",), "PASILLO": ("PASILLO",),
    }
    for base, candidates in by_type.items():
        if len(candidates) != 1:
            continue
        if any(token in item_key for token in tokens_by_type.get(base, (base,))):
            generic.append(candidates[0])
    return generic


def _m2_explicitos_cerca_de_area(project_data: dict, area: str) -> float | None:
    source = str(project_data.get("description") or "")
    source_norm = normalizar_texto(source).upper()
    area_norm = normalizar_texto(area).upper()
    start = source_norm.find(area_norm)
    if start < 0:
        return None

    # Cortar la búsqueda antes de la siguiente área explícita. Así una medida de
    # Cocina no puede terminar asignándose accidentalmente a Baño 1, por ejemplo.
    next_starts = []
    for other in detectar_areas_explicitas(project_data):
        other_norm = normalizar_texto(other).upper()
        pos = source_norm.find(other_norm, start + len(area_norm))
        if pos > start:
            next_starts.append(pos)
    end = min(next_starts) if next_starts else min(len(source), start + 120)
    end = min(end, start + 120)
    window = source[start:end]

    direct = re.search(
        r"(\d+(?:[.,]\d+)?)\s*(?:m2|m²|metros?\s+cuadrados?)",
        window, flags=re.I,
    )
    if direct:
        return float(direct.group(1).replace(",", "."))

    dims = re.search(
        r"(\d+(?:[.,]\d+)?)\s*[x×]\s*(\d+(?:[.,]\d+)?)\s*m\b",
        window, flags=re.I,
    )
    if dims:
        return float(dims.group(1).replace(",", ".")) * float(dims.group(2).replace(",", "."))
    return None


def asignar_areas_deterministicamente(project_data: dict, item: dict) -> list[dict]:
    """
    1) una sola área explícita -> 100 %;
    2) varias áreas + concepto M2 + m² explícitos conciliables -> proporcional;
    3) cualquier caso ambiguo -> General 100 %.
    """
    matches = _areas_mencionadas_en_item(project_data, item)

    if len(matches) == 1:
        return [{
            "area": matches[0], "porcentaje": 100.0,
            "cantidad_referencia": float(item.get("quantity") or 0.0),
            "criterio": "Área indicada explícitamente en el alcance del concepto.",
            "confianza": "Alta",
        }]

    if len(matches) > 1 and normalizar_unidad(item.get("unit")) == "M2":
        measures = {area: _m2_explicitos_cerca_de_area(project_data, area) for area in matches}
        if all(v is not None and v > 0 for v in measures.values()):
            total_measure = sum(measures.values())
            item_qty = float(item.get("quantity") or 0.0)
            tolerance = max(1.0, item_qty * 0.15)
            if item_qty > 0 and abs(total_measure - item_qty) <= tolerance:
                result, accumulated = [], 0.0
                for idx, area in enumerate(matches):
                    if idx == len(matches) - 1:
                        pct = max(100.0 - accumulated, 0.0)
                    else:
                        pct = round(measures[area] / total_measure * 100.0, 6)
                        accumulated += pct
                    result.append({
                        "area": area, "porcentaje": pct,
                        "cantidad_referencia": measures[area],
                        "criterio": "Reparto calculado con m² explícitos del texto inicial.",
                        "confianza": "Alta",
                    })
                return result

    return [{
        "area": AREA_GENERAL, "porcentaje": 100.0,
        "cantidad_referencia": None,
        "criterio": "No existe una asignación por área verificable con los datos iniciales.",
        "confianza": "Alta",
    }]


def normalizar_asignaciones_area(asignaciones) -> list[dict]:
    if isinstance(asignaciones, str):
        try:
            asignaciones = json.loads(asignaciones)
        except Exception:
            asignaciones = []

    rows = []
    for raw in asignaciones or []:
        if hasattr(raw, "model_dump"):
            raw = raw.model_dump()
        if not isinstance(raw, dict):
            continue
        try:
            pct = max(float(raw.get("porcentaje") or 0.0), 0.0)
        except (TypeError, ValueError):
            pct = 0.0
        if pct <= 0:
            continue
        rows.append({
            "area": normalizar_nombre_area(raw.get("area")),
            "porcentaje": pct,
            "cantidad_referencia": raw.get("cantidad_referencia"),
            "criterio": str(raw.get("criterio") or ""),
            "confianza": str(raw.get("confianza") or "Alta"),
        })

    if not rows:
        return [{
            "area": AREA_GENERAL, "porcentaje": 100.0,
            "cantidad_referencia": None,
            "criterio": "Sin asignación verificable.", "confianza": "Alta",
        }]

    total, running = sum(x["porcentaje"] for x in rows), 0.0
    for idx, row in enumerate(rows):
        if idx == len(rows) - 1:
            row["porcentaje"] = max(100.0 - running, 0.0)
        else:
            row["porcentaje"] = round(row["porcentaje"] / total * 100.0, 6)
            running += row["porcentaje"]
    return rows


def obtener_asignaciones_area_item(item: dict) -> list[dict]:
    if item.get("area_allocations"):
        return normalizar_asignaciones_area(item.get("area_allocations"))
    if item.get("area_allocations_json"):
        return normalizar_asignaciones_area(item.get("area_allocations_json"))
    return normalizar_asignaciones_area([])


def recalcular_areas_items(project_data: dict, items: list[dict]) -> list[dict]:
    output = []
    for item in items:
        out = dict(item)
        out["area_allocations"] = asignar_areas_deterministicamente(project_data, out)
        output.append(out)
    return output


def descripcion_areas_item(item: dict) -> str:
    return " · ".join(x["area"] for x in obtener_asignaciones_area_item(item))


def _detectar_areas_en_texto(texto: str) -> list[str]:
    """Extrae áreas explícitas de cualquier texto técnico/comercial."""
    pseudo = {"description": str(texto or "")}
    return detectar_areas_explicitas(pseudo)


def resolver_area_actividad(project_data: dict, area_propuesta: str, titulo: str = "", descripcion: str = "", subpartida: str = "") -> str:
    allowed=areas_validas_proyecto(project_data)
    specific=[x for x in allowed if x!=AREA_GENERAL]
    proposed=normalizar_nombre_area(area_propuesta)
    pkey=normalizar_texto(proposed).upper()
    for known in specific:
        if normalizar_texto(known).upper()==pkey:
            return known
    if proposed!=AREA_GENERAL:
        base=tipo_base_area(proposed)
        matches=[x for x in specific if tipo_base_area(x)==base]
        if len(matches)==1: return matches[0]
    activity=normalizar_texto(" ".join(str(x or "") for x in (titulo,descripcion,subpartida))).upper()
    exact=[]
    for known in specific:
        k=normalizar_texto(known).upper()
        if k and k in activity: exact.append(known)
    if len(exact)==1: return exact[0]
    type_matches=[]
    for known in specific:
        base=tipo_base_area(known)
        aliases={"BANO":("BANO",),"DORMITORIO":("RECAMARA","HABITACION"),"LAVADO":("LAVADO","LAVANDERIA"),"ESTACIONAMIENTO":("ESTACIONAMIENTO","COCHERA"),"VESTIDOR":("VESTIDOR","CLOSET")}.get(base,(base,))
        if any(token in activity for token in aliases): type_matches.append(known)
    unique=[]
    for x in type_matches:
        if x not in unique: unique.append(x)
    if len(unique)==1: return unique[0]
    return AREA_GENERAL



def validar_area_actividad(project_data: dict, area: str) -> str:
    """Compatibilidad con llamadas antiguas."""
    return resolver_area_actividad(project_data, area)


def area_item(item: dict) -> str:
    direct = str(item.get("area") or "").strip()
    if direct:
        return normalizar_nombre_area(direct)

    allocations = obtener_asignaciones_area_item(item)
    specific = []
    for allocation in allocations:
        value = allocation.get("area")
        if value and value != AREA_GENERAL and value not in specific:
            specific.append(value)
    return specific[0] if len(specific) == 1 else AREA_GENERAL


def item_included(item: dict) -> bool:
    value = item.get("included", True)
    if isinstance(value, str):
        return normalizar_texto(value).upper() not in {"NO", "FALSE", "0"}
    return bool(value)


def confianza_cantidad_verificada(item: dict) -> str:
    """
    Evita que una cantidad estimada aparezca como Alta solo porque Gemini lo dijo.
    Alta se reserva a cantidades directamente respaldadas por datos explícitos.
    """
    criterion = normalizar_texto(item.get("quantity_criterion") or "").upper()
    raw = normalizar_texto(item.get("quantity_confidence") or "Media").upper()

    uncertainty_terms = (
        "ESTIM", "APROX", "SUPUEST", "PROMED", "PROPORC",
        "REFERENCIA", "POR CONFIRMAR", "A CONFIRMAR",
    )
    explicit_terms = (
        "ESPECIFICAD", "INDICAD", "REPORTAD", "PROPORCIONAD",
        "LONGITUD", "SUPERFICIE", "DIMENSION", "MEDIDA",
        "CANTIDAD SOLICITADA",
    )

    if any(term in criterion for term in uncertainty_terms):
        return "Baja" if "SUPUEST" in criterion or "POR CONFIRMAR" in criterion else "Media"

    has_number = bool(re.search(r"\\d", criterion))
    has_explicit_basis = any(term in criterion for term in explicit_terms)

    if raw == "ALTA" and not (has_number or has_explicit_basis):
        return "Media"

    if raw == "BAJA":
        return "Baja"
    if raw == "ALTA":
        return "Alta"
    return "Media"


def confianza_precio_verificada(item: dict) -> str:
    """
    La fuente del precio manda sobre la autoevaluación de Gemini.

    Una estimación pura de IA nunca se presenta como confianza Alta.
    """
    source = normalizar_texto(item.get("price_source") or "").upper()
    declared = normalizar_texto(item.get("price_confidence") or "Media").upper()

    if source == "IA ESTIMADO":
        return "Baja"

    if source == "HISTORICO IA":
        return "Media" if declared == "ALTA" else (
            "Baja" if declared == "BAJA" else "Media"
        )

    if source in {"REFERENCIA CDMX", "HISTORICO EXTERNO"}:
        return "Alta" if declared == "ALTA" else "Media"

    if source == "BASE INTERNA":
        return "Alta" if declared == "ALTA" else "Media"

    if source == "EXCEL IMPORTADO":
        return "Media"

    return "Baja" if declared == "BAJA" else "Media"


def confianza_general_item(item: dict) -> str:
    rank = {"BAJA": 0, "MEDIA": 1, "ALTA": 2}
    reverse = {0: "Baja", 1: "Media", 2: "Alta"}
    q = rank[normalizar_texto(confianza_cantidad_verificada(item)).upper()]
    if item.get("price_evidence_score") is not None:
        p_label = etiqueta_confianza_score(float(item["price_evidence_score"]))
    else:
        p_label = confianza_precio_verificada(item)
    p = rank[normalizar_texto(p_label).upper()]
    return reverse[min(q, p)]


def consideraciones_consolidadas_item(item: dict) -> str:
    parts = []
    quantity = str(item.get("quantity_criterion") or "").strip()
    inclusion = str(item.get("inclusion_basis") or "").strip()
    considerations = str(item.get("considerations") or "").strip()

    if quantity:
        parts.append(f"Cantidad: {quantity}")
    if inclusion:
        parts.append(f"Inclusión: {inclusion}")
    if considerations:
        parts.append(considerations)
    evidence_score = item.get("price_evidence_score")
    if evidence_score is not None:
        parts.append(f"Evidencia de precio: {int(round(float(evidence_score)))} / 100")
    return " | ".join(parts)


def original_unit_sale_item(item: dict) -> float:
    value = item.get("original_unit_sale")
    if value is None:
        value = item.get("unit_sale")
    return float(value or 0.0)


MODELOS_GEMINI_V16 = [
    "gemini-3.8-flash",
    "gemini-3.7-flash",
    "gemini-3.6-flash",
    "gemini-3.5-flash",
    "gemini-3.5-flash-lite",
]


def modelos_gemini(model_name: str | None = None) -> list[str]:
    modelos = []
    for model in [model_name, *MODELOS_GEMINI_V16]:
        if model and model not in modelos:
            modelos.append(model)
    return modelos


def normalizar_lista_areas(values) -> list[str]:
    areas, seen = [], set()
    for raw in values or []:
        area = normalizar_nombre_area(raw)
        key = normalizar_texto(area).upper()
        if not key or key == "GENERAL":
            continue
        if key not in seen:
            seen.add(key)
            areas.append(area)
    return [AREA_GENERAL] + areas


def areas_validas_proyecto(project_data: dict) -> list[str]:
    detected = project_data.get("areas_detectadas")
    if detected:
        return normalizar_lista_areas(detected)
    return normalizar_lista_areas(detectar_areas_explicitas(project_data))


def score_confianza_precio(source: str, detail: str = "", refs_count: int = 0) -> int:
    source_n = normalizar_texto(source).upper()
    detail_n = normalizar_texto(detail).upper()
    if source_n == "BASE INTERNA":
        if any(x in detail_n for x in ("COSTO REAL", "COTIZADO", "VALIDADO")):
            return 98
        return 92
    if source_n == "REFERENCIA CDMX":
        return 85
    if source_n == "WEB FUNDAMENTADO":
        if refs_count >= 4: return 86
        if refs_count == 3: return 82
        if refs_count == 2: return 72
        return 55
    if source_n == "HISTORICO EXTERNO": return 72
    if source_n == "HISTORICO IA": return 50
    if source_n == "EXCEL IMPORTADO": return 65
    if source_n == "IA ESTIMADO": return 30
    return 50


def etiqueta_confianza_score(score: float) -> str:
    if score >= 85: return "Alta"
    if score >= 65: return "Media"
    return "Baja"


def extraer_formula_geometrica(texto: str, unit: str) -> float | None:
    raw = str(texto or "").replace(",", ".")
    normalized_unit = normalizar_unidad(unit)
    if normalized_unit == "M2":
        m = re.search(r"(\d+(?:\.\d+)?)\s*(?:m)?\s*[x×]\s*(\d+(?:\.\d+)?)\s*(?:m)?", raw, flags=re.I)
        if m:
            return float(m.group(1))*float(m.group(2))
    if normalized_unit == "M3":
        m = re.search(r"(\d+(?:\.\d+)?)\s*(?:m)?\s*[x×]\s*(\d+(?:\.\d+)?)\s*(?:m)?\s*[x×]\s*(\d+(?:\.\d+)?)\s*(?:m)?", raw, flags=re.I)
        if m:
            return float(m.group(1))*float(m.group(2))*float(m.group(3))
    return None


def verificar_cantidades_python(result: PresupuestoIA) -> PresupuestoIA:
    corrected=[]
    for act in result.actividades:
        formula_text = " ".join([str(act.criterio_cantidad or ""), str(act.descripcion_tecnica or "")])
        calculated = extraer_formula_geometrica(formula_text, act.unidad)
        if calculated is not None and calculated > 0:
            current=float(act.cantidad or 0.0)
            if current <= 0 or abs(calculated-current)/max(calculated,1e-9)>0.03:
                act=act.model_copy(update={
                    "cantidad": round(calculated,4),
                    "criterio_cantidad": f"{act.criterio_cantidad} | Verificación Python: {calculated:.4f} {act.unidad}",
                })
        corrected.append(act)
    return result.model_copy(update={"actividades": corrected})


def _median_filter(values: list[float]) -> list[float]:
    clean=[float(x) for x in values if x is not None and float(x)>0]
    if len(clean)<=2: return clean
    med=statistics.median(clean)
    filtered=[x for x in clean if med*0.45 <= x <= med*2.20]
    return filtered if len(filtered)>=2 else clean


def calcular_precio_web(investigation: InvestigacionPrecioWebIA, actividad: ActividadIA, params: dict) -> dict | None:
    unit=normalizar_unidad(actividad.unidad)
    refs=[]
    for ref in investigation.referencias:
        if not ref.compatible or ref.precio_unitario<=0: continue
        if normalizar_unidad(ref.unidad)!=unit: continue
        if normalizar_texto(ref.tipo_precio).upper()=="MATERIAL SOLO": continue
        refs.append(ref)
    if not refs: return None
    prices=_median_filter([r.precio_unitario for r in refs])
    if not prices: return None
    market_median=statistics.median(prices)
    direct_refs=[r.precio_unitario for r in refs if normalizar_texto(r.tipo_precio).upper()=="SUBCONTRATISTA"]
    if direct_refs:
        unit_cost=statistics.median(_median_filter(direct_refs))
        note="referencia explícita de subcontratista"
    else:
        factor=(1+params["indirect_pct"]/100.0)*(1+params["profit_pct"]/100.0)
        unit_cost=market_median/factor if factor>0 else market_median
        note="mediana de mercado instalado convertida a costo base equivalente"
    source_lines=[]
    for ref in refs[:5]:
        source_lines.append(f"{ref.fuente_nombre}: ${ref.precio_unitario:,.2f}/{ref.unidad} ({ref.fuente_url})")
    detail=f"{len(prices)} referencias web compatibles | mediana ${market_median:,.2f}/{actividad.unidad} | {note} | " + " ; ".join(source_lines)
    score=score_confianza_precio("WEB_FUNDAMENTADO",detail,refs_count=len(prices))
    return {"unit_cost":float(unit_cost),"source":"WEB_FUNDAMENTADO","source_detail":detail,"status":"REFERENCIA_WEB","confidence":etiqueta_confianza_score(score),"evidence_score":score,"references_count":len(prices)}


NIVELES_PRESUPUESTO = ["Económico", "Medio", "Medio-alto", "Alto"]


def criterio_nivel_presupuesto(nivel: str) -> str:
    criterios = {
        "Económico": (
            "Prioriza soluciones funcionales, comerciales y de costo contenido. "
            "Evita especificaciones premium salvo que el alcance las exija."
        ),
        "Medio": (
            "Utiliza soluciones comerciales de calidad media, durables y comunes "
            "en remodelación residencial y comercial."
        ),
        "Medio-alto": (
            "Utiliza especificaciones de buena calidad, acabados cuidados y soluciones "
            "comerciales superiores al promedio, sin llegar automáticamente a opciones premium."
        ),
        "Alto": (
            "Prioriza especificaciones, acabados y soluciones de gama alta cuando sean "
            "coherentes con el proyecto; identifica trabajos que requieran proveedor especializado."
        ),
    }
    return criterios.get(nivel, criterios["Medio-alto"])


def normalizar_composicion_costos(
    materiales_pct: float,
    mano_obra_pct: float,
    otros_pct: float,
) -> tuple[float, float, float]:
    """Normaliza la composición estimada para que sume 100 %."""
    material = max(float(materiales_pct or 0), 0.0)
    labor = max(float(mano_obra_pct or 0), 0.0)
    other = max(float(otros_pct or 0), 0.0)
    total = material + labor + other

    if total <= 0:
        # Para datos históricos sin desglose preferimos no inventar.
        return 0.0, 0.0, 100.0

    return (
        material / total * 100.0,
        labor / total * 100.0,
        other / total * 100.0,
    )


def aplicar_composicion_costo(item: dict) -> dict:
    """
    Calcula una descomposición informativa del costo integrado.
    No modifica el costo ni agrega nuevamente el desperdicio.
    """
    out = dict(item)
    material_pct, labor_pct, other_pct = normalizar_composicion_costos(
        out.get("material_share_pct", 0.0),
        out.get("labor_share_pct", 0.0),
        out.get("other_share_pct", 0.0),
    )
    waste_pct = max(float(out.get("waste_reference_pct", 0.0) or 0.0), 0.0)
    unit_cost = max(float(out.get("unit_cost", 0.0) or 0.0), 0.0)

    out["material_share_pct"] = material_pct
    out["labor_share_pct"] = labor_pct
    out["other_share_pct"] = other_pct
    out["waste_reference_pct"] = waste_pct
    out["material_unit_est"] = unit_cost * material_pct / 100.0
    out["labor_unit_est"] = unit_cost * labor_pct / 100.0
    out["other_unit_est"] = unit_cost * other_pct / 100.0
    out["waste_reference_unit"] = out["material_unit_est"] * waste_pct / 100.0
    return out


# =========================================================
# CATÁLOGOS EXTERNOS - CDMX
# =========================================================


CDMX_TABULADOR_PAGE = (
    "https://www.obras.cdmx.gob.mx/normas-tabulador/"
    "tabulador-general-de-precios-unitarios"
)

# Fallback oficial conocido al momento de construir esta versión. La app no
# depende de este enlace para siempre: primero intenta descubrir dinámicamente
# la edición más reciente publicada en la página oficial.
CDMX_FALLBACK_CATALOG = {
    "source": "CDMX",
    "version_label": "Mayo 2026",
    "year": 2026,
    "month": 5,
    "url": (
        "https://obras.cdmx.gob.mx/storage/app/media/"
        "Tabulador%20General%20de%20Precios%20Unitarios%20Mayo%202026/"
        "tabulador-general-de-precios-unitarios-del-gobierno-de-la-ciudad-"
        "de-mexico-actualizacion-de-mayo-2026.pdf"
    ),
}

SPANISH_MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "SETIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}
SPANISH_MONTH_NAMES = {
    value: key.title() for key, value in SPANISH_MONTHS.items()
    if key != "SETIEMBRE"
}

CDMX_HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; PresupuestadorEmpresa/1.0; "
        "+https://streamlit.io)"
    ),
    "Accept": "text/html,application/pdf,*/*",
}


def normalizar_unidad(unidad: str) -> str:
    # Los superíndices deben convertirse antes de normalizar texto; de lo
    # contrario "m²" podría quedar reducido a "m".
    raw = str(unidad or "").strip()
    raw = raw.replace("²", "2").replace("³", "3")
    value = normalizar_texto(raw).upper()
    value = (
        value.replace("M²", "M2")
        .replace("M³", "M3")
        .replace("MTS2", "M2")
        .replace("MTS3", "M3")
        .replace("METROS CUADRADOS", "M2")
        .replace("METRO CUADRADO", "M2")
        .replace("METROS CUBICOS", "M3")
        .replace("METRO CUBICO", "M3")
        .replace("METROS LINEALES", "ML")
        .replace("METRO LINEAL", "ML")
        .replace("PIEZAS", "PZA")
        .replace("PIEZA", "PZA")
        .replace("PZAS", "PZA")
        .replace("PZ", "PZA")
        .replace("KILOGRAMOS", "KG")
        .replace("KILOGRAMO", "KG")
        .replace("TONELADAS", "TON")
        .replace("TONELADA", "TON")
        .replace("LITROS", "L")
        .replace("LITRO", "L")
    )
    value = re.sub(r"[^A-Z0-9%]+", "", value)
    aliases = {
        "M2": "M2",
        "M3": "M3",
        "ML": "ML",
        "M": "M",
        "PZA": "PZA",
        "PZA.": "PZA",
        "KG": "KG",
        "TON": "TON",
        "L": "L",
        "LT": "L",
        "H": "H",
        "HR": "H",
        "HRA": "H",
        "HORA": "H",
        "DIA": "DIA",
        "MES": "MES",
        "LOTE": "LOTE",
        "JGO": "JGO",
        "JUEGO": "JGO",
        "PTO": "PTO",
        "PUNTO": "PTO",
        "SERV": "SERV",
        "SERVICIO": "SERV",
        "%": "%",
    }
    return aliases.get(value, value[:16])


def _mes_y_anio_desde_texto(texto: str) -> tuple[int | None, int | None]:
    normalized = normalizar_texto(texto).upper()
    year_match = re.search(r"\b(20\d{2})\b", normalized)
    year = int(year_match.group(1)) if year_match else None
    month = None
    for name, number in SPANISH_MONTHS.items():
        if re.search(rf"\b{name}\b", normalized):
            month = number
            break
    return month, year


def descubrir_ultimo_tabulador_cdmx(timeout: int = 30) -> dict:
    """
    Busca en la página oficial todos los enlaces PDF del Tabulador General y
    selecciona el de año/mes más reciente. Si la página cambia o falla, utiliza
    el último enlace oficial conocido como fallback.
    """
    try:
        response = requests.get(
            CDMX_TABULADOR_PAGE,
            headers=CDMX_HTTP_HEADERS,
            timeout=timeout,
        )
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        candidates = []
        for a in soup.find_all("a", href=True):
            href = str(a.get("href") or "").strip()
            label = " ".join(a.stripped_strings)
            combined = f"{label} {href}"
            normalized = normalizar_texto(combined).upper()

            if "TABULADOR GENERAL DE PRECIOS UNITARIOS" not in normalized:
                continue
            if "NOTA" in normalized or "ANEXO" in normalized:
                continue
            if ".PDF" not in normalized:
                continue

            month, year = _mes_y_anio_desde_texto(combined)
            if not year:
                continue

            # Una edición anual sin mes es anterior a sus actualizaciones
            # mensuales del mismo año.
            month_sort = month or 0
            candidates.append(
                {
                    "source": "CDMX",
                    "version_label": (
                        f"{SPANISH_MONTH_NAMES.get(month, 'Edición')} {year}"
                        if month
                        else f"Edición {year}"
                    ),
                    "year": year,
                    "month": month or 0,
                    "url": urljoin(CDMX_TABULADOR_PAGE, href),
                    "_sort": (year, month_sort),
                }
            )

        if candidates:
            latest = max(candidates, key=lambda x: x["_sort"])
            latest.pop("_sort", None)
            return latest
    except Exception:
        pass

    return dict(CDMX_FALLBACK_CATALOG)


def _parse_price(value: str) -> float | None:
    raw = str(value or "").strip()
    raw = raw.replace("$", "").replace(",", "").replace(" ", "")
    raw = raw.replace("\u00a0", "")
    if not re.fullmatch(r"\d+(?:\.\d+)?", raw):
        return None
    try:
        number = float(raw)
    except ValueError:
        return None
    if number <= 0 or number > 1_000_000_000:
        return None
    return number


def _looks_like_source_code(value: str) -> bool:
    value = str(value or "").strip().upper()
    if not value or len(value) > 45:
        return False
    if value in {"CLAVE", "PAGINA", "PÁGINA"}:
        return False
    if " " in value:
        return False
    # Los códigos del tabulador pueden contener letras, números, punto,
    # guion, diagonal y guion bajo.
    return bool(re.fullmatch(r"[A-Z0-9][A-Z0-9._/\-]{1,44}", value))


def _clean_pdf_cell(value) -> str:
    text = str(value or "")
    text = text.replace("\x00", " ").replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def _concept_from_cells(
    code: str,
    description: str,
    unit: str,
    price,
    chapter: str = "",
) -> dict | None:
    code = _clean_pdf_cell(code).upper()
    description = _clean_pdf_cell(description)
    unit = _clean_pdf_cell(unit).upper()
    numeric_price = _parse_price(price)

    if not _looks_like_source_code(code):
        return None
    if not description or len(description) < 8:
        return None
    normalized_desc = normalizar_texto(description)
    if any(
        bad in normalized_desc
        for bad in [
            "tabulador general de precios",
            "concepto de obra",
            "plaza de la constitucion",
            "gobierno de la ciudad",
        ]
    ):
        return None
    normalized_unit = normalizar_unidad(unit)
    if not normalized_unit or numeric_price is None:
        return None

    return {
        "source_code": code,
        "description": description,
        "normalized_description": normalized_desc,
        "unit": unit,
        "normalized_unit": normalized_unit,
        "unit_price": numeric_price,
        "chapter": _clean_pdf_cell(chapter),
    }


def _parse_layout_line(line: str) -> dict | None:
    """
    Parser principal para texto extraído en modo layout.
    El tabulador oficial utiliza cuatro columnas: Clave, Concepto, Unidad, P.U.
    """
    raw = line.rstrip()
    if not raw.strip():
        return None

    # Primero se aprovechan los bloques de 2+ espacios producidos por layout.
    parts = [p.strip() for p in re.split(r"\s{2,}", raw.strip()) if p.strip()]
    if len(parts) >= 4:
        code = parts[0]
        price = parts[-1]
        unit = parts[-2]
        description = " ".join(parts[1:-2])
        concept = _concept_from_cells(code, description, unit, price)
        if concept:
            return concept

    # Fallback para PDFs donde las columnas colapsan a un solo espacio.
    m = re.match(
        r"^\s*(?P<code>\S+)\s+"
        r"(?P<desc>.+?)\s+"
        r"(?P<unit>[A-Za-zÁÉÍÓÚÜÑáéíóúüñ0-9²³%./_-]{1,16})\s+"
        r"\$?\s*(?P<price>\d[\d,]*(?:\.\d+)?)\s*$",
        raw,
    )
    if not m:
        return None
    return _concept_from_cells(
        m.group("code"),
        m.group("desc"),
        m.group("unit"),
        m.group("price"),
    )


def parsear_tabulador_cdmx_pdf(pdf_bytes: bytes) -> list[dict]:
    """
    Convierte el PDF oficial en registros estructurados.

    Se usa texto en modo layout para conservar las cuatro columnas. Como
    protección adicional, se prueba una estrategia de líneas acumuladas para
    descripciones que se hayan dividido entre renglones.
    """
    reader = PdfReader(BytesIO(pdf_bytes))
    found: dict[str, dict] = {}

    for page in reader.pages:
        try:
            text_layout = page.extract_text(extraction_mode="layout") or ""
        except TypeError:
            text_layout = page.extract_text() or ""
        except Exception:
            continue

        lines = [line.rstrip() for line in text_layout.splitlines()]
        pending = ""

        for line in lines:
            concept = _parse_layout_line(line)
            if concept:
                found[concept["source_code"]] = concept
                pending = ""
                continue

            stripped = _clean_pdf_cell(line)
            if not stripped:
                pending = ""
                continue

            # Una fila puede partir la descripción en dos renglones. Solo se
            # acumula cuando el primer token parece código.
            first = stripped.split()[0] if stripped.split() else ""
            if _looks_like_source_code(first):
                pending = stripped
                continue

            if pending:
                merged = f"{pending} {stripped}"
                concept = _parse_layout_line(merged)
                if concept:
                    found[concept["source_code"]] = concept
                    pending = ""
                elif len(merged) < 1200:
                    pending = merged
                else:
                    pending = ""

    concepts = list(found.values())
    concepts.sort(key=lambda x: x["source_code"])
    return concepts


def descargar_tabulador_cdmx(catalog_info: dict, timeout: int = 90) -> bytes:
    response = requests.get(
        catalog_info["url"],
        headers=CDMX_HTTP_HEADERS,
        timeout=timeout,
    )
    response.raise_for_status()
    content_type = str(response.headers.get("content-type") or "").lower()
    payload = response.content
    if not payload.startswith(b"%PDF") and "pdf" not in content_type:
        raise RuntimeError(
            "La URL oficial encontrada no devolvió un archivo PDF válido."
        )
    return payload


def actualizar_catalogo_cdmx(db) -> dict:
    info = descubrir_ultimo_tabulador_cdmx()
    current = db.get_active_external_catalog("CDMX")

    if (
        current
        and int(current.get("year") or 0) == int(info.get("year") or 0)
        and int(current.get("month") or 0) == int(info.get("month") or 0)
        and int(current.get("concept_count") or 0) > 0
    ):
        return {
            "status": "already_current",
            "catalog": current,
            "concept_count": int(current["concept_count"]),
        }

    pdf_bytes = descargar_tabulador_cdmx(info)
    concepts = parsear_tabulador_cdmx_pdf(pdf_bytes)

    # Una edición completa normalmente contiene miles de filas. Este umbral
    # evita reemplazar un catálogo bueno por una extracción rota.
    if len(concepts) < 500:
        raise RuntimeError(
            "La extracción del PDF produjo muy pocos conceptos "
            f"({len(concepts)}). No se modificó el catálogo existente."
        )

    catalog = db.replace_external_catalog(
        source="CDMX",
        version_label=info["version_label"],
        year=int(info["year"]),
        month=int(info.get("month") or 0),
        source_url=info["url"],
        concepts=concepts,
    )
    return {
        "status": "updated",
        "catalog": catalog,
        "concept_count": len(concepts),
    }


def _tokens_utiles(texto: str) -> set[str]:
    stop = {
        "PARA", "CON", "DEL", "LAS", "LOS", "UNA", "UNO", "UNAS", "UNOS",
        "QUE", "POR", "COMO", "MAS", "INCLUYE", "INCLUYENDO", "SUMINISTRO",
        "INSTALACION", "COLOCACION", "TRABAJO", "TRABAJOS", "SERVICIO",
        "COMPLETO", "COMPLETA", "SEGUN", "MEDIANTE", "HASTA", "DESDE",
    }
    return {
        token
        for token in normalizar_texto(texto).upper().split()
        if len(token) >= 4 and token not in stop
    }


def score_similitud_externa(a: str, b: str) -> float:
    a_n = normalizar_texto(a).upper()
    b_n = normalizar_texto(b).upper()
    if not a_n or not b_n:
        return 0.0

    seq = SequenceMatcher(None, a_n, b_n).ratio()
    ta = _tokens_utiles(a_n)
    tb = _tokens_utiles(b_n)
    union = ta | tb
    intersection = ta & tb
    jaccard = len(intersection) / len(union) if union else 0.0
    coverage = len(intersection) / len(ta) if ta else 0.0

    # La cobertura de palabras de la actividad tiene bastante peso para evitar
    # coincidencias visualmente similares pero técnicamente distintas.
    return 0.45 * seq + 0.25 * jaccard + 0.30 * coverage


# =========================================================
# BASE DE DATOS
# =========================================================


class Database:
    """Base relacional con PostgreSQL opcional y SQLite como respaldo local."""

    def __init__(self, database_url: str | None = None):
        self.database_url = (database_url or "").strip() or None
        self.kind = "postgres" if self.database_url else "sqlite"
        self.sqlite_path = Path(__file__).with_name("presupuestador_empresa.db")
        self._init_schema()

    @property
    def persistent(self) -> bool:
        return self.kind == "postgres"

    def _connect(self):
        if self.kind == "postgres":
            if psycopg is None:
                raise RuntimeError("psycopg no está instalado.")
            return psycopg.connect(
                self.database_url,
                row_factory=dict_row,
                prepare_threshold=None,
                connect_timeout=15,
            )

        conn = sqlite3.connect(self.sqlite_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

    def _adapt(self, sql: str) -> str:
        return sql.replace("?", "%s") if self.kind == "postgres" else sql

    def _ensure_column(self, table: str, column: str, sql_type: str):
        """Agrega una columna a una base existente sin destruir información."""
        allowed_tables = {"projects", "budgets", "concepts", "price_history", "budget_items"}
        allowed_columns = {
            "parent_budget_id",
            "revision_instruction",
            "commercial_title",
            "budget_level",
            "material_share_pct",
            "labor_share_pct",
            "other_share_pct",
            "waste_reference_pct",
            "execution_order",
            "area_allocations_json",
            "area",
            "included",
            "original_unit_sale",
            "package_group",
            "price_evidence_score",
        }
        if table not in allowed_tables or column not in allowed_columns:
            raise ValueError("Migración de columna no permitida.")

        if self.kind == "postgres":
            exists = self.fetchone(
                """
                SELECT 1 AS ok
                FROM information_schema.columns
                WHERE table_schema = 'public'
                  AND table_name = ?
                  AND column_name = ?
                """,
                (table, column),
            )
        else:
            rows = self.fetchall(f"PRAGMA table_info({table})")
            exists = any(str(r.get("name")) == column for r in rows)

        if not exists:
            self.execute(f"ALTER TABLE {table} ADD COLUMN {column} {sql_type}")

    def execute(self, sql: str, params=()):
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute(self._adapt(sql), params)
            conn.commit()

    def executemany(self, sql: str, seq_params):
        with self._connect() as conn:
            cur = conn.cursor()
            cur.executemany(self._adapt(sql), seq_params)
            conn.commit()

    def fetchone(self, sql: str, params=()):
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute(self._adapt(sql), params)
            row = cur.fetchone()
            return dict(row) if row is not None else None

    def fetchall(self, sql: str, params=()):
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute(self._adapt(sql), params)
            rows = cur.fetchall()
            return [dict(r) for r in rows]

    def _init_schema(self):
        schema = [
            """
            CREATE TABLE IF NOT EXISTS projects (
                id TEXT PRIMARY KEY,
                code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                project_type TEXT NOT NULL,
                budget_level TEXT,
                location TEXT,
                dimension_mode TEXT,
                dimensions_text TEXT,
                description TEXT,
                guide_text TEXT,
                main_activity TEXT,
                created_at TEXT NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS budgets (
                id TEXT PRIMARY KEY,
                project_id TEXT NOT NULL,
                version INTEGER NOT NULL DEFAULT 1,
                status TEXT NOT NULL,
                indirect_pct REAL NOT NULL,
                profit_pct REAL NOT NULL,
                iva_pct REAL NOT NULL,
                waste_pct REAL NOT NULL,
                direct_cost REAL NOT NULL,
                indirect_cost REAL NOT NULL,
                profit REAL NOT NULL,
                sale_before_tax REAL NOT NULL,
                iva_amount REAL NOT NULL,
                total REAL NOT NULL,
                scope_summary TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS concepts (
                id TEXT PRIMARY KEY,
                code TEXT,
                category TEXT,
                subcategory TEXT,
                description TEXT NOT NULL,
                unit TEXT NOT NULL,
                normalized_description TEXT NOT NULL,
                created_budget_id TEXT,
                created_at TEXT NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS price_history (
                id TEXT PRIMARY KEY,
                concept_id TEXT NOT NULL,
                unit_cost REAL NOT NULL,
                source TEXT NOT NULL,
                source_detail TEXT,
                status TEXT NOT NULL,
                confidence TEXT,
                project_id TEXT,
                budget_id TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(concept_id) REFERENCES concepts(id) ON DELETE CASCADE,
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE SET NULL,
                FOREIGN KEY(budget_id) REFERENCES budgets(id) ON DELETE CASCADE
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS budget_items (
                id TEXT PRIMARY KEY,
                budget_id TEXT NOT NULL,
                concept_id TEXT,
                category TEXT,
                subcategory TEXT,
                code TEXT,
                commercial_title TEXT,
                description TEXT NOT NULL,
                unit TEXT NOT NULL,
                quantity REAL NOT NULL,
                unit_cost REAL NOT NULL,
                direct_amount REAL NOT NULL,
                unit_indirect REAL NOT NULL,
                unit_profit REAL NOT NULL,
                unit_sale REAL NOT NULL,
                sale_amount REAL NOT NULL,
                sale_margin_pct REAL NOT NULL,
                benefit_amount REAL NOT NULL,
                price_source TEXT,
                price_source_detail TEXT,
                price_confidence TEXT,
                material_share_pct REAL,
                labor_share_pct REAL,
                other_share_pct REAL,
                waste_reference_pct REAL,
                execution_order INTEGER,
                area_allocations_json TEXT,
                area TEXT,
                included INTEGER,
                original_unit_sale REAL,
                package_group TEXT,
                price_evidence_score REAL,
                quantity_criterion TEXT,
                inclusion_basis TEXT,
                considerations TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(budget_id) REFERENCES budgets(id) ON DELETE CASCADE,
                FOREIGN KEY(concept_id) REFERENCES concepts(id) ON DELETE SET NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS external_catalogs (
                id TEXT PRIMARY KEY,
                source TEXT NOT NULL,
                version_label TEXT NOT NULL,
                year INTEGER NOT NULL,
                month INTEGER NOT NULL DEFAULT 0,
                source_url TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                concept_count INTEGER NOT NULL DEFAULT 0,
                active INTEGER NOT NULL DEFAULT 1
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS external_concepts (
                id TEXT PRIMARY KEY,
                catalog_id TEXT NOT NULL,
                source TEXT NOT NULL,
                source_code TEXT NOT NULL,
                description TEXT NOT NULL,
                normalized_description TEXT NOT NULL,
                unit TEXT NOT NULL,
                normalized_unit TEXT NOT NULL,
                unit_price REAL NOT NULL,
                chapter TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(catalog_id) REFERENCES external_catalogs(id) ON DELETE CASCADE
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_concepts_norm ON concepts(normalized_description)",
            "CREATE INDEX IF NOT EXISTS idx_ext_catalog_source ON external_catalogs(source, active)",
            "CREATE INDEX IF NOT EXISTS idx_ext_concept_catalog ON external_concepts(catalog_id)",
            "CREATE INDEX IF NOT EXISTS idx_ext_concept_unit ON external_concepts(source, normalized_unit)",
            "CREATE INDEX IF NOT EXISTS idx_ext_concept_norm ON external_concepts(normalized_description)",
            "CREATE INDEX IF NOT EXISTS idx_price_concept ON price_history(concept_id)",
            "CREATE INDEX IF NOT EXISTS idx_items_budget ON budget_items(budget_id)",
        ]
        for statement in schema:
            self.execute(statement)

        # Migraciones no destructivas para instalaciones creadas con versiones
        # anteriores de la app.
        self._ensure_column("budgets", "parent_budget_id", "TEXT")
        self._ensure_column("budgets", "revision_instruction", "TEXT")
        self._ensure_column("budget_items", "commercial_title", "TEXT")
        self._ensure_column("projects", "budget_level", "TEXT")
        self._ensure_column("budget_items", "material_share_pct", "REAL")
        self._ensure_column("budget_items", "labor_share_pct", "REAL")
        self._ensure_column("budget_items", "other_share_pct", "REAL")
        self._ensure_column("budget_items", "waste_reference_pct", "REAL")
        self._ensure_column("budget_items", "execution_order", "INTEGER")
        self._ensure_column("budget_items", "area_allocations_json", "TEXT")
        self._ensure_column("budget_items", "area", "TEXT")
        self._ensure_column("budget_items", "included", "INTEGER")
        self._ensure_column("budget_items", "original_unit_sale", "REAL")
        self._ensure_column("budget_items", "package_group", "TEXT")
        self._ensure_column("budget_items", "price_evidence_score", "REAL")

    def stats(self) -> dict:
        return {
            "projects": self.fetchone("SELECT COUNT(*) AS n FROM projects")["n"],
            "budgets": self.fetchone("SELECT COUNT(*) AS n FROM budgets")["n"],
            "concepts": self.fetchone("SELECT COUNT(*) AS n FROM concepts")["n"],
        }

    def get_latest_project_record(self) -> dict | None:
        """Devuelve el último proyecto guardado y su último total conocido."""
        return self.fetchone(
            """
            SELECT p.*,
                   (SELECT COUNT(*) FROM budgets b WHERE b.project_id = p.id) AS budget_count,
                   (SELECT b.total FROM budgets b
                    WHERE b.project_id = p.id
                    ORDER BY b.created_at DESC LIMIT 1) AS latest_total
            FROM projects p
            ORDER BY p.created_at DESC
            LIMIT 1
            """
        )

    def get_active_external_catalog(self, source: str) -> dict | None:
        return self.fetchone(
            """
            SELECT *
            FROM external_catalogs
            WHERE UPPER(source) = UPPER(?) AND active = 1
            ORDER BY year DESC, month DESC, imported_at DESC
            LIMIT 1
            """,
            (source,),
        )

    def list_external_catalogs(self, source: str | None = None) -> list[dict]:
        if source:
            return self.fetchall(
                """
                SELECT *
                FROM external_catalogs
                WHERE UPPER(source) = UPPER(?)
                ORDER BY year DESC, month DESC, imported_at DESC
                """,
                (source,),
            )
        return self.fetchall(
            """
            SELECT *
            FROM external_catalogs
            ORDER BY source, year DESC, month DESC, imported_at DESC
            """
        )

    def replace_external_catalog(
        self,
        source: str,
        version_label: str,
        year: int,
        month: int,
        source_url: str,
        concepts: list[dict],
    ) -> dict:
        """
        Importa una edición completa en una sola transacción. Las ediciones
        anteriores se conservan como metadatos, pero solo la nueva queda activa.
        """
        source = source.strip().upper()
        created = ahora_iso()
        catalog_id = str(uuid.uuid4())

        with self._connect() as conn:
            cur = conn.cursor()

            cur.execute(
                self._adapt(
                    "UPDATE external_catalogs SET active = 0 WHERE UPPER(source) = UPPER(?)"
                ),
                (source,),
            )

            # Si la misma edición fue importada parcialmente antes, se elimina
            # para que el nuevo lote sea consistente.
            cur.execute(
                self._adapt(
                    """
                    SELECT id FROM external_catalogs
                    WHERE UPPER(source) = UPPER(?) AND year = ? AND month = ?
                    """
                ),
                (source, year, month),
            )
            duplicate_rows = cur.fetchall()
            for dup in duplicate_rows:
                dup_id = dup["id"] if isinstance(dup, dict) else dup[0]
                cur.execute(
                    self._adapt("DELETE FROM external_catalogs WHERE id = ?"),
                    (dup_id,),
                )

            cur.execute(
                self._adapt(
                    """
                    INSERT INTO external_catalogs (
                        id, source, version_label, year, month, source_url,
                        imported_at, concept_count, active
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """
                ),
                (
                    catalog_id,
                    source,
                    version_label,
                    year,
                    month,
                    source_url,
                    created,
                    len(concepts),
                    1,
                ),
            )

            insert_sql = self._adapt(
                """
                INSERT INTO external_concepts (
                    id, catalog_id, source, source_code, description,
                    normalized_description, unit, normalized_unit, unit_price,
                    chapter, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
            )
            rows = [
                (
                    str(uuid.uuid4()),
                    catalog_id,
                    source,
                    c["source_code"],
                    c["description"],
                    c["normalized_description"],
                    c["unit"],
                    c["normalized_unit"],
                    float(c["unit_price"]),
                    c.get("chapter") or "",
                    created,
                )
                for c in concepts
            ]
            cur.executemany(insert_sql, rows)
            conn.commit()

        return self.get_active_external_catalog(source)

    def external_candidates(
        self,
        source: str,
        unit: str,
        description: str,
        limit: int = 350,
    ) -> list[dict]:
        normalized_unit = normalizar_unidad(unit)
        if not normalized_unit:
            return []

        keywords = sorted(
            _tokens_utiles(description),
            key=lambda x: (-len(x), x),
        )[:5]

        base_sql = """
            SELECT ec.*, c.version_label, c.year, c.month, c.source_url
            FROM external_concepts ec
            JOIN external_catalogs c ON c.id = ec.catalog_id
            WHERE c.active = 1
              AND UPPER(ec.source) = UPPER(?)
              AND ec.normalized_unit = ?
        """
        params: list = [source, normalized_unit]

        if keywords:
            clauses = []
            for word in keywords:
                clauses.append("UPPER(ec.normalized_description) LIKE ?")
                params.append(f"%{word}%")
            sql = base_sql + " AND (" + " OR ".join(clauses) + ") LIMIT ?"
            params.append(limit)
            rows = self.fetchall(sql, tuple(params))
            if rows:
                return rows

        return self.fetchall(
            base_sql + " LIMIT ?",
            (source, normalized_unit, limit),
        )

    def search_external_concepts(
        self,
        source: str = "CDMX",
        search: str = "",
        limit: int = 250,
    ) -> list[dict]:
        q = normalizar_texto(search).upper()
        if q:
            return self.fetchall(
                """
                SELECT ec.*, c.version_label, c.source_url
                FROM external_concepts ec
                JOIN external_catalogs c ON c.id = ec.catalog_id
                WHERE c.active = 1
                  AND UPPER(ec.source) = UPPER(?)
                  AND (
                    UPPER(ec.source_code) LIKE ?
                    OR UPPER(ec.normalized_description) LIKE ?
                  )
                ORDER BY ec.source_code
                LIMIT ?
                """,
                (source, f"%{q}%", f"%{q}%", limit),
            )
        return self.fetchall(
            """
            SELECT ec.*, c.version_label, c.source_url
            FROM external_concepts ec
            JOIN external_catalogs c ON c.id = ec.catalog_id
            WHERE c.active = 1 AND UPPER(ec.source) = UPPER(?)
            ORDER BY ec.source_code
            LIMIT ?
            """,
            (source, limit),
        )

    def delete_external_source(self, source: str):
        # Cascada: borrar catálogo elimina sus conceptos.
        self.execute(
            "DELETE FROM external_catalogs WHERE UPPER(source) = UPPER(?)",
            (source,),
        )

    def clear_all_data(self):
        """
        Elimina todos los datos empresariales de las tablas de la aplicación.
        Conserva el esquema para que la app siga funcionando inmediatamente.
        """
        # El orden evita conflictos de claves foráneas tanto en PostgreSQL como SQLite.
        for table in [
            "external_concepts",
            "external_catalogs",
            "budget_items",
            "price_history",
            "budgets",
            "concepts",
            "projects",
        ]:
            self.execute(f"DELETE FROM {table}")

    def next_project_code(self, client_name: str, location: str) -> str:
        """
        Código corporativo: abreviatura del cliente + ubicación + consecutivo.
        Ejemplo: Desarrollos de la Vega / Farallón -> DDV-FAR-0001.
        """
        prefix = f"{abreviar_cliente(client_name or 'Cliente')}-{abreviar(location or 'Ubicacion')}"
        rows = self.fetchall("SELECT code FROM projects WHERE code LIKE ?", (f"{prefix}-%",))
        max_num = 0
        for row in rows:
            m = re.search(r"-(\d+)$", row["code"] or "")
            if m:
                max_num = max(max_num, int(m.group(1)))
        return f"{prefix}-{max_num + 1:04d}"

    def price_candidates(self, unit: str, limit: int = 600) -> list[dict]:
        rows = self.fetchall(
            """
            SELECT
                c.id AS concept_id,
                c.code,
                c.category,
                c.subcategory,
                c.description,
                c.unit,
                c.normalized_description,
                ph.unit_cost,
                ph.source,
                ph.source_detail,
                ph.status,
                ph.confidence,
                ph.created_at
            FROM concepts c
            JOIN price_history ph ON ph.concept_id = c.id
            WHERE UPPER(c.unit) = UPPER(?)
            ORDER BY ph.created_at DESC
            """,
            (unit,),
        )

        # Conserva solamente el precio más reciente de cada concepto.
        unique = []
        seen = set()
        for row in rows:
            if row["concept_id"] in seen:
                continue
            seen.add(row["concept_id"])
            unique.append(row)
            if len(unique) >= limit:
                break
        return unique

    def save_generation(
        self,
        project_code: str,
        project_data: dict,
        result,
        items: list[dict],
        params: dict,
        financials: dict,
    ) -> tuple[str, str]:
        project_id = str(uuid.uuid4())
        budget_id = str(uuid.uuid4())
        created = ahora_iso()

        self.execute(
            """
            INSERT INTO projects (
                id, code, name, project_type, budget_level, location, dimension_mode,
                dimensions_text, description, guide_text, main_activity, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                project_id,
                project_code,
                project_data["name"],
                project_data["project_type"],
                project_data.get("budget_level", "Medio-alto"),
                project_data["location"],
                project_data["dimension_mode"],
                project_data["dimensions_text"],
                project_data["description"],
                project_data["guide_text"],
                result.actividad_principal,
                created,
            ),
        )

        self.execute(
            """
            INSERT INTO budgets (
                id, project_id, version, status, indirect_pct, profit_pct,
                iva_pct, waste_pct, direct_cost, indirect_cost, profit,
                sale_before_tax, iva_amount, total, scope_summary, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                budget_id,
                project_id,
                1,
                "GENERADO",
                params["indirect_pct"],
                params["profit_pct"],
                params["iva_pct"],
                params["waste_pct"],
                financials["direct_cost"],
                financials["indirect_cost"],
                financials["profit"],
                financials["sale_before_tax"],
                financials["iva_amount"],
                financials["total"],
                result.alcance_resumido,
                created,
            ),
        )

        for item in items:
            concept_id = item.get("concept_id")
            concept_was_existing = bool(concept_id)

            if not concept_id:
                concept_id = str(uuid.uuid4())
                item["concept_id"] = concept_id
                self.execute(
                    """
                    INSERT INTO concepts (
                        id, code, category, subcategory, description, unit,
                        normalized_description, created_budget_id, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        concept_id,
                        item["code"],
                        item["category"],
                        item["subcategory"],
                        item["description"],
                        item["unit"],
                        normalizar_texto(item["description"]),
                        budget_id,
                        created,
                    ),
                )

                # Solo se crea historial nuevo cuando el sistema generó o encontró
                # una referencia externa nueva. Un precio interno reutilizado ya
                # cuenta con historial propio.
                if item["price_source"] not in {"BASE_INTERNA", "HISTORICO_IA"}:
                    self.execute(
                        """
                        INSERT INTO price_history (
                            id, concept_id, unit_cost, source, source_detail,
                            status, confidence, project_id, budget_id, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            str(uuid.uuid4()),
                            concept_id,
                            item["unit_cost"],
                            item["price_source"],
                            item["price_source_detail"],
                            item["price_status"],
                            item["price_confidence"],
                            project_id,
                            budget_id,
                            created,
                        ),
                    )

            if concept_was_existing and item.get("record_new_price"):
                self.execute(
                    """
                    INSERT INTO price_history (
                        id, concept_id, unit_cost, source, source_detail,
                        status, confidence, project_id, budget_id, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid.uuid4()),
                        concept_id,
                        item["unit_cost"],
                        item["price_source"],
                        item["price_source_detail"],
                        item["price_status"],
                        item["price_confidence"],
                        project_id,
                        budget_id,
                        created,
                    ),
                )

            self.execute(
                """
                INSERT INTO budget_items (
                    id, budget_id, concept_id, category, subcategory, code,
                    commercial_title, description, unit, quantity, unit_cost, direct_amount,
                    unit_indirect, unit_profit, unit_sale, sale_amount,
                    sale_margin_pct, benefit_amount, price_source,
                    price_source_detail, price_confidence,
                    material_share_pct, labor_share_pct, other_share_pct, waste_reference_pct,
                    execution_order, area_allocations_json, area, included,
                    original_unit_sale, package_group, price_evidence_score,
                    quantity_criterion, inclusion_basis, considerations, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(uuid.uuid4()),
                    budget_id,
                    concept_id,
                    item["category"],
                    item["subcategory"],
                    item["code"],
                    titulo_comercial_item(item),
                    item["description"],
                    item["unit"],
                    item["quantity"],
                    item["unit_cost"],
                    item["direct_amount"],
                    item["unit_indirect"],
                    item["unit_profit"],
                    item["unit_sale"],
                    item["sale_amount"],
                    item["sale_margin_pct"],
                    item["benefit_amount"],
                    item["price_source"],
                    item["price_source_detail"],
                    item["price_confidence"],
                    item.get("material_share_pct", 0.0),
                    item.get("labor_share_pct", 0.0),
                    item.get("other_share_pct", 100.0),
                    item.get("waste_reference_pct", 0.0),
                    int(item.get("execution_order") or 500),
                    json.dumps(
                        obtener_asignaciones_area_item(item),
                        ensure_ascii=False,
                        separators=(",", ":"),
                    ),
                    area_item(item),
                    1 if item_included(item) else 0,
                    original_unit_sale_item(item),
                    str(item.get("package_group") or ""),
                    float(item.get("price_evidence_score") or 0.0),
                    item["quantity_criterion"],
                    item["inclusion_basis"],
                    item["considerations"],
                    created,
                ),
            )

        return project_id, budget_id

    def save_revision(
        self,
        project_id: str,
        parent_budget_id: str,
        result,
        items: list[dict],
        params: dict,
        financials: dict,
        revision_instruction: str,
    ) -> tuple[str, int]:
        """
        Guarda una revisión como NUEVO presupuesto del mismo proyecto.
        El presupuesto anterior se conserva para mantener trazabilidad.
        """
        row = self.fetchone(
            "SELECT COALESCE(MAX(version), 0) AS max_version FROM budgets WHERE project_id = ?",
            (project_id,),
        )
        version = int(row["max_version"] or 0) + 1
        budget_id = str(uuid.uuid4())
        created = ahora_iso()

        self.execute(
            """
            INSERT INTO budgets (
                id, project_id, version, status, indirect_pct, profit_pct,
                iva_pct, waste_pct, direct_cost, indirect_cost, profit,
                sale_before_tax, iva_amount, total, scope_summary, created_at,
                parent_budget_id, revision_instruction
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                budget_id,
                project_id,
                version,
                "REVISION",
                params["indirect_pct"],
                params["profit_pct"],
                params["iva_pct"],
                params["waste_pct"],
                financials["direct_cost"],
                financials["indirect_cost"],
                financials["profit"],
                financials["sale_before_tax"],
                financials["iva_amount"],
                financials["total"],
                result.alcance_resumido,
                created,
                parent_budget_id,
                revision_instruction.strip(),
            ),
        )

        for item in items:
            concept_id = item.get("concept_id")
            concept_was_existing = bool(concept_id)

            if not concept_id:
                concept_id = str(uuid.uuid4())
                item["concept_id"] = concept_id
                self.execute(
                    """
                    INSERT INTO concepts (
                        id, code, category, subcategory, description, unit,
                        normalized_description, created_budget_id, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        concept_id,
                        item["code"],
                        item["category"],
                        item["subcategory"],
                        item["description"],
                        item["unit"],
                        normalizar_texto(item["description"]),
                        budget_id,
                        created,
                    ),
                )

                if item["price_source"] not in {"BASE_INTERNA", "HISTORICO_IA"}:
                    self.execute(
                        """
                        INSERT INTO price_history (
                            id, concept_id, unit_cost, source, source_detail,
                            status, confidence, project_id, budget_id, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            str(uuid.uuid4()),
                            concept_id,
                            item["unit_cost"],
                            item["price_source"],
                            item["price_source_detail"],
                            item["price_status"],
                            item["price_confidence"],
                            project_id,
                            budget_id,
                            created,
                        ),
                    )

            if concept_was_existing and item.get("record_new_price"):
                self.execute(
                    """
                    INSERT INTO price_history (
                        id, concept_id, unit_cost, source, source_detail,
                        status, confidence, project_id, budget_id, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid.uuid4()),
                        concept_id,
                        item["unit_cost"],
                        item["price_source"],
                        item["price_source_detail"],
                        item["price_status"],
                        item["price_confidence"],
                        project_id,
                        budget_id,
                        created,
                    ),
                )

            self.execute(
                """
                INSERT INTO budget_items (
                    id, budget_id, concept_id, category, subcategory, code,
                    commercial_title, description, unit, quantity, unit_cost, direct_amount,
                    unit_indirect, unit_profit, unit_sale, sale_amount,
                    sale_margin_pct, benefit_amount, price_source,
                    price_source_detail, price_confidence,
                    material_share_pct, labor_share_pct, other_share_pct, waste_reference_pct,
                    execution_order, area_allocations_json, area, included,
                    original_unit_sale, package_group, price_evidence_score,
                    quantity_criterion, inclusion_basis, considerations, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(uuid.uuid4()),
                    budget_id,
                    concept_id,
                    item["category"],
                    item["subcategory"],
                    item["code"],
                    titulo_comercial_item(item),
                    item["description"],
                    item["unit"],
                    item["quantity"],
                    item["unit_cost"],
                    item["direct_amount"],
                    item["unit_indirect"],
                    item["unit_profit"],
                    item["unit_sale"],
                    item["sale_amount"],
                    item["sale_margin_pct"],
                    item["benefit_amount"],
                    item["price_source"],
                    item["price_source_detail"],
                    item["price_confidence"],
                    item.get("material_share_pct", 0.0),
                    item.get("labor_share_pct", 0.0),
                    item.get("other_share_pct", 100.0),
                    item.get("waste_reference_pct", 0.0),
                    int(item.get("execution_order") or 500),
                    json.dumps(
                        obtener_asignaciones_area_item(item),
                        ensure_ascii=False,
                        separators=(",", ":"),
                    ),
                    area_item(item),
                    1 if item_included(item) else 0,
                    original_unit_sale_item(item),
                    str(item.get("package_group") or ""),
                    float(item.get("price_evidence_score") or 0.0),
                    item["quantity_criterion"],
                    item["inclusion_basis"],
                    item["considerations"],
                    created,
                ),
            )

        return budget_id, version

    def delete_generation(self, project_id: str, budget_id: str):
        # La eliminación del presupuesto borra partidas e historial vinculado
        # mediante ON DELETE CASCADE.
        self.execute("DELETE FROM budgets WHERE id = ?", (budget_id,))

        # Elimina conceptos creados exclusivamente por el presupuesto descartado.
        self.execute(
            """
            DELETE FROM concepts
            WHERE created_budget_id = ?
              AND NOT EXISTS (
                  SELECT 1 FROM budget_items bi WHERE bi.concept_id = concepts.id
              )
              AND NOT EXISTS (
                  SELECT 1 FROM price_history ph WHERE ph.concept_id = concepts.id
              )
            """,
            (budget_id,),
        )

        remaining = self.fetchone(
            "SELECT COUNT(*) AS n FROM budgets WHERE project_id = ?",
            (project_id,),
        )["n"]
        if remaining == 0:
            self.execute("DELETE FROM projects WHERE id = ?", (project_id,))


    # -----------------------------------------------------
    # Administración de la base interna
    # -----------------------------------------------------

    def list_concepts(self, search: str = "", limit: int = 500) -> list[dict]:
        search_n = normalizar_texto(search)
        if search_n:
            rows = self.fetchall(
                """
                SELECT c.*,
                       (SELECT ph.unit_cost FROM price_history ph
                        WHERE ph.concept_id = c.id
                        ORDER BY ph.created_at DESC LIMIT 1) AS latest_cost,
                       (SELECT ph.source FROM price_history ph
                        WHERE ph.concept_id = c.id
                        ORDER BY ph.created_at DESC LIMIT 1) AS latest_source,
                       (SELECT ph.status FROM price_history ph
                        WHERE ph.concept_id = c.id
                        ORDER BY ph.created_at DESC LIMIT 1) AS latest_status,
                       (SELECT COUNT(*) FROM budget_items bi
                        WHERE bi.concept_id = c.id) AS usage_count
                FROM concepts c
                WHERE c.normalized_description LIKE ?
                   OR LOWER(COALESCE(c.code, '')) LIKE ?
                   OR LOWER(COALESCE(c.category, '')) LIKE ?
                   OR LOWER(COALESCE(c.subcategory, '')) LIKE ?
                ORDER BY c.description
                LIMIT ?
                """,
                (f"%{search_n}%", f"%{search.lower()}%", f"%{search.lower()}%", f"%{search.lower()}%", limit),
            )
        else:
            rows = self.fetchall(
                """
                SELECT c.*,
                       (SELECT ph.unit_cost FROM price_history ph
                        WHERE ph.concept_id = c.id
                        ORDER BY ph.created_at DESC LIMIT 1) AS latest_cost,
                       (SELECT ph.source FROM price_history ph
                        WHERE ph.concept_id = c.id
                        ORDER BY ph.created_at DESC LIMIT 1) AS latest_source,
                       (SELECT ph.status FROM price_history ph
                        WHERE ph.concept_id = c.id
                        ORDER BY ph.created_at DESC LIMIT 1) AS latest_status,
                       (SELECT COUNT(*) FROM budget_items bi
                        WHERE bi.concept_id = c.id) AS usage_count
                FROM concepts c
                ORDER BY c.created_at DESC
                LIMIT ?
                """,
                (limit,),
            )
        return rows

    def get_concept(self, concept_id: str) -> dict | None:
        return self.fetchone("SELECT * FROM concepts WHERE id = ?", (concept_id,))

    def create_concept(self, code: str, category: str, subcategory: str, description: str, unit: str) -> str:
        concept_id = str(uuid.uuid4())
        self.execute(
            """
            INSERT INTO concepts (
                id, code, category, subcategory, description, unit,
                normalized_description, created_budget_id, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                concept_id,
                limpiar_codigo(code, "MAN-001"),
                category.strip(),
                subcategory.strip(),
                description.strip(),
                unit.strip().upper(),
                normalizar_texto(description),
                None,
                ahora_iso(),
            ),
        )
        return concept_id

    def update_concept(self, concept_id: str, code: str, category: str, subcategory: str, description: str, unit: str):
        self.execute(
            """
            UPDATE concepts
            SET code = ?, category = ?, subcategory = ?, description = ?,
                unit = ?, normalized_description = ?
            WHERE id = ?
            """,
            (
                limpiar_codigo(code, "CON-001"),
                category.strip(),
                subcategory.strip(),
                description.strip(),
                unit.strip().upper(),
                normalizar_texto(description),
                concept_id,
            ),
        )

    def concept_usage(self, concept_id: str) -> dict:
        return {
            "budget_items": self.fetchone(
                "SELECT COUNT(*) AS n FROM budget_items WHERE concept_id = ?", (concept_id,)
            )["n"],
            "prices": self.fetchone(
                "SELECT COUNT(*) AS n FROM price_history WHERE concept_id = ?", (concept_id,)
            )["n"],
        }

    def delete_concept(self, concept_id: str):
        self.execute("DELETE FROM concepts WHERE id = ?", (concept_id,))

    def list_prices(self, concept_id: str) -> list[dict]:
        return self.fetchall(
            """
            SELECT * FROM price_history
            WHERE concept_id = ?
            ORDER BY created_at DESC
            """,
            (concept_id,),
        )

    def add_price(
        self,
        concept_id: str,
        unit_cost: float,
        source: str,
        source_detail: str,
        status: str,
        confidence: str,
    ) -> str:
        price_id = str(uuid.uuid4())
        self.execute(
            """
            INSERT INTO price_history (
                id, concept_id, unit_cost, source, source_detail,
                status, confidence, project_id, budget_id, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                price_id,
                concept_id,
                float(unit_cost),
                source.strip().upper(),
                source_detail.strip(),
                status.strip().upper(),
                confidence.strip(),
                None,
                None,
                ahora_iso(),
            ),
        )
        return price_id

    def delete_price(self, price_id: str):
        self.execute("DELETE FROM price_history WHERE id = ?", (price_id,))

    def list_projects(self, limit: int = 500) -> list[dict]:
        return self.fetchall(
            """
            SELECT p.*,
                   (SELECT COUNT(*) FROM budgets b WHERE b.project_id = p.id) AS budget_count,
                   (SELECT MAX(b.total) FROM budgets b WHERE b.project_id = p.id) AS latest_total
            FROM projects p
            ORDER BY p.created_at DESC
            LIMIT ?
            """,
            (limit,),
        )

    def get_project(self, project_id: str) -> dict | None:
        return self.fetchone("SELECT * FROM projects WHERE id = ?", (project_id,))

    def update_project(
        self,
        project_id: str,
        name: str,
        project_type: str,
        location: str,
        main_activity: str,
        dimensions_text: str,
        description: str,
        guide_text: str,
    ):
        self.execute(
            """
            UPDATE projects
            SET name = ?, project_type = ?, location = ?, main_activity = ?,
                dimensions_text = ?, description = ?, guide_text = ?
            WHERE id = ?
            """,
            (
                name.strip(), project_type.strip(), location.strip(), main_activity.strip(),
                dimensions_text.strip(), description.strip(), guide_text.strip(), project_id,
            ),
        )

    def delete_project(self, project_id: str):
        budget_rows = self.fetchall("SELECT id FROM budgets WHERE project_id = ?", (project_id,))
        budget_ids = [r["id"] for r in budget_rows]
        self.execute("DELETE FROM projects WHERE id = ?", (project_id,))

        # Limpieza de conceptos que nacieron en presupuestos del proyecto eliminado
        # y que ya no cuentan con uso ni historial.
        for budget_id in budget_ids:
            self.execute(
                """
                DELETE FROM concepts
                WHERE created_budget_id = ?
                  AND NOT EXISTS (
                      SELECT 1 FROM budget_items bi WHERE bi.concept_id = concepts.id
                  )
                  AND NOT EXISTS (
                      SELECT 1 FROM price_history ph WHERE ph.concept_id = concepts.id
                  )
                """,
                (budget_id,),
            )

    def list_budgets(self, project_id: str | None = None, limit: int = 500) -> list[dict]:
        if project_id:
            return self.fetchall(
                """
                SELECT b.*, p.code AS project_code, p.name AS project_name,
                       p.location AS project_location
                FROM budgets b
                JOIN projects p ON p.id = b.project_id
                WHERE b.project_id = ?
                ORDER BY b.created_at DESC
                LIMIT ?
                """,
                (project_id, limit),
            )
        return self.fetchall(
            """
            SELECT b.*, p.code AS project_code, p.name AS project_name,
                   p.location AS project_location
            FROM budgets b
            JOIN projects p ON p.id = b.project_id
            ORDER BY b.created_at DESC
            LIMIT ?
            """,
            (limit,),
        )

    def get_budget(self, budget_id: str) -> dict | None:
        return self.fetchone(
            """
            SELECT b.*, p.code AS project_code, p.name AS project_name
            FROM budgets b
            JOIN projects p ON p.id = b.project_id
            WHERE b.id = ?
            """,
            (budget_id,),
        )

    def list_budget_items(self, budget_id: str) -> list[dict]:
        return self.fetchall(
            """
            SELECT * FROM budget_items
            WHERE budget_id = ?
            ORDER BY category, subcategory, code, created_at
            """,
            (budget_id,),
        )

    def delete_budget(self, budget_id: str):
        budget = self.fetchone("SELECT project_id FROM budgets WHERE id = ?", (budget_id,))
        if not budget:
            return
        self.delete_generation(budget["project_id"], budget_id)

    def export_table(self, table_name: str) -> list[dict]:
        allowed = {
            "projects", "budgets", "concepts", "price_history", "budget_items",
            "external_catalogs", "external_concepts",
        }
        if table_name not in allowed:
            raise ValueError("Tabla no permitida.")
        return self.fetchall(f"SELECT * FROM {table_name}")


DATABASE_CACHE_VERSION = "2026-09-03-v16-multipass-precios-evidencia"


@st.cache_resource(show_spinner=False)
def get_database(database_url: str | None, cache_version: str):
    """
    Usa PostgreSQL si DATABASE_URL existe; SQLite solo para desarrollo sin URL.

    cache_version forma parte deliberadamente de la clave de caché. Esto evita
    reutilizar una instancia de Database creada con una versión anterior de la
    clase después de actualizar app.py en Streamlit Community Cloud.
    """
    _ = cache_version
    if database_url:
        # En producción no se hace fallback silencioso: si Supabase falla,
        # se detiene para evitar guardar datos en un SQLite efímero por accidente.
        return Database(database_url), None
    return Database(None), None


# =========================================================
# MODELOS DE RESPUESTA ESTRUCTURADA
# =========================================================


class AreasProyectoIA(BaseModel):
    areas: list[str] = Field(
        description=(
            "Lista cerrada de espacios físicos o funcionales del proyecto. "
            "Debe contener General y después las áreas específicas."
        )
    )


class ReferenciaPrecioWebIA(BaseModel):
    precio_unitario: float = Field(ge=0)
    unidad: str
    tipo_precio: str = Field(
        description=(
            "SUBCONTRATISTA, SUMINISTRO_INSTALACION, PRECIO_PUBLICO_INSTALADO "
            "o MATERIAL_SOLO"
        )
    )
    descripcion_referencia: str
    fuente_nombre: str
    fuente_url: str
    fecha_referencia: str = ""
    compatible: bool = True


class InvestigacionPrecioWebIA(BaseModel):
    referencias: list[ReferenciaPrecioWebIA]
    observaciones: str = ""


class ActividadIA(BaseModel):
    area: str = Field(
        description=(
            "Área física explícita del proyecto donde se ejecuta la actividad. "
            "Usa General solo para trabajos realmente comunes a toda la obra."
        )
    )
    partida: str = Field(
        description="Sección comercial amplia, por ejemplo ALBAÑILERÍA Y ESTRUCTURA"
    )
    subpartida: str = Field(
        description="Subpartida breve, legible y sin numeración, normalmente de 1 a 5 palabras"
    )
    codigo_sugerido: str = Field(description="Código interno breve y único")
    orden_ejecucion: int = Field(
        ge=1, le=999,
        description="Orden relativo según dependencias constructivas",
    )
    titulo_comercial: str = Field(description="Título corto que identifica el trabajo cotizado")
    descripcion_tecnica: str = Field(
        description="Descripción clara de qué se hace, dónde, especificación principal y qué incluye"
    )
    unidad: str = Field(description="Unidad: LOTE, PZA, M2, M3, ML, PTO, JGO, etc.")
    cantidad: float = Field(ge=0, description="Cantidad justificable")
    costo_unitario_estimado: float = Field(
        ge=0,
        description=(
            "Costo unitario integrado estimado que cobraría el subcontratista a la empresa, "
            "antes de indirectos, utilidad, 30% de marca e IVA."
        ),
    )
    criterio_cantidad: str = Field(description="Criterio verificable y breve de cantidad")
    fundamento_inclusion: str = Field(description="Razón breve para incluir la actividad")
    nivel_confianza_cantidad: str = Field(description="Alta, Media o Baja")
    nivel_confianza_precio: str = Field(description="Alta, Media o Baja")
    requiere_cotizacion: bool = Field(
        description="True cuando conviene confirmar el costo con proveedor/subcontratista"
    )
    consideraciones: str = Field(
        description="Supuestos, exclusiones, información faltante o condiciones por confirmar"
    )


class PresupuestoIA(BaseModel):
    nombre_proyecto: str
    actividad_principal: str
    alcance_resumido: str
    consideraciones_generales: list[str]
    datos_faltantes: list[str]
    actividades: list[ActividadIA]


class ClasificacionActividadIA(BaseModel):
    codigo: str = Field(description="Código exacto de la actividad recibida")
    partida: str = Field(description="Partida comercial corregida")
    subpartida: str = Field(description="Subpartida breve corregida, sin numeración")
    orden_ejecucion: int = Field(
        ge=1,
        le=999,
        description="Orden relativo corregido según la secuencia constructiva",
    )


class AuditoriaEstructuraIA(BaseModel):
    actividades: list[ClasificacionActividadIA]


class OperacionRevisionIA(BaseModel):
    accion: str = Field(description="AGREGAR, MODIFICAR o ELIMINAR")
    codigo_objetivo: str = Field(
        default="",
        description="Código exacto de la actividad actual para MODIFICAR o ELIMINAR"
    )
    actividad: ActividadIA | None = Field(
        default=None,
        description="Actividad completa resultante para AGREGAR o MODIFICAR; null para ELIMINAR"
    )
    recalcular_precio: bool = Field(
        default=True,
        description="True si el cambio altera materialmente el alcance o la base del costo unitario"
    )
    motivo: str = Field(description="Resumen técnico breve del cambio solicitado")


class RevisionPresupuestoIA(BaseModel):
    resumen_revision: str
    actividad_principal_actualizada: str
    alcance_resumido_actualizado: str
    consideraciones_generales_actualizadas: list[str]
    datos_faltantes_actualizados: list[str]
    operaciones: list[OperacionRevisionIA]


# =========================================================
# GEMINI
# =========================================================


def get_api_key() -> str | None:
    return get_secret("GEMINI_API_KEY")



def detectar_areas_proyecto_ia(api_key: str, model_name: str, project_data: dict) -> list[str]:
    """PASS 1 V16: crea una lista cerrada de espacios antes de generar actividades."""
    client=genai.Client(api_key=api_key)
    prompt=f"""
Actúa como ANALISTA DE ALCANCE DE REMODELACIÓN.

DESCRIPCIÓN
{project_data['description']}

Devuelve una lista de ÁREAS FÍSICAS O FUNCIONALES.
REGLAS
1. Incluye General.
2. Después incluye cada espacio específico explícito.
3. No juntes baños distintos bajo 'Baños' ni recámaras distintas bajo 'Recámaras'.
4. Conserva calificadores: Baño principal, Baño de visitas, Recámara 1, Cocina, Sala, Roof top, Despacho, etc.
5. No incluyas oficios como Albañilería, Carpintería o Pintura.
6. Si dice '3 baños' sin nombres, normaliza Baño 1, Baño 2 y Baño 3.
7. No inventes espacios ajenos al alcance.
8. Esta lista será CERRADA para todo el presupuesto.
"""
    last=None
    for model in modelos_gemini(model_name):
        try:
            response=client.models.generate_content(model=model,contents=prompt,config=types.GenerateContentConfig(response_mime_type="application/json",response_schema=AreasProyectoIA,temperature=0.1))
            if response.text:
                parsed=AreasProyectoIA.model_validate_json(response.text)
                areas=normalizar_lista_areas(parsed.areas)
                if areas: return areas
        except Exception as exc:
            last=exc
    fallback=normalizar_lista_areas(detectar_areas_explicitas(project_data))
    if fallback: return fallback
    raise RuntimeError(f"No fue posible detectar áreas. Último error: {last}")


def generar_presupuesto_ia(
    api_key: str,
    model_name: str,
    project_data: dict,
    params: dict,
) -> PresupuestoIA:
    client = genai.Client(api_key=api_key)
    year = datetime.now().year
    budget_level = project_data.get("budget_level", "Medio-alto")
    level_criterion = criterio_nivel_presupuesto(budget_level)
    allowed_areas = areas_validas_proyecto(project_data)
    allowed_areas_text = "\n".join(f"- {x}" for x in allowed_areas)

    prompt = f"""
Actúa como un INGENIERO DE COSTOS SENIOR de una empresa de remodelación de alto
nivel que opera principalmente en Ciudad de México y SUBCONTRATA prácticamente
todas las actividades.

Tu objetivo es transformar el alcance en ACTIVIDADES COTIZABLES Y NEGOCIABLES.

CONFIGURACIÓN FIJA
- Referencia de mercado: Ciudad de México, {year}.
- Nivel comercial: {budget_level}.
- Criterio del nivel: {level_criterion}
- costo_unitario_estimado representa el costo integrado estimado del
  SUBCONTRATISTA para la empresa.
- No calcules 30% de marca ni IVA. Python/Excel lo harán después.

DATOS DEL PROYECTO
Cliente: {project_data['name']}
Ubicación: {project_data['location'] or 'No indicada'}
Tipo de obra: {project_data['project_type']}
Nivel: {budget_level}

DESCRIPCIÓN DEL USUARIO
{project_data['description']}

TEXTO GUÍA
{project_data['guide_text'] or 'Sin consideraciones adicionales.'}

ÁREAS VÁLIDAS DEL PROYECTO — LISTA CERRADA
{allowed_areas_text}

REGLAS DE ÁREA
- area DEBE ser exactamente uno de los valores anteriores.
- General NO es el valor predeterminado.
- Usa General solo para actividades verdaderamente comunes a toda la obra.
- Si título/descripción dicen Cocina, Baño principal, Recámara 1, etc., area debe ser ese espacio.
- Antes de asignar General, comprueba que ninguna área específica sea aplicable.
- En remodelaciones integrales la mayoría de actividades deben quedar en un espacio específico.

REGLA CENTRAL DE GRANULARIDAD
1. UNA ACTIVIDAD = UN TRABAJO COTIZABLE EN UNA SOLA ÁREA.
2. NO agrupes en una actividad trabajos correspondientes a áreas distintas,
   aunque sean del mismo oficio, misma partida o técnicamente similares.
3. Si existen Cocina, Baño 1, Baño 2 y Baño 3, sus trabajos deben permanecer
   separados cuando puedan cotizarse, modificarse, excluirse o contratarse
   independientemente.
4. Usa General únicamente para trabajos verdaderamente comunes a toda la obra.
5. NO inventes áreas. Usa solo espacios explícitos del usuario; si no puedes
   identificar uno de manera justificada, usa General.
6. Solo agrupa unidades si son equivalentes: misma área, mismo trabajo, misma
   especificación, misma unidad y misma base de precio.
7. Carpintería/mobiliario: muebles de tipo, diseño, función, especificación o
   dimensiones diferentes SIEMPRE son actividades independientes. Unidades
   idénticas pueden usar cantidad mayor a 1.
8. No conviertas el presupuesto en APU: no hagas filas por material, herramienta,
   cuadrilla o consumible.

ALCANCE
9. Incluye trabajos explícitos, previos indispensables y complementarios necesarios.
10. Incluye proyecto/ingenierías/licencias/permisos solo cuando sean previsibles.
11. No agregues trabajos opcionales ajenos al alcance.
12. Si una inclusión es inferida, indícalo brevemente.

PARTIDAS
13. Usa preferentemente:
   PROYECTO Y TRÁMITES; PRELIMINARES Y PROTECCIONES;
   DESMONTAJES Y DEMOLICIONES; ALBAÑILERÍA Y ESTRUCTURA;
   INSTALACIONES ELÉCTRICAS; INSTALACIONES HIDROSANITARIAS;
   ACABADOS Y RECUBRIMIENTOS; CARPINTERÍA; CANCELERÍA Y HERRERÍA;
   EXTERIORES Y AMENIDADES; LIMPIEZA Y ENTREGA.
14. Clasifica por la naturaleza principal del trabajo.
15. Protecciones temporales son PRELIMINARES; limpieza final es LIMPIEZA Y ENTREGA.

SECUENCIA Y PRESENTACIÓN
16. orden_ejecucion debe seguir la secuencia constructiva real.
17. subpartida debe ser corta.
18. titulo_comercial debe identificar el trabajo sin depender de la descripción.
19. descripcion_tecnica debe explicar qué se hace, dónde y qué incluye.
20. codigo_sugerido debe ser único y breve.

CANTIDADES
21. Calcula M2, ML, M3, PZA u otras cantidades cuando los datos lo permitan.
22. Si faltan dimensiones, usa PZA/LOTE/JGO cuando sea más profesional que inventar.
23. Si un mismo trabajo aparece en varias áreas, cuantifica cada área por separado.

COSTOS
24. costo_unitario_estimado es el costo integrado esperado del subcontratista.
25. Debe ser razonable para CDMX {year} y el nivel {budget_level}.
26. Si es variable/especializado, requiere_cotizacion=True.
27. No desarrolles materiales/mano de obra/otros/desperdicio por separado.
28. No calcules indirectos, utilidad, precio final, margen, 30% de marca ni IVA.

CONTROL
29. No dupliques actividades equivalentes dentro de la misma área.
30. Verifica que ninguna actividad mezcle cocina, baños distintos, recámaras
    distintas u otras áreas independientes.
31. Criterio, fundamento y consideraciones deben ser breves y revisables.
32. NIVEL DE CONFIANZA DE CANTIDAD:
    - Alta: únicamente cuando la cantidad sale directamente de números,
      dimensiones o conteos explícitos del usuario, sin supuestos relevantes.
    - Media: cuando parte de datos explícitos pero requiere interpretación,
      geometría aproximada o una decisión técnica menor.
    - Baja: cuando la cantidad es estimada, supuesta, promediada, inferida o
      falta información importante.
33. NIVEL DE CONFIANZA DE PRECIO:
    costo_unitario_estimado es una estimación; no uses Alta solo porque el trabajo
    sea común. La aplicación elevará después la confianza si encuentra una base
    interna validada o una referencia externa sólida.
34. No expongas cadenas de pensamiento.
"""

    modelos = modelos_gemini(model_name)

    last_error = None
    for model in modelos:
        try:
            response = client.models.generate_content(
                model=model,
                contents=prompt,
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_schema=PresupuestoIA,
                ),
            )
            if not response.text:
                raise RuntimeError(f"Gemini ({model}) devolvió una respuesta vacía.")
            return PresupuestoIA.model_validate_json(response.text)
        except Exception as exc:
            last_error = exc
            msg = str(exc).lower()
            model_error = (
                "404" in msg
                or "not_found" in msg
                or "no longer available" in msg
                or ("model" in msg and "not available" in msg)
            )
            if not model_error:
                raise

    raise RuntimeError(
        f"No fue posible usar un modelo Gemini disponible. Último error: {last_error}"
    )



def auditar_estructura_presupuesto_ia(
    api_key: str,
    model_name: str,
    project_data: dict,
    result: PresupuestoIA,
) -> PresupuestoIA:
    """
    Auditoría V15 de granularidad, área, clasificación y secuencia.
    Puede dividir una actividad que mezcle áreas antes de resolver precios.
    """
    if not result.actividades:
        return result

    client = genai.Client(api_key=api_key)
    current = [x.model_dump() for x in result.actividades]

    prompt = f"""
Actúa como AUDITOR SENIOR DE PRESUPUESTOS DE REMODELACIÓN.

PROYECTO
Tipo: {project_data['project_type']}
Ubicación: {project_data['location']}
Descripción original:
{project_data['description']}

ÁREAS VÁLIDAS — LISTA CERRADA
{chr(10).join(f"- {x}" for x in areas_validas_proyecto(project_data))}

ACTIVIDADES
{json.dumps(current, ensure_ascii=False, separators=(',', ':'))}

Devuelve un PresupuestoIA completo corregido.

REGLAS
1. UNA ACTIVIDAD = UN TRABAJO COTIZABLE EN UNA SOLA ÁREA.
2. Puedes DIVIDIR una actividad si mezcla áreas o trabajos negociables por separado.
3. No juntes Cocina con baños, baños distintos entre sí ni recámaras distintas
   solo porque sean del mismo oficio.
4. Carpintería/mobiliario: muebles distintos permanecen separados.
5. Solo agrupa unidades realmente equivalentes dentro de la misma área.
6. area debe ser EXACTAMENTE una de las áreas válidas. General es excepcional; si el concepto reside en un espacio específico, asígnalo a ese espacio.
7. Si divides una actividad, no dupliques metrajes/cantidades/costos. Si no hay
   datos suficientes para dividir responsablemente, conserva General y explica
   la incertidumbre en consideraciones.
8. No conviertas el presupuesto en APU.
9. Corrige partida/subpartida según la naturaleza del trabajo.
10. Corrige orden_ejecucion según dependencias constructivas.
11. costo_unitario_estimado sigue siendo costo del subcontratista.
12. No calcules 30% de marca ni IVA.
13. Códigos únicos.
14. No incluyas explicación fuera del JSON.
"""

    models = []
    for model in modelos_gemini(model_name):
        if model and model not in models:
            models.append(model)

    for model in models:
        try:
            response = client.models.generate_content(
                model=model,
                contents=prompt,
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_schema=PresupuestoIA,
                ),
            )
            if not response.text:
                continue

            audited = PresupuestoIA.model_validate_json(response.text)
            fixed, used = [], set()
            for idx, act in enumerate(audited.actividades, 1):
                code = limpiar_codigo(act.codigo_sugerido, f"CON-{idx:03d}")
                base, counter = code, 2
                while code.upper() in used:
                    code = f"{base}-{counter}"
                    counter += 1
                used.add(code.upper())
                fixed.append(
                    act.model_copy(
                        update={
                            "area": resolver_area_actividad(
                                project_data,
                                act.area,
                                act.titulo_comercial,
                                act.descripcion_tecnica,
                                act.subpartida,
                            ),
                            "partida": normalizar_seccion_comercial(act.partida),
                            "codigo_sugerido": code,
                        }
                    )
                )
            return audited.model_copy(update={"actividades": fixed})
        except Exception:
            continue

    fallback = [
        act.model_copy(
            update={
                "area": resolver_area_actividad(
                    project_data,
                    act.area,
                    act.titulo_comercial,
                    act.descripcion_tecnica,
                    act.subpartida,
                )
            }
        )
        for act in result.actividades
    ]
    return result.model_copy(update={"actividades": fallback})


def sincronizar_items_con_estructura(
    result: PresupuestoIA,
    items: list[dict],
) -> list[dict]:
    by_code = {
        str(act.codigo_sugerido or "").strip().upper(): act
        for act in result.actividades
    }
    output = []
    for item in items:
        out = dict(item)
        act = by_code.get(str(out.get("code") or "").strip().upper())
        if act is not None:
            out["area"] = act.area
            out["category"] = normalizar_seccion_comercial(act.partida)
            out["subcategory"] = act.subpartida.strip()
            out["execution_order"] = int(act.orden_ejecucion)
        output.append(out)
    return output


def revisar_presupuesto_ia(
    api_key: str,
    model_name: str,
    project_data: dict,
    params: dict,
    current_result: PresupuestoIA,
    current_items: list[dict],
    revision_request: str,
) -> RevisionPresupuestoIA:
    """
    Ajusta el presupuesto vigente sin regenerarlo por completo.
    La solicitud puede referirse a alcance, cantidades, descripciones o precios.
    """
    client = genai.Client(api_key=api_key)

    presupuesto_actual = [
        {
            "codigo": x["code"],
            "area": area_item(x),
            "partida": x["category"],
            "subpartida": x["subcategory"],
            "orden_ejecucion": x.get("execution_order", 500),
            "titulo_comercial": titulo_comercial_item(x),
            "descripcion": x["description"],
            "unidad": x["unit"],
            "cantidad": x["quantity"],
            "costo_subcontratista_unitario": x["unit_cost"],
            "precio_interno_vigente": x["unit_sale"],
            "fuente_precio": x["price_source"],
            "criterio_cantidad": x["quantity_criterion"],
            "consideraciones": x["considerations"],
        }
        for x in current_items
    ]

    prompt = f"""
Actúa como revisor técnico y de costos de un presupuesto de remodelación e interiorismo.
NO vuelvas a generar el presupuesto desde cero. Trabaja únicamente sobre las actividades
que necesiten cambiar.

DATOS DEL PROYECTO
Cliente: {project_data['name']}
Ubicación: {project_data['location']}
Tipo de obra: {project_data['project_type']}
Nivel de presupuesto: {project_data.get('budget_level', 'Medio-alto')}

DESCRIPCIÓN ORIGINAL
{project_data['description']}

TEXTO GUÍA
{project_data['guide_text'] or 'Sin texto guía adicional.'}

PRESUPUESTO ACTUAL
{json.dumps(presupuesto_actual, ensure_ascii=False, separators=(',', ':'))}

ALCANCE ACTUAL
{current_result.alcance_resumido}

PETICIÓN DEL USUARIO
{revision_request}

INSTRUCCIONES
1. Cambia solamente lo necesario.
2. Puedes AGREGAR, MODIFICAR o ELIMINAR actividades.
3. Para MODIFICAR/ELIMINAR usa exactamente codigo_objetivo.
4. Para AGREGAR/MODIFICAR devuelve la ActividadIA completa.
5. UNA ACTIVIDAD = UN TRABAJO COTIZABLE EN UNA SOLA ÁREA.
6. Trabajos iguales en áreas distintas deben ser actividades distintas.
7. Si una actividad actual mezcla áreas, puedes usar ELIMINAR + varias AGREGAR.
8. No inventes áreas; usa espacios explícitos o General.
9. Carpintería/mobiliario: muebles distintos van separados; idénticos pueden usar cantidad N.
10. Si cambia el costo del subcontratista, modifica costo_unitario_estimado y
    usa recalcular_precio=True.
11. Si cambia cantidad/descripción sin alterar costo, usa recalcular_precio=False.
12. No modifiques precios no mencionados.
13. No desarrolles APU ni desgloses de materiales/mano de obra.
14. Conserva área/partida/subpartida/título/orden cuando no estén afectados.
15. No calcules 30% de marca ni IVA.
16. Devuelve alcance/consideraciones/datos faltantes completos y actualizados.
17. En motivo usa una explicación breve, sin razonamiento interno.

"""

    modelos = modelos_gemini(model_name)

    last_error = None
    for model in modelos:
        try:
            response = client.models.generate_content(
                model=model,
                contents=prompt,
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_schema=RevisionPresupuestoIA,
                ),
            )
            if not response.text:
                raise RuntimeError(f"Gemini ({model}) devolvió una revisión vacía.")
            return RevisionPresupuestoIA.model_validate_json(response.text)
        except Exception as exc:
            last_error = exc
            msg = str(exc).lower()
            model_error = (
                "404" in msg
                or "not_found" in msg
                or "no longer available" in msg
                or ("model" in msg and "not available" in msg)
            )
            if not model_error:
                raise

    raise RuntimeError(
        f"No fue posible usar un modelo Gemini para la revisión. Último error: {last_error}"
    )


# =========================================================
# MOTOR DE PRECIOS
# =========================================================

def investigar_precio_web_ia(api_key: str, model_name: str, actividad: ActividadIA, project_data: dict, intento: int = 1) -> InvestigacionPrecioWebIA | None:
    if not api_key: return None
    unit=normalizar_unidad(actividad.unidad)
    if unit in {"","%"}: return None
    client=genai.Client(api_key=api_key)
    year=datetime.now().year
    strategy=("Busca referencias muy comparables de precio unitario en Ciudad de México o zona metropolitana. Prioriza servicios instalados y contratistas." if intento==1 else "Amplía a México, fabricantes, instaladores y catálogos comerciales, manteniendo unidad y alcance técnico.")
    prompt=f"""
Investiga el PRECIO UNITARIO ACTUAL de esta actividad.
Área: {actividad.area}
Partida: {actividad.partida}
Título: {actividad.titulo_comercial}
Descripción: {actividad.descripcion_tecnica}
Unidad obligatoria: {actividad.unidad}
Nivel: {project_data.get('budget_level','Medio-alto')}
Ubicación: Ciudad de México
Año: {year}
{strategy}
REGLAS
1. Devuelve 2 a 5 referencias cuando existan.
2. El precio debe corresponder exactamente a {actividad.unidad}.
3. tipo_precio: SUBCONTRATISTA, SUMINISTRO_INSTALACION, PRECIO_PUBLICO_INSTALADO o MATERIAL_SOLO.
4. Para servicio integrado no uses material suelto como referencia principal.
5. Incluye URL pública concreta; no inventes URLs ni precios.
6. Si no encuentras evidencia útil, referencias=[].
7. No agregues 30% de franquicia ni IVA.
"""
    for model in modelos_gemini(model_name):
        try:
            interaction=client.interactions.create(model=model,input=prompt,tools=[{"type":"google_search"}],response_format={"type":"text","mime_type":"application/json","schema":InvestigacionPrecioWebIA.model_json_schema()})
            raw=getattr(interaction,'output_text',None)
            if raw: return InvestigacionPrecioWebIA.model_validate_json(raw)
        except Exception:
            try:
                grounding_tool=types.Tool(google_search=types.GoogleSearch())
                response=client.models.generate_content(model=model,contents=prompt,config=types.GenerateContentConfig(tools=[grounding_tool],response_mime_type="application/json",response_schema=InvestigacionPrecioWebIA,temperature=0.1))
                if response.text: return InvestigacionPrecioWebIA.model_validate_json(response.text)
            except Exception:
                continue
    return None




def buscar_precio_interno(db: Database, actividad: ActividadIA) -> dict | None:
    candidatos = db.price_candidates(actividad.unidad)
    best = None
    best_score = 0.0

    for row in candidatos:
        score = score_similitud(actividad.descripcion_tecnica, row["description"])
        if score > best_score:
            best_score = score
            best = row

    if best is None or best_score < 0.82:
        return None

    original_source = (best.get("source") or "").upper()
    original_status = (best.get("status") or "").upper()

    if original_status in {"VALIDADO", "COSTO_REAL", "COTIZADO_PROVEEDOR"}:
        source = "BASE_INTERNA"
        confidence = "Alta"
    elif original_source == "IA_ESTIMADO" or original_status == "ESTIMADO_IA":
        source = "HISTORICO_IA"
        confidence = "Media" if best_score >= 0.9 else "Baja"
    elif (
        original_source in {"REFERENCIA_CDMX", "REFERENCIA_EXTERNA", "HISTORICO_EXTERNO"}
        or original_status == "REFERENCIA_EXTERNA"
    ):
        source = "HISTORICO_EXTERNO"
        confidence = "Media" if best_score >= 0.9 else "Baja"
    else:
        source = "BASE_INTERNA"
        confidence = best.get("confidence") or "Media"

    return {
        "concept_id": best["concept_id"],
        "unit_cost": float(best["unit_cost"]),
        "source": source,
        "source_detail": (
            f"Coincidencia {best_score:.0%} con: {best['description']} "
            f"| origen previo: {best.get('source') or 'sin dato'}"
        ),
        "status": original_status or "HISTORICO",
        "confidence": confidence,
        "match_score": best_score,
    }


def buscar_precio_externo(
    db: Database,
    actividad: ActividadIA,
    project_data: dict,
    params: dict,
) -> dict | None:
    """
    Busca la actividad en el catálogo CDMX activo.

    El P.U. oficial se trata como una referencia integrada de venta antes de IVA,
    NO como costo directo de la empresa. Cuando la coincidencia es fuerte se
    calcula un costo base equivalente que, después de aplicar los indirectos y la
    utilidad configurados en la app, reproduce el P.U. de referencia CDMX.
    """
    unit = normalizar_unidad(actividad.unidad)

    # Las partidas globales son demasiado ambiguas para cruzarlas
    # automáticamente contra un catálogo de precios unitarios.
    if unit in {"", "LOTE", "JGO", "PTO", "SERV", "%"}:
        return None

    candidates = db.external_candidates(
        source="CDMX",
        unit=actividad.unidad,
        description=actividad.descripcion_tecnica,
        limit=350,
    )
    if not candidates:
        return None

    scored = []
    for row in candidates:
        score = score_similitud_externa(
            actividad.descripcion_tecnica,
            row["description"],
        )
        scored.append((score, row))

    scored.sort(key=lambda x: x[0], reverse=True)
    best_score, best = scored[0]

    # Por debajo de 72 % se considera que no existe una referencia útil.
    if best_score < 0.72:
        return None

    reference_price = float(best["unit_price"])
    use_automatically = best_score >= 0.86

    # El flujo financiero interno define:
    # venta = costo * (1 + indirectos) * (1 + utilidad)
    # Por lo tanto se despeja el costo base para que la venta antes de IVA
    # coincida con el P.U. CDMX cuando sí se decide utilizarlo.
    factor = (
        (1.0 + params["indirect_pct"] / 100.0)
        * (1.0 + params["profit_pct"] / 100.0)
    )
    equivalent_cost = reference_price / factor if factor > 0 else reference_price

    month = int(best.get("month") or 0)
    year = int(best.get("year") or 0)
    edition = best.get("version_label") or (
        f"{SPANISH_MONTH_NAMES.get(month, '')} {year}".strip()
    )

    detail = (
        f"CDMX {edition} | clave {best['source_code']} | "
        f"P.U. oficial ${reference_price:,.2f}/{best['unit']} | "
        f"coincidencia {best_score:.0%} | "
        f"{best['description']}"
    )

    return {
        "unit_cost": equivalent_cost,
        "reference_unit_price": reference_price,
        "source": "REFERENCIA_CDMX",
        "source_detail": detail,
        "status": "REFERENCIA_EXTERNA",
        "confidence": "Alta" if best_score >= 0.92 else "Media",
        "match_score": best_score,
        "use_automatically": use_automatically,
        "source_code": best["source_code"],
        "catalog": edition,
        "source_url": best.get("source_url") or "",
    }


def resolver_items(
    db: Database,
    result: PresupuestoIA,
    project_data: dict,
    params: dict,
    force_new_price_codes: set[str] | None = None,
    api_key: str | None = None,
    model_name: str = "gemini-3.8-flash",
) -> list[dict]:
    """Base interna > CDMX > Google Search > IA último recurso."""
    items=[]
    force_new_price_codes={str(x).strip().upper() for x in (force_new_price_codes or set())}
    for idx,act in enumerate(result.actividades,start=1):
        requested_code=limpiar_codigo(act.codigo_sugerido,f"CON-{idx:03d}")
        force_new_price=requested_code.upper() in force_new_price_codes
        internal=None if force_new_price else buscar_precio_interno(db,act)
        external=None if internal or force_new_price else buscar_precio_externo(db,act,project_data,params)
        web_price=None
        if not internal and not (external and external.get("use_automatically")):
            inv=investigar_precio_web_ia(api_key or "",model_name,act,project_data,1)
            if inv: web_price=calcular_precio_web(inv,act,params)
            if not web_price or web_price.get("evidence_score",0)<65:
                inv2=investigar_precio_web_ia(api_key or "",model_name,act,project_data,2)
                if inv2:
                    cand=calcular_precio_web(inv2,act,params)
                    if cand and cand.get("evidence_score",0)>(web_price or {}).get("evidence_score",0): web_price=cand
        if internal:
            unit_cost=float(internal["unit_cost"]); concept_id=internal["concept_id"]; source=internal["source"]; source_detail=internal["source_detail"]; price_status=internal["status"]
            evidence_score=score_confianza_precio(source,source_detail); price_confidence=etiqueta_confianza_score(evidence_score)
        elif external and external.get("use_automatically"):
            unit_cost=float(external["unit_cost"]); concept_id=None; source="REFERENCIA_CDMX"; source_detail=external["source_detail"]; price_status="REFERENCIA_EXTERNA"
            evidence_score=88 if external.get("match_score",0)>=0.92 else 76; price_confidence=etiqueta_confianza_score(evidence_score)
        elif web_price:
            unit_cost=float(web_price["unit_cost"]); concept_id=None; source="WEB_FUNDAMENTADO"; source_detail=web_price["source_detail"]; price_status="REFERENCIA_WEB"; evidence_score=int(web_price["evidence_score"]); price_confidence=etiqueta_confianza_score(evidence_score)
            if external: source_detail += " | CDMX secundario: " + external["source_detail"]
        else:
            unit_cost=float(act.costo_unitario_estimado); concept_id=None; source="IA_ESTIMADO"; source_detail="Último recurso: sin referencia interna, CDMX fuerte ni evidencia web consolidable."; price_status="ESTIMADO_IA"; evidence_score=30; price_confidence="Baja"
            if external: source_detail += " | CDMX débil: " + external["source_detail"]
        quantity=max(float(act.cantidad),0.0)
        indirect_unit=unit_cost*params["indirect_pct"]/100.0
        profit_unit=(unit_cost+indirect_unit)*params["profit_pct"]/100.0
        internal_unit=unit_cost+indirect_unit+profit_unit
        contractor_amount=quantity*unit_cost; internal_amount=quantity*internal_unit
        benefit=internal_amount-contractor_amount; margin=benefit/internal_amount*100.0 if internal_amount else 0.0
        considerations=act.consideraciones.strip()
        if evidence_score<65:
            considerations=((considerations+" | ") if considerations else "")+"Precio con evidencia insuficiente; recomendable confirmar con proveedor."
        if act.requiere_cotizacion and evidence_score<85:
            note="Actividad especializada: conviene cotización directa."
            if note not in considerations: considerations=((considerations+" | ") if considerations else "")+note
        area=resolver_area_actividad(project_data,act.area,act.titulo_comercial,act.descripcion_tecnica,act.subpartida)
        items.append({
            "concept_id":concept_id,"area":area,"included":True,"package_group":"","category":normalizar_seccion_comercial(act.partida),"subcategory":act.subpartida.strip(),"code":requested_code,"execution_order":int(act.orden_ejecucion),"commercial_title":act.titulo_comercial.strip(),"description":act.descripcion_tecnica.strip(),"unit":act.unidad.strip().upper(),"quantity":quantity,"unit_cost":unit_cost,"direct_amount":contractor_amount,"unit_indirect":indirect_unit,"unit_profit":profit_unit,"unit_sale":internal_unit,"original_unit_sale":internal_unit,"sale_amount":internal_amount,"benefit_amount":benefit,"sale_margin_pct":margin,"price_source":source,"price_source_detail":source_detail,"price_status":price_status,"price_confidence":price_confidence,"price_evidence_score":evidence_score,"quantity_confidence":act.nivel_confianza_cantidad,"quantity_criterion":act.criterio_cantidad.strip(),"inclusion_basis":act.fundamento_inclusion.strip(),"considerations":considerations,"material_share_pct":0.0,"labor_share_pct":0.0,"other_share_pct":100.0,"waste_reference_pct":0.0,"area_allocations":[{"area":area,"porcentaje":100.0,"cantidad_referencia":quantity,"criterio":"Área directa V16.","confianza":"Alta"}],
        })
    return items


def recalcular_item_financiero(item: dict, params: dict) -> dict:
    out = dict(item)
    unit_cost = float(out["unit_cost"])
    quantity = max(float(out["quantity"]), 0.0)

    indirect_unit = unit_cost * params["indirect_pct"] / 100.0
    profit_unit = (unit_cost + indirect_unit) * params["profit_pct"] / 100.0
    internal_unit = unit_cost + indirect_unit + profit_unit

    contractor_amount = quantity * unit_cost
    internal_amount = quantity * internal_unit
    benefit = internal_amount - contractor_amount
    margin = benefit / internal_amount * 100.0 if internal_amount else 0.0

    out.update({
        "quantity": quantity,
        "unit_cost": unit_cost,
        "direct_amount": contractor_amount,
        "unit_indirect": indirect_unit,
        "unit_profit": profit_unit,
        "unit_sale": internal_unit,
        "sale_amount": internal_amount,
        "benefit_amount": benefit,
        "sale_margin_pct": margin,
    })
    out.setdefault("original_unit_sale", internal_unit)
    out.setdefault("included", True)
    out.setdefault("package_group", "")
    out.setdefault("area", AREA_GENERAL)
    return out


def item_a_actividad(item: dict) -> ActividadIA:
    return ActividadIA(
        area=area_item(item),
        partida=item["category"],
        subpartida=item["subcategory"],
        codigo_sugerido=item["code"],
        orden_ejecucion=int(item.get("execution_order") or 500),
        titulo_comercial=titulo_comercial_item(item),
        descripcion_tecnica=item["description"],
        unidad=item["unit"],
        cantidad=float(item["quantity"]),
        costo_unitario_estimado=float(item["unit_cost"]),
        criterio_cantidad=item.get("quantity_criterion") or "Cantidad de la versión vigente.",
        fundamento_inclusion=item.get("inclusion_basis") or "Actividad incluida en el alcance vigente.",
        nivel_confianza_cantidad=item.get("quantity_confidence") or "Media",
        nivel_confianza_precio=item.get("price_confidence") or "Media",
        requiere_cotizacion="requiere cotización" in (item.get("considerations") or "").lower(),
        consideraciones=item.get("considerations") or "",
    )


def aplicar_revision_estructural(
    db: Database,
    current_result: PresupuestoIA,
    current_items: list[dict],
    revision: RevisionPresupuestoIA,
    project_data: dict,
    params: dict,
    api_key: str | None = None,
    model_name: str = "gemini-3.8-flash",
) -> tuple[PresupuestoIA, list[dict], list[str]]:
    """
    Aplica las operaciones devueltas por Gemini. Las actividades no mencionadas
    se copian literalmente desde la versión anterior.
    """
    if not revision.operaciones:
        raise RuntimeError(
            "Gemini no identificó operaciones estructurales. "
            "Revise que la solicitud describa un cambio importante de alcance."
        )

    operations = []
    activities_to_resolve = []
    force_codes = set()

    for op in revision.operaciones:
        action = (op.accion or "").strip().upper()
        if action not in {"AGREGAR", "MODIFICAR", "ELIMINAR"}:
            raise RuntimeError(f"Acción de revisión no válida: {op.accion}")

        if action in {"AGREGAR", "MODIFICAR"}:
            if op.actividad is None:
                raise RuntimeError(f"La acción {action} no contiene una actividad completa.")
            activities_to_resolve.append(op.actividad)
            if action == "AGREGAR" or op.recalcular_precio:
                force_codes.add(
                    limpiar_codigo(
                        op.actividad.codigo_sugerido,
                        f"REV-{len(activities_to_resolve):03d}",
                    )
                )

        operations.append((action, op))

    resolved_changes = []
    if activities_to_resolve:
        temp_result = PresupuestoIA(
            nombre_proyecto=current_result.nombre_proyecto,
            actividad_principal=revision.actividad_principal_actualizada,
            alcance_resumido=revision.alcance_resumido_actualizado,
            consideraciones_generales=revision.consideraciones_generales_actualizadas,
            datos_faltantes=revision.datos_faltantes_actualizados,
            actividades=activities_to_resolve,
        )
        resolved_changes = resolver_items(
            db, verificar_cantidades_python(temp_result), project_data, params,
            force_new_price_codes=force_codes, api_key=api_key, model_name=model_name,
        )

    items = [dict(x) for x in current_items]
    resolved_index = 0
    change_log = []

    def find_index(code: str) -> int:
        code_n = (code or "").strip().upper()
        for i, item in enumerate(items):
            if str(item.get("code") or "").strip().upper() == code_n:
                return i
        raise RuntimeError(
            f"La revisión hizo referencia al código '{code}', "
            "pero ese código no existe en el presupuesto actual."
        )

    for action, op in operations:
        if action == "ELIMINAR":
            idx = find_index(op.codigo_objetivo)
            removed = items.pop(idx)
            change_log.append(
                f"ELIMINADO {removed['code']}: {titulo_comercial_item(removed)}"
            )
            continue

        new_item = dict(resolved_changes[resolved_index])
        resolved_index += 1

        if action == "MODIFICAR":
            idx = find_index(op.codigo_objetivo)
            old_item = items[idx]
            new_item["original_unit_sale"] = old_item.get(
                "original_unit_sale", old_item.get("unit_sale")
            )
            new_item["included"] = old_item.get("included", True)
            new_item["package_group"] = old_item.get("package_group", "")

            same_unit = (
                str(new_item.get("unit") or "").strip().upper()
                == str(old_item.get("unit") or "").strip().upper()
            )
            desc_similarity = SequenceMatcher(
                None,
                normalizar_texto(str(old_item.get("description") or "")),
                normalizar_texto(str(new_item.get("description") or "")),
            ).ratio()

            preserve_old_price = not op.recalcular_precio and same_unit
            if preserve_old_price:
                for key in [
                    "concept_id",
                    "unit_cost",
                    "price_source",
                    "price_source_detail",
                    "price_status",
                    "price_confidence",
                ]:
                    new_item[key] = old_item.get(key)
                new_item = recalcular_item_financiero(new_item, params)
            elif same_unit and desc_similarity >= 0.72:
                # Es el mismo concepto con una nueva valoración (por ejemplo,
                # "este precio está muy bajo"). Conservamos su identidad para
                # registrar un nuevo precio histórico en lugar de duplicarlo.
                new_item["concept_id"] = old_item.get("concept_id")
                new_item["record_new_price"] = True

            # Evita que una modificación cambie el código a uno que ya pertenece
            # a otra actividad.
            proposed = str(new_item["code"]).strip().upper()
            conflicts = {
                str(x.get("code") or "").strip().upper()
                for j, x in enumerate(items)
                if j != idx
            }
            if proposed in conflicts:
                new_item["code"] = old_item["code"]

            items[idx] = new_item
            change_log.append(
                f"MODIFICADO {old_item['code']}: {op.motivo.strip() or 'Cambio de alcance'}"
            )
            continue

        # AGREGAR
        existing_codes = {
            str(x.get("code") or "").strip().upper() for x in items
        }
        base_code = str(new_item["code"]).strip().upper() or "REV"
        candidate = base_code
        counter = 2
        while candidate in existing_codes:
            candidate = f"{base_code}-{counter}"
            counter += 1
        new_item["code"] = candidate

        # Inserta después de la última actividad de la misma partida para
        # mantener el Excel ordenado por bloques.
        insert_at = len(items)
        same_category = [
            i for i, x in enumerate(items)
            if str(x.get("category") or "").strip().upper()
            == str(new_item.get("category") or "").strip().upper()
        ]
        if same_category:
            insert_at = max(same_category) + 1
        items.insert(insert_at, new_item)
        change_log.append(
            f"AGREGADO {new_item['code']}: {titulo_comercial_item(new_item)}"
        )

    revised_result = PresupuestoIA(
        nombre_proyecto=current_result.nombre_proyecto,
        actividad_principal=revision.actividad_principal_actualizada,
        alcance_resumido=revision.alcance_resumido_actualizado,
        consideraciones_generales=revision.consideraciones_generales_actualizadas,
        datos_faltantes=revision.datos_faltantes_actualizados,
        actividades=[item_a_actividad(x) for x in items],
    )

    return revised_result, items, change_log


def calcular_financieros(items: list[dict], params: dict) -> dict:
    active = [x for x in items if item_included(x)]
    direct_cost = sum(float(x.get("direct_amount") or 0.0) for x in active)
    indirect_cost = direct_cost * params["indirect_pct"] / 100.0
    profit = (direct_cost + indirect_cost) * params["profit_pct"] / 100.0
    sale_before_tax = sum(float(x.get("sale_amount") or 0.0) for x in active)
    iva_amount = sale_before_tax * params["iva_pct"] / 100.0
    total = sale_before_tax + iva_amount
    return {
        "direct_cost": direct_cost,
        "indirect_cost": indirect_cost,
        "profit": profit,
        "sale_before_tax": sale_before_tax,
        "iva_amount": iva_amount,
        "total": total,
    }


# =========================================================
# IMPORTAR PRESUPUESTO DESDE EXCEL
# =========================================================


def _valor_float_excel(value, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip()
    raw = raw.replace("$", "").replace(",", "")
    raw = raw.replace("%", "")
    try:
        return float(raw)
    except Exception:
        return default


def _pct_excel_a_porcentaje(value, default: float) -> float:
    number = _valor_float_excel(value, default)
    # Excel suele guardar 12 % como 0.12.
    if 0 <= number <= 1:
        return number * 100.0
    return number


def _quitar_numeracion_excel(value: str) -> str:
    raw = str(value or "").strip()
    raw = re.sub(r"^\s*\d+(?:\.\d+)*\.?\s*", "", raw)
    return raw.strip()


def _buscar_encabezado_presupuesto(ws) -> int | None:
    """Busca la fila de Partida/Subpartida/... sin depender de una versión."""
    expected = {
        "PARTIDA",
        "SUBPARTIDA",
        "DESCRIPCION TECNICA",
        "UNIDAD",
    }

    for row in range(1, min(ws.max_row, 150) + 1):
        values = {
            normalizar_texto(ws.cell(row, col).value).upper()
            for col in range(1, min(ws.max_column, 12) + 1)
        }
        if expected.issubset(values):
            return row
    return None


def _buscar_hoja_presupuesto(workbook):
    preferred = [
        "01 Presupuesto",
        "Presupuesto",
        "PRESUPUESTO",
    ]
    for name in preferred:
        if name in workbook.sheetnames:
            ws = workbook[name]
            header = _buscar_encabezado_presupuesto(ws)
            if header:
                return ws, header

    for ws in workbook.worksheets:
        header = _buscar_encabezado_presupuesto(ws)
        if header:
            return ws, header

    raise RuntimeError(
        "No encontré una tabla de presupuesto reconocible. "
        "Se requieren las columnas Partida, Subpartida, Descripción Técnica y Unidad."
    )


def _mapear_columnas_excel(ws, header_row: int) -> dict:
    aliases = {
        "code": {"CODIGO", "CLAVE"},
        "area": {"AREA"},
        "partida": {"PARTIDA", "CATEGORIA"},
        "subpartida": {"SUBPARTIDA"},
        "title": {"TITULO COMERCIAL", "TITULO"},
        "description": {"DESCRIPCION TECNICA", "DESCRIPCION", "CONCEPTO"},
        "unit": {"UNIDAD"},
        "quantity": {"CANT", "CANTIDAD", "CANT."},
        "unit_price": {"PRECIO UNITARIO MXN", "PRECIO UNITARIO", "P U", "P U VENTA"},
        "amount": {
            "IMPORTE INTERNO MXN", "IMPORTE INTERNO",
            "IMPORTE TOTAL MXN", "IMPORTE TOTAL", "IMPORTE"
        },
        "final_amount": {"IMPORTE FINAL MXN", "IMPORTE FINAL"},
        "included": {"CONSIDERAR", "INCLUIR"},
    }
    mapping = {}
    for col in range(1, ws.max_column + 1):
        key = normalizar_texto(ws.cell(header_row, col).value).upper()
        for field, options in aliases.items():
            if key in options and field not in mapping:
                mapping[field] = col

    required = {"partida", "description", "unit"}
    missing = required - set(mapping)
    if missing:
        raise RuntimeError("Faltan columnas necesarias: " + ", ".join(sorted(missing)))
    return mapping


def _leer_parametros_control(workbook, fallback_params: dict) -> tuple[dict, list[dict]]:
    params = dict(fallback_params)
    control_rows = []
    if "02 Control Interno" not in workbook.sheetnames:
        return params, control_rows

    ws = workbook["02 Control Interno"]

    # Compatibilidad legacy.
    for row in range(1, min(ws.max_row, 30) + 1):
        label = normalizar_texto(ws.cell(row, 1).value).upper()
        value = ws.cell(row, 2).value
        if label == "INDIRECTOS":
            params["indirect_pct"] = _pct_excel_a_porcentaje(value, params["indirect_pct"])
        elif label == "UTILIDAD":
            params["profit_pct"] = _pct_excel_a_porcentaje(value, params["profit_pct"])
        elif label == "IVA":
            params["iva_pct"] = _pct_excel_a_porcentaje(value, params["iva_pct"])
        elif "DESPERDICIO" in label:
            params["waste_pct"] = _pct_excel_a_porcentaje(value, params["waste_pct"])

    header_row, headers = None, {}
    for row in range(1, min(ws.max_row, 100) + 1):
        current = {}
        for col in range(1, ws.max_column + 1):
            value = normalizar_texto(ws.cell(row, col).value).upper()
            if value:
                current[value] = col
        has_code = "CODIGO" in current or "CLAVE" in current
        has_cost = "COSTO SUBCONTRATISTA UNIT" in current or "COSTO BASE UNIT" in current
        if has_code and has_cost:
            header_row, headers = row, current
            break

    if header_row is None:
        return params, control_rows

    def col(*names):
        for name in names:
            if name in headers:
                return headers[name]
        return None

    c_code = col("CODIGO", "CLAVE")
    c_title = col("TITULO COMERCIAL", "TITULO")
    c_cost = col("COSTO SUBCONTRATISTA UNIT", "COSTO BASE UNIT")
    c_original = col("P U INTERNO ORIGINAL")
    c_package = col("LOTE PAQUETE", "LOTE", "PAQUETE")

    for row in range(header_row + 1, ws.max_row + 1):
        code = str(ws.cell(row, c_code).value or "").strip() if c_code else ""
        title = str(ws.cell(row, c_title).value or "").strip() if c_title else ""
        cost = _valor_float_excel(ws.cell(row, c_cost).value) if c_cost else 0.0
        if not code and not title and cost == 0:
            continue
        control_rows.append({
            "code": code,
            "title": title,
            "unit_cost": cost,
            "original_unit_sale": (
                _valor_float_excel(ws.cell(row, c_original).value) if c_original else 0.0
            ),
            "package_group": (
                str(ws.cell(row, c_package).value or "").strip() if c_package else ""
            ),
        })

    return params, control_rows


def _leer_metadatos_excel(ws, fallback_name: str = "") -> dict:
    """
    Lee el encabezado propio de la app cuando está disponible.

    Si el archivo es un presupuesto anterior o externo cuya primera fila ya es
    la tabla de conceptos, NO interpreta las primeras actividades como nombre,
    tipo de obra o ubicación.
    """
    safe_name = Path(str(fallback_name or "")).stem.strip() or "Proyecto importado"

    project_type = "Remodelación interior general"
    budget_level = "Medio-alto"
    location = "Ubicación por confirmar"
    project_code = "IMPORTADO"
    version = 1
    name = safe_name

    first_cell = normalizar_texto(ws["A1"].value or "").upper()
    meta = str(ws["A3"].value or "").strip()

    # Solo interpretar A2/A3 como metadatos si realmente existe el encabezado
    # producido por esta aplicación.
    has_app_header = (
        first_cell == "PRESUPUESTO"
        and "·" in meta
    )

    if has_app_header:
        name = str(ws["A2"].value or safe_name).strip() or safe_name
        parts = [x.strip() for x in meta.split("·") if x.strip()]

        if len(parts) >= 1:
            project_type = parts[0]
        if len(parts) >= 2 and parts[1] in NIVELES_PRESUPUESTO:
            budget_level = parts[1]

        version_index = None
        for idx, value in enumerate(parts):
            match = re.fullmatch(r"V(\d+)", value, flags=re.I)
            if match:
                version_index = idx
                version = max(int(match.group(1)), 1)
                if idx >= 1:
                    project_code = parts[idx - 1]
                break

        if version_index is not None:
            start_location = (
                2
                if len(parts) >= 2 and parts[1] in NIVELES_PRESUPUESTO
                else 1
            )
            end_location = max(version_index - 1, start_location)
            location_parts = parts[start_location:end_location]
            if location_parts:
                location = " · ".join(location_parts)
        elif len(parts) >= 3:
            location = parts[2]

    return {
        "name": name,
        "project_type": project_type,
        "budget_level": budget_level,
        "location": location,
        "project_code": project_code,
        "version": version,
    }


def importar_presupuesto_excel(
    excel_bytes: bytes,
    fallback_params: dict,
    file_name: str = "",
) -> dict:
    """
    Reconstruye un presupuesto editable desde un .xlsx.

    Prioridades:
    1. 01 Presupuesto = alcance, cantidades e importes vigentes.
    2. 02 Control Interno = costos internos y parámetros, si existe.
    3. Si Control Interno no existe, se reconstruye el costo base a partir del
       importe interno y los porcentajes actuales.
    """
    if not excel_bytes:
        raise RuntimeError("El archivo está vacío.")

    try:
        wb_values = load_workbook(
            filename=BytesIO(excel_bytes),
            data_only=True,
            read_only=False,
        )
    except Exception as exc:
        raise RuntimeError("No fue posible leer el archivo Excel.") from exc

    ws, header_row = _buscar_hoja_presupuesto(wb_values)
    columns = _mapear_columnas_excel(ws, header_row)
    metadata = _leer_metadatos_excel(
        ws,
        fallback_name=file_name,
    )

    params, control_rows = _leer_parametros_control(
        wb_values,
        fallback_params,
    )

    raw_rows = []
    blank_streak = 0

    for row in range(header_row + 1, ws.max_row + 1):
        part_raw = ws.cell(row, columns["partida"]).value
        sub_raw = (
            ws.cell(row, columns["subpartida"]).value
            if columns.get("subpartida") else ""
        )
        desc_raw = ws.cell(row, columns["description"]).value
        unit_raw = ws.cell(row, columns["unit"]).value

        part_text = str(part_raw or "").strip()
        if "PRESUPUESTO INTERNO" in normalizar_texto(part_text).upper():
            break

        has_content = any(
            str(value or "").strip()
            for value in (part_raw, sub_raw, desc_raw, unit_raw)
        )
        if not has_content:
            blank_streak += 1
            if blank_streak >= 3 and raw_rows:
                break
            continue
        blank_streak = 0

        description = str(desc_raw or "").strip()
        unit = str(unit_raw or "").strip().upper()
        if not description or not unit:
            continue

        quantity = (
            _valor_float_excel(ws.cell(row, columns["quantity"]).value, 1.0)
            if columns.get("quantity")
            else 1.0
        )
        quantity = max(quantity, 0.0)

        amount = (
            _valor_float_excel(ws.cell(row, columns["amount"]).value, 0.0)
            if columns.get("amount")
            else 0.0
        )
        unit_price = (
            _valor_float_excel(ws.cell(row, columns["unit_price"]).value, 0.0)
            if columns.get("unit_price")
            else 0.0
        )

        if amount <= 0 and unit_price > 0 and quantity > 0:
            amount = unit_price * quantity
        if unit_price <= 0 and amount > 0 and quantity > 0:
            unit_price = amount / quantity

        raw_rows.append(
            {
                "code": (
                    str(ws.cell(row, columns["code"]).value or "").strip()
                    if columns.get("code") else ""
                ),
                "area": (
                    normalizar_nombre_area(ws.cell(row, columns["area"]).value)
                    if columns.get("area") else AREA_GENERAL
                ),
                "category": normalizar_seccion_comercial(_quitar_numeracion_excel(part_raw)),
                "subcategory": _quitar_numeracion_excel(sub_raw),
                "title": (
                    str(ws.cell(row, columns["title"]).value or "").strip()
                    if columns.get("title") else ""
                ),
                "description": description,
                "unit": unit,
                "quantity": quantity,
                "unit_sale": unit_price,
                "sale_amount": amount,
                "included": (
                    item_included({"included": ws.cell(row, columns["included"]).value})
                    if columns.get("included") else True
                ),
            }
        )

    if not raw_rows:
        raise RuntimeError(
            "No encontré conceptos utilizables dentro de la tabla del presupuesto."
        )

    # Descripción reconstruida: no inventa información; simplemente convierte
    # los renglones actuales en contexto para los ajustes posteriores con Gemini.
    description_lines = ["PRESUPUESTO RECARGADO DESDE EXCEL."]
    for row in raw_rows:
        description_lines.append(
            f"- [{row['area']} | {row['category']} / {row['subcategory']}] "
            f"{row['description']} | {row['quantity']:g} {row['unit']}"
        )

    if metadata["project_code"] == "IMPORTADO":
        metadata["project_code"] = (
            f"{abreviar_cliente(metadata['name'])}-IMP-0001"
        )

    imported_areas = []
    for row_data in raw_rows:
        area = normalizar_nombre_area(row_data.get("area") or AREA_GENERAL)
        if area != AREA_GENERAL and area not in imported_areas:
            imported_areas.append(area)

    project_data = {
        "name": metadata["name"],
        "project_type": metadata["project_type"],
        "budget_level": metadata["budget_level"],
        "location": metadata["location"],
        "dimension_mode": "Recuperadas de Excel",
        "dimensions_text": "",
        "description": "\n".join(description_lines),
        "guide_text": DEFAULT_GUIDE_TEXT,
        "areas_detectadas": [AREA_GENERAL] + imported_areas,
    }

    factor = (
        (1.0 + float(params["indirect_pct"]) / 100.0)
        * (1.0 + float(params["profit_pct"]) / 100.0)
    )
    if factor <= 0:
        factor = 1.0

    items = []
    for idx, row in enumerate(raw_rows):
        control = control_rows[idx] if idx < len(control_rows) else {}

        quantity = float(row["quantity"])
        sale_amount = float(row["sale_amount"])
        commercial_unit = (
            sale_amount / quantity
            if quantity > 0
            else float(row["unit_sale"] or 0.0)
        )

        unit_cost = float(control.get("unit_cost") or 0.0)
        if unit_cost <= 0:
            unit_cost = commercial_unit / factor if factor else commercial_unit

        code = str(row.get("code") or control.get("code") or "").strip()
        if not code:
            code = f"IMP-{idx + 1:03d}"

        title = str(row.get("title") or control.get("title") or "").strip()
        if not title:
            title = row["subcategory"] or re.split(
                r"[.;:]",
                row["description"],
                maxsplit=1,
            )[0][:80]

        direct_amount = quantity * unit_cost
        indirect_unit = unit_cost * params["indirect_pct"] / 100.0
        profit_unit = (
            unit_cost + indirect_unit
        ) * params["profit_pct"] / 100.0

        material_unit = float(control.get("material_unit") or 0.0)
        labor_unit = float(control.get("labor_unit") or 0.0)
        other_unit = float(control.get("other_unit") or 0.0)
        split_total = material_unit + labor_unit + other_unit

        if split_total > 0 and unit_cost > 0:
            material_share = material_unit / split_total * 100.0
            labor_share = labor_unit / split_total * 100.0
            other_share = other_unit / split_total * 100.0
        else:
            material_share = 0.0
            labor_share = 0.0
            other_share = 100.0

        benefit_amount = sale_amount - direct_amount
        margin = (
            benefit_amount / sale_amount * 100.0
            if sale_amount else 0.0
        )

        item = {
            "concept_id": None,
            "area": row.get("area") or AREA_GENERAL,
            "included": bool(row.get("included", True)),
            "package_group": str(control.get("package_group") or ""),
            "category": row["category"],
            "subcategory": row["subcategory"],
            "code": limpiar_codigo(code, f"IMP-{idx + 1:03d}"),
            "execution_order": (idx + 1) * 10,
            "commercial_title": title,
            "description": row["description"],
            "unit": row["unit"],
            "quantity": quantity,
            "unit_cost": unit_cost,
            "direct_amount": direct_amount,
            "unit_indirect": indirect_unit,
            "unit_profit": profit_unit,
            "unit_sale": commercial_unit,
            "original_unit_sale": float(control.get("original_unit_sale") or commercial_unit),
            "sale_amount": sale_amount,
            "benefit_amount": benefit_amount,
            "sale_margin_pct": margin,
            "price_source": "EXCEL_IMPORTADO",
            "price_source_detail": (
                "Precio recuperado del archivo Excel cargado por el usuario."
            ),
            "price_status": "IMPORTADO",
            "price_confidence": "Media",
            "price_evidence_score": 65,
            "material_share_pct": material_share,
            "labor_share_pct": labor_share,
            "other_share_pct": other_share,
            "waste_reference_pct": float(control.get("waste_pct") or 0.0),
            "quantity_confidence": "Alta",
            "quantity_criterion": "Cantidad recuperada del archivo Excel.",
            "inclusion_basis": "Concepto existente en el presupuesto importado.",
            "considerations": "",
        }
        item = aplicar_composicion_costo(item)
        items.append(item)

    activities = [item_a_actividad(item) for item in items]
    result = PresupuestoIA(
        nombre_proyecto=project_data["name"],
        actividad_principal=project_data["project_type"],
        alcance_resumido=(
            "Presupuesto reconstruido desde el archivo Excel cargado. "
            "Los conceptos existentes se conservan como alcance vigente."
        ),
        consideraciones_generales=[
            "Presupuesto recargado desde Excel para continuar con modificaciones."
        ],
        datos_faltantes=[],
        actividades=activities,
    )

    financials = calcular_financieros(items, params)

    excel_bytes_rebuilt = crear_excel(
        project_code=metadata["project_code"],
        project_data=project_data,
        result=result,
        items=items,
        params=params,
        version=metadata["version"],
    )

    return {
        "project_code": metadata["project_code"],
        "version": metadata["version"],
        "project_data": project_data,
        "params": params,
        "result": result,
        "items": items,
        "financials": financials,
        "excel_bytes": excel_bytes_rebuilt,
    }


# =========================================================
# EXCEL
# =========================================================


def crear_excel(
    project_code: str,
    project_data: dict,
    result: PresupuestoIA,
    items: list[dict],
    params: dict,
    version: int = 1,
) -> bytes:
    """
    V15.1: dos hojas.

    01 Presupuesto:
      Área | Partida | Subpartida | Descripción Técnica | Unidad | Cant. |
      Precio Unitario | Importe interno | Importe Final | Considerar

    02 Control Interno:
      conserva Código y Título comercial, además del control de subcontratación,
      margen, confianza, fuente y consideraciones.
    """
    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcOnSave = True

    # Paleta recuperada del formato anterior.
    brown = "4A342B"
    brown_mid = "6A4B3C"
    brown_light = "EDE7E3"
    beige = "F5F0EC"
    cream = "FFF8E7"
    gray = "E7E7E7"
    gray_light = "F5F5F5"
    formula_fill = "F2F2F2"
    white = "FFFFFF"
    dark_gray = "555555"
    thin_gray = Side(style="thin", color="D9D9D9")

    ordered_items = estructura_partidas_excel(items)
    for item in ordered_items:
        item["area"] = area_item(item)
        item["included"] = item_included(item)
        item["package_group"] = str(item.get("package_group") or "")
        item["original_unit_sale"] = original_unit_sale_item(item)

    # =====================================================
    # 01 PRESUPUESTO
    # =====================================================
    ws = wb.active
    ws.title = "01 Presupuesto"
    ws.sheet_view.showGridLines = False

    # Título como en el formato anterior: fondo blanco + texto café.
    ws.merge_cells("A1:J1")
    ws["A1"] = "PRESUPUESTO"
    ws["A1"].font = Font(size=16, bold=True, color=brown)

    ws.merge_cells("A2:J2")
    ws["A2"] = project_data.get("name") or "Cliente"
    ws["A2"].font = Font(size=11, bold=True, color=dark_gray)

    ws.merge_cells("A3:J3")
    ws["A3"] = (
        f"{project_data.get('project_type', '')} · "
        f"{project_data.get('budget_level', 'Medio-alto')} · "
        f"{project_data.get('location', '')} · {project_code} · V{version:02d}"
    )
    ws["A3"].font = Font(size=9, color=dark_gray)

    # ---------------- RESUMEN ----------------
    ws.merge_cells("A5:I5")
    ws["A5"] = "RESUMEN"
    ws["A5"].font = Font(bold=True, color=brown)

    summary_internal_row = 6
    summary_brand_row = 7
    summary_iva_row = 8
    summary_final_row = 9

    labels = [
        (6, "Presupuesto interno considerado (sin 30% / sin IVA)", gray),
        (7, f"Presupuesto + {BRAND_MARKUP_PCT:.0f}% marca (sin IVA)", beige),
        (8, f"IVA {params['iva_pct']:.0f}%", gray_light),
        (9, "Total final considerado", brown_light),
    ]
    for rr, label, fill in labels:
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
        ws.cell(rr, 1, label)
        ws.cell(rr, 1).font = Font(
            bold=True,
            color=brown if rr == 9 else dark_gray,
        )
        ws.cell(rr, 1).fill = PatternFill("solid", fgColor=fill)
        ws.cell(rr, 9).fill = PatternFill("solid", fgColor=fill)
        ws.cell(rr, 9).font = Font(
            bold=True,
            color=brown if rr == 9 else dark_gray,
        )
        ws.cell(rr, 9).number_format = '$#,##0.00'

    # ---------------- RESUMEN POR ÁREA ----------------
    areas = []
    for item in ordered_items:
        area = area_item(item)
        if area not in areas:
            areas.append(area)
    if AREA_GENERAL in areas:
        areas = [x for x in areas if x != AREA_GENERAL] + [AREA_GENERAL]
    if not areas:
        areas = [AREA_GENERAL]

    ws.merge_cells("A11:C11")
    ws["A11"] = "RESUMEN POR ÁREA"
    ws["A11"].font = Font(bold=True, color=brown)

    area_headers = ["Área", "Importe interno considerado", "Importe final considerado"]
    for col, header in enumerate(area_headers, 1):
        cell = ws.cell(12, col, header)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=brown_mid)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    area_rows = {}
    rr = 13
    for idx, area in enumerate(areas):
        area_rows[area] = rr
        ws.cell(rr, 1, area)
        fill = beige if idx % 2 == 0 else gray_light
        for col in range(1, 4):
            ws.cell(rr, col).fill = PatternFill("solid", fgColor=fill)
        rr += 1

    # ---------------- TABLA ----------------
    table_header_row = rr + 1
    headers = [
        "Área",
        "Partida",
        "Subpartida",
        "Descripción Técnica",
        "Unidad",
        "Cant.",
        "Precio Unitario (MXN)",
        "Importe interno (MXN)",
        "Importe Final (MXN)",
        "Considerar",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(table_header_row, col, header)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=brown)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    first_item_row = table_header_row + 1
    commercial_row_map = {}

    validation = DataValidation(
        type="list",
        formula1='"Sí,No"',
        allow_blank=False,
    )
    ws.add_data_validation(validation)

    for offset, item in enumerate(ordered_items):
        row = first_item_row + offset
        commercial_row_map[item["code"]] = row

        values = [
            area_item(item),
            item.get("partida_excel") or nombre_partida_excel(item.get("category")),
            item.get("subpartida_excel") or nombre_subpartida_excel(item),
            item["description"],
            item["unit"],
            float(item["quantity"]),
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)

        # H = Importe interno editable. G = P.U. derivado de H/F.
        ws.cell(row, 8, float(item.get("sale_amount") or 0.0))
        ws.cell(row, 7, f'=IF(F{row}=0,0,H{row}/F{row})')
        ws.cell(
            row,
            9,
            f'=IF(J{row}="Sí",H{row}*(1+{BRAND_MARKUP_PCT / 100.0:.6f})'
            f'*(1+{params["iva_pct"] / 100.0:.6f}),0)',
        )
        ws.cell(row, 10, "Sí" if item_included(item) else "No")
        validation.add(ws.cell(row, 10))

        ws.cell(row, 6).number_format = "0.00"
        for col in (7, 8, 9):
            ws.cell(row, col).number_format = '$#,##0.00'

        # Recuperar lectura visual del Excel anterior:
        # fórmula de P.U. gris, importe editable crema, final beige.
        ws.cell(row, 7).fill = PatternFill("solid", fgColor=formula_fill)
        ws.cell(row, 8).fill = PatternFill("solid", fgColor=cream)
        ws.cell(row, 9).fill = PatternFill("solid", fgColor=beige)
        ws.cell(row, 10).fill = PatternFill("solid", fgColor=cream)

        for col in range(1, 11):
            ws.cell(row, col).alignment = Alignment(
                vertical="top",
                wrap_text=col in {1, 2, 3, 4},
            )
            ws.cell(row, col).border = Border(bottom=thin_gray)

    last_item_row = first_item_row + len(ordered_items) - 1
    if not ordered_items:
        last_item_row = first_item_row

    # Resumen general.
    ws.cell(
        summary_internal_row,
        9,
        f'=SUMIF(J{first_item_row}:J{last_item_row},"Sí",H{first_item_row}:H{last_item_row})',
    )
    ws.cell(
        summary_brand_row,
        9,
        f'=I{summary_internal_row}*(1+{BRAND_MARKUP_PCT / 100.0:.6f})',
    )
    ws.cell(
        summary_iva_row,
        9,
        f'=I{summary_brand_row}*{params["iva_pct"] / 100.0:.6f}',
    )
    ws.cell(
        summary_final_row,
        9,
        f'=I{summary_brand_row}+I{summary_iva_row}',
    )

    # Resumen por área.
    for area, row in area_rows.items():
        safe = str(area).replace('"', '""')
        ws.cell(
            row,
            2,
            f'=SUMIFS(H{first_item_row}:H{last_item_row},'
            f'A{first_item_row}:A{last_item_row},"{safe}",'
            f'J{first_item_row}:J{last_item_row},"Sí")',
        )
        ws.cell(
            row,
            3,
            f'=SUMIFS(I{first_item_row}:I{last_item_row},'
            f'A{first_item_row}:A{last_item_row},"{safe}",'
            f'J{first_item_row}:J{last_item_row},"Sí")',
        )
        ws.cell(row, 2).number_format = '$#,##0.00'
        ws.cell(row, 3).number_format = '$#,##0.00'

    widths = [20, 27, 23, 72, 10, 11, 21, 22, 22, 13]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[table_header_row].height = 32
    if ordered_items:
        ws.auto_filter.ref = f"A{table_header_row}:J{last_item_row}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45

    # =====================================================
    # 02 CONTROL INTERNO + TRAZABILIDAD
    # =====================================================
    wc = wb.create_sheet("02 Control Interno")
    wc.sheet_view.showGridLines = False

    wc.merge_cells("A1:U1")
    wc["A1"] = "CONTROL INTERNO Y TRAZABILIDAD"
    wc["A1"].font = Font(size=15, bold=True, color=brown)

    control_headers = [
        "Código",
        "Área",
        "Partida",
        "Subpartida",
        "Título comercial",
        "Descripción",
        "Unidad",
        "Lote / Paquete",
        "Cant.",
        "Considerar",
        "Costo subcontratista unit.",
        "Costo subcontratista total",
        "P.U. interno original",
        "P.U. interno vigente",
        "Importe interno vigente",
        "Dif. P.U. vs original",
        "Beneficio",
        "Margen %",
        "Confianza",
        "Fuente precio",
        "Consideraciones",
    ]

    header_row = 3
    for col, header in enumerate(control_headers, 1):
        cell = wc.cell(header_row, col, header)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=brown)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for offset, item in enumerate(ordered_items):
        row = header_row + 1 + offset
        main_row = commercial_row_map[item["code"]]

        values = [
            item["code"],
            area_item(item),
            item.get("partida_excel") or nombre_partida_excel(item.get("category")),
            item.get("subpartida_excel") or nombre_subpartida_excel(item),
            titulo_comercial_item(item),
            item["description"],
            item["unit"],
            str(item.get("package_group") or ""),
        ]
        for col, value in enumerate(values, 1):
            wc.cell(row, col, value)

        # Enlaces con 01 después de retirar Código/Título de esa hoja.
        wc.cell(row, 9, f"='01 Presupuesto'!F{main_row}")
        wc.cell(row, 10, f"='01 Presupuesto'!J{main_row}")

        # Editable para negociación.
        wc.cell(row, 11, float(item.get("unit_cost") or 0.0))
        wc.cell(row, 12, f"=I{row}*K{row}")

        wc.cell(row, 13, original_unit_sale_item(item))
        wc.cell(row, 14, f"='01 Presupuesto'!G{main_row}")
        wc.cell(row, 15, f"='01 Presupuesto'!H{main_row}")
        wc.cell(row, 16, f"=N{row}-M{row}")
        wc.cell(row, 17, f"=O{row}-L{row}")
        wc.cell(row, 18, f'=IF(O{row}=0,0,Q{row}/O{row})')
        wc.cell(row, 19, confianza_general_item(item))
        wc.cell(row, 20, item.get("price_source") or "")
        wc.cell(row, 21, consideraciones_consolidadas_item(item))

        wc.cell(row, 9).number_format = "0.00"
        for col in (11, 12, 13, 14, 15, 16, 17):
            wc.cell(row, col).number_format = '$#,##0.00'
        wc.cell(row, 18).number_format = "0.00%"

        # Misma jerarquía visual del presupuesto anterior.
        wc.cell(row, 8).fill = PatternFill("solid", fgColor=cream)
        wc.cell(row, 11).fill = PatternFill("solid", fgColor=cream)
        for col in (12, 14, 15, 16, 17, 18):
            wc.cell(row, col).fill = PatternFill("solid", fgColor=formula_fill)

        # Confianza con semáforo muy discreto.
        confidence = confianza_general_item(item)
        conf_fill = {
            "Alta": "E8EFE3",
            "Media": "FFF2CC",
            "Baja": "F4CCCC",
        }.get(confidence, gray_light)
        wc.cell(row, 19).fill = PatternFill("solid", fgColor=conf_fill)

        for col in range(1, 22):
            wc.cell(row, col).alignment = Alignment(
                vertical="top",
                wrap_text=col in {2, 3, 4, 5, 6, 8, 20, 21},
            )
            wc.cell(row, col).border = Border(bottom=thin_gray)

    widths = [
        15, 20, 28, 22, 28, 58, 10, 20, 11, 12, 21,
        22, 20, 20, 21, 19, 19, 14, 13, 22, 70,
    ]
    for col, width in enumerate(widths, 1):
        wc.column_dimensions[get_column_letter(col)].width = width

    wc.row_dimensions[header_row].height = 34
    if ordered_items:
        wc.auto_filter.ref = f"A{header_row}:U{header_row + len(ordered_items)}"

    wc.page_setup.orientation = "landscape"
    wc.page_setup.fitToWidth = 1
    wc.page_setup.fitToHeight = 0
    wc.page_margins.left = 0.2
    wc.page_margins.right = 0.2
    wc.page_margins.top = 0.4
    wc.page_margins.bottom = 0.4

    out = BytesIO()
    wb.save(out)
    return out.getvalue()



def dataframe_resumen(items: list[dict]) -> pd.DataFrame:
    rows = []
    for x in estructura_partidas_excel(items):
        rows.append({
            "Área": area_item(x),
            "Partida": x["partida_excel"],
            "Subpartida": x["subpartida_excel"],
            "Unidad": x["unit"],
            "Cant.": x["quantity"],
            "Precio Unitario": x["unit_sale"],
            "Importe interno": x["sale_amount"],
            "Considerar": "Sí" if item_included(x) else "No",
        })
    return pd.DataFrame(rows)



# =========================================================
# API DE GEMINI
# =========================================================


def get_api_key_runtime() -> str | None:
    """La clave de empresa vive únicamente en Streamlit Secrets."""
    value = get_secret("GEMINI_API_KEY")
    return str(value).strip() if value else None


# =========================================================
# PANEL DE ADMINISTRACIÓN DE BASE INTERNA
# =========================================================


def _df_or_empty(rows: list[dict], columns: list[str] | None = None) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=columns or [])
    return pd.DataFrame(rows)



def render_admin_database(db: Database):
    st.header("Catálogo e historial")
    st.caption(
        "Consulta y administra conceptos, precios históricos, proyectos y presupuestos "
        "sin editar directamente las tablas de la base de datos."
    )

    if db.persistent:
        st.success("Base persistente PostgreSQL conectada.")
    else:
        st.warning(
            "SQLite local activo. Úselo únicamente para desarrollo. "
            "En Streamlit Community Cloud configure DATABASE_URL para persistencia."
        )

    try:
        stats = db.stats()
        m1, m2, m3 = st.columns(3)
        m1.metric("Proyectos", stats["projects"])
        m2.metric("Presupuestos", stats["budgets"])
        m3.metric("Conceptos", stats["concepts"])
    except Exception as exc:
        st.error(f"No fue posible consultar la base: {exc}")
        return

    st.divider()

    (
        tab_concepts,
        tab_prices,
        tab_projects,
        tab_budgets,
        tab_external,
        tab_maintenance,
        tab_export,
    ) = st.tabs(
        [
            "Conceptos",
            "Precios",
            "Proyectos",
            "Presupuestos",
            "Fuentes externas",
            "Mantenimiento",
            "Exportar",
        ]
    )

    source_labels = {
        "COTIZACION_PROVEEDOR": "Cotización de proveedor",
        "COSTO_REAL": "Costo real de obra",
        "REFERENCIA_EXTERNA": "Referencia externa",
        "REFERENCIA_CDMX": "Referencia CDMX",
        "HISTORICO_EXTERNO": "Histórico de referencia externa",
        "IA_ESTIMADO": "Estimación de IA",
        "MANUAL": "Registro manual",
        "BASE_INTERNA": "Base interna",
        "HISTORICO_IA": "Histórico generado por IA",
    }
    status_labels = {
        "VALIDADO": "Validado",
        "COTIZADO_PROVEEDOR": "Cotizado por proveedor",
        "COSTO_REAL": "Costo real",
        "REFERENCIA_EXTERNA": "Referencia externa",
        "ESTIMADO_IA": "Estimado por IA",
    }

    def friendly_source(value):
        return source_labels.get(str(value or "").upper(), value or "Sin fuente")

    def friendly_status(value):
        return status_labels.get(str(value or "").upper(), value or "Sin estado")

    # =====================================================
    # CONCEPTOS
    # =====================================================
    with tab_concepts:
        st.subheader("Catálogo de conceptos")
        st.caption(
            "Busque un concepto, abra su ficha y, si es necesario, modifique sus datos. "
            "Los precios se administran por separado en la pestaña Precios."
        )

        all_concepts = db.list_concepts("", limit=1000)
        categories = sorted(
            {str(c.get("category") or "").strip() for c in all_concepts if str(c.get("category") or "").strip()}
        )

        f1, f2 = st.columns([2, 1])
        with f1:
            search = st.text_input(
                "Buscar",
                placeholder="Ej. cancelería, demolición, pintura, PRE-01",
                key="catalog_search",
            )
        with f2:
            category_filter = st.selectbox(
                "Partida",
                ["Todas"] + categories,
                key="catalog_category",
            )

        concepts = db.list_concepts(search, limit=1000)
        if category_filter != "Todas":
            concepts = [
                c for c in concepts
                if str(c.get("category") or "").strip() == category_filter
            ]

        st.caption(f"{len(concepts)} concepto(s) encontrados.")

        if concepts:
            catalog_df = pd.DataFrame([
                {
                    "Código": c.get("code") or "",
                    "Concepto": c.get("description") or "",
                    "Partida": c.get("category") or "",
                    "Subpartida": c.get("subcategory") or "",
                    "Unidad": c.get("unit") or "",
                    "Último costo": c.get("latest_cost"),
                    "Fuente": friendly_source(c.get("latest_source")),
                    "Usos": int(c.get("usage_count") or 0),
                }
                for c in concepts
            ])
            st.dataframe(
                catalog_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Último costo": st.column_config.NumberColumn(format="$ %.2f"),
                    "Usos": st.column_config.NumberColumn(format="%d"),
                },
            )

            concept_map = {
                c["id"]: f"{c.get('code') or 'SIN-COD'} — {c.get('description') or ''}"
                for c in concepts
            }
            selected_id = st.selectbox(
                "Abrir concepto",
                options=list(concept_map.keys()),
                format_func=lambda x: concept_map[x],
                key="catalog_open_concept",
            )

            concept = db.get_concept(selected_id)
            if concept:
                usage = db.concept_usage(selected_id)
                prices = db.list_prices(selected_id)

                with st.container(border=True):
                    st.markdown(f"### {concept.get('description') or 'Concepto'}")
                    st.caption(f"Código: {concept.get('code') or 'Sin código'}")

                    d1, d2, d3 = st.columns(3)
                    d1.metric("Unidad", concept.get("unit") or "—")
                    d2.metric("Usos en presupuestos", int(usage.get("budget_items") or 0))
                    d3.metric("Precios registrados", int(usage.get("prices") or 0))

                    st.markdown("**Partida**")
                    st.write(concept.get("category") or "Sin partida")
                    st.markdown("**Subpartida**")
                    st.write(concept.get("subcategory") or "Sin subpartida")

                    if prices:
                        last = prices[0]
                        st.markdown("**Precio más reciente**")
                        st.write(
                            f"{formato_moneda(float(last['unit_cost']))} · "
                            f"{friendly_source(last.get('source'))} · "
                            f"{friendly_status(last.get('status'))}"
                        )
                    else:
                        st.info("Este concepto todavía no tiene precios registrados.")

                edit_key = f"editing_concept_{selected_id}"
                if edit_key not in st.session_state:
                    st.session_state[edit_key] = False

                b1, b2 = st.columns([1, 3])
                with b1:
                    if st.button(
                        "Editar concepto",
                        key=f"open_edit_{selected_id}",
                        use_container_width=True,
                    ):
                        st.session_state[edit_key] = True
                        st.rerun()
                with b2:
                    st.caption(
                        "Para agregar, revisar o eliminar costos históricos use la pestaña Precios."
                    )

                if st.session_state.get(edit_key):
                    with st.container(border=True):
                        st.markdown("#### Editar concepto")
                        st.caption(
                            "Modificar estos datos no altera los precios históricos ni los presupuestos ya generados."
                        )
                        with st.form(f"edit_concept_form_{selected_id}"):
                            ec1, ec2 = st.columns(2)
                            with ec1:
                                c_code = st.text_input(
                                    "Código",
                                    value=concept.get("code") or "",
                                )
                                c_category = st.text_input(
                                    "Partida",
                                    value=concept.get("category") or "",
                                )
                                c_subcategory = st.text_input(
                                    "Subpartida",
                                    value=concept.get("subcategory") or "",
                                )
                            with ec2:
                                c_unit = st.text_input(
                                    "Unidad",
                                    value=concept.get("unit") or "",
                                )
                                c_description = st.text_area(
                                    "Descripción",
                                    value=concept.get("description") or "",
                                    height=145,
                                )
                            save_col, cancel_col = st.columns(2)
                            save_edit = save_col.form_submit_button(
                                "Guardar cambios",
                                type="primary",
                                use_container_width=True,
                            )
                            cancel_edit = cancel_col.form_submit_button(
                                "Cancelar",
                                use_container_width=True,
                            )

                        if cancel_edit:
                            st.session_state[edit_key] = False
                            st.rerun()

                        if save_edit:
                            if not c_description.strip() or not c_unit.strip():
                                st.error("Descripción y unidad son obligatorias.")
                            else:
                                db.update_concept(
                                    selected_id,
                                    c_code,
                                    c_category,
                                    c_subcategory,
                                    c_description,
                                    c_unit,
                                )
                                st.session_state[edit_key] = False
                                st.success("Concepto actualizado.")
                                st.rerun()

                if prices:
                    st.markdown("#### Últimos precios")
                    recent_prices = prices[:5]
                    recent_df = pd.DataFrame([
                        {
                            "Costo": p["unit_cost"],
                            "Fuente": friendly_source(p.get("source")),
                            "Estado": friendly_status(p.get("status")),
                            "Confianza": p.get("confidence") or "",
                            "Fecha": p.get("created_at") or "",
                        }
                        for p in recent_prices
                    ])
                    st.dataframe(
                        recent_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Costo": st.column_config.NumberColumn(format="$ %.2f")
                        },
                    )

                with st.expander("Opciones avanzadas"):
                    st.warning(
                        "Eliminar un concepto es una acción administrativa. "
                        "Los presupuestos históricos conservan los datos de la actividad, "
                        "pero pueden perder el vínculo con el catálogo."
                    )
                    st.write(
                        f"Usos en presupuestos: {usage['budget_items']} · "
                        f"Precios históricos: {usage['prices']}"
                    )
                    confirm_concept = st.checkbox(
                        "Confirmo que deseo eliminar este concepto.",
                        key=f"confirm_delete_concept_{selected_id}",
                    )
                    if st.button(
                        "Eliminar concepto",
                        key=f"delete_concept_{selected_id}",
                        disabled=not confirm_concept,
                    ):
                        db.delete_concept(selected_id)
                        st.success("Concepto eliminado.")
                        st.rerun()
        else:
            st.info("No se encontraron conceptos con esos filtros.")

        st.divider()

        create_flag = "admin_create_concept_open"
        if create_flag not in st.session_state:
            st.session_state[create_flag] = False

        if not st.session_state[create_flag]:
            if st.button("Crear concepto nuevo", use_container_width=True):
                st.session_state[create_flag] = True
                st.rerun()
        else:
            with st.container(border=True):
                st.markdown("### Nuevo concepto")
                st.caption(
                    "Úselo para registrar manualmente una actividad que todavía no existe en el catálogo."
                )
                with st.form("catalog_create_concept"):
                    nc1, nc2 = st.columns(2)
                    with nc1:
                        a_code = st.text_input("Código", value="MAN-001")
                        a_category = st.text_input("Partida")
                        a_subcategory = st.text_input("Subpartida")
                    with nc2:
                        a_unit = st.text_input("Unidad", value="LOTE")
                        a_description = st.text_area("Descripción", height=145)

                    add_col, cancel_col = st.columns(2)
                    create_btn = add_col.form_submit_button(
                        "Crear concepto",
                        type="primary",
                        use_container_width=True,
                    )
                    cancel_create = cancel_col.form_submit_button(
                        "Cancelar",
                        use_container_width=True,
                    )

                if cancel_create:
                    st.session_state[create_flag] = False
                    st.rerun()

                if create_btn:
                    if not a_description.strip() or not a_unit.strip():
                        st.error("Descripción y unidad son obligatorias.")
                    else:
                        new_id = db.create_concept(
                            a_code,
                            a_category,
                            a_subcategory,
                            a_description,
                            a_unit,
                        )
                        st.session_state[create_flag] = False
                        st.success("Concepto creado. Puede agregarle un precio desde la pestaña Precios.")
                        st.rerun()

    # =====================================================
    # PRECIOS
    # =====================================================
    with tab_prices:
        st.subheader("Historial de precios")
        st.caption(
            "Los precios se guardan como registros históricos. Agregar un precio nuevo "
            "no elimina ni reemplaza automáticamente los anteriores."
        )

        concepts_for_prices = db.list_concepts("", limit=1500)

        if not concepts_for_prices:
            st.info("Primero debe existir al menos un concepto en el catálogo.")
        else:
            price_search = st.text_input(
                "Buscar concepto para administrar sus precios",
                placeholder="Ej. pintura, carpintería, cancel",
                key="price_concept_search",
            )
            if price_search.strip():
                filtered_price_concepts = db.list_concepts(price_search, limit=500)
            else:
                filtered_price_concepts = concepts_for_prices

            if not filtered_price_concepts:
                st.info("No se encontraron conceptos.")
            else:
                price_concept_map = {
                    c["id"]: f"{c.get('code') or 'SIN-COD'} — {c.get('description') or ''} [{c.get('unit') or ''}]"
                    for c in filtered_price_concepts
                }
                price_concept_id = st.selectbox(
                    "Concepto",
                    options=list(price_concept_map.keys()),
                    format_func=lambda x: price_concept_map[x],
                    key="price_selected_concept",
                )

                concept = db.get_concept(price_concept_id)
                prices = db.list_prices(price_concept_id)

                with st.container(border=True):
                    st.markdown(f"### {concept.get('description') or 'Concepto'}")
                    pinfo1, pinfo2, pinfo3 = st.columns(3)
                    pinfo1.metric("Código", concept.get("code") or "—")
                    pinfo2.metric("Unidad", concept.get("unit") or "—")
                    pinfo3.metric("Registros", len(prices))

                    st.caption(
                        f"{concept.get('category') or 'Sin partida'} / "
                        f"{concept.get('subcategory') or 'Sin subpartida'}"
                    )

                if prices:
                    price_df = pd.DataFrame([
                        {
                            "Costo unitario": p["unit_cost"],
                            "Fuente": friendly_source(p.get("source")),
                            "Estado": friendly_status(p.get("status")),
                            "Confianza": p.get("confidence") or "",
                            "Detalle": p.get("source_detail") or "",
                            "Fecha": p.get("created_at") or "",
                        }
                        for p in prices
                    ])
                    st.dataframe(
                        price_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Costo unitario": st.column_config.NumberColumn(format="$ %.2f")
                        },
                    )
                else:
                    st.info("No hay precios registrados para este concepto.")

                st.markdown("#### Agregar precio")
                st.caption(
                    "Registre una nueva referencia. El historial anterior se conserva."
                )

                source_options = [
                    "COTIZACION_PROVEEDOR",
                    "COSTO_REAL",
                    "REFERENCIA_EXTERNA",
                    "IA_ESTIMADO",
                    "MANUAL",
                ]
                status_options = [
                    "VALIDADO",
                    "COTIZADO_PROVEEDOR",
                    "COSTO_REAL",
                    "REFERENCIA_EXTERNA",
                    "ESTIMADO_IA",
                ]

                with st.form(f"add_price_form_{price_concept_id}"):
                    ap1, ap2 = st.columns(2)
                    with ap1:
                        new_cost = st.number_input(
                            "Costo unitario",
                            min_value=0.0,
                            step=100.0,
                            format="%.2f",
                        )
                        source_label_choice = st.selectbox(
                            "Origen del precio",
                            options=source_options,
                            format_func=lambda x: source_labels[x],
                        )
                        confidence = st.selectbox(
                            "Nivel de confianza",
                            ["Alta", "Media", "Baja"],
                        )
                    with ap2:
                        status_choice = st.selectbox(
                            "Estado",
                            options=status_options,
                            format_func=lambda x: status_labels[x],
                        )
                        detail = st.text_area(
                            "Referencia / proveedor / nota",
                            height=125,
                            placeholder="Ej. Cotización Proveedor X, agosto 2026.",
                        )
                    add_price = st.form_submit_button(
                        "Agregar al historial",
                        type="primary",
                        use_container_width=True,
                    )

                if add_price:
                    if new_cost <= 0:
                        st.error("El costo debe ser mayor que cero.")
                    else:
                        db.add_price(
                            price_concept_id,
                            new_cost,
                            source_label_choice,
                            detail,
                            status_choice,
                            confidence,
                        )
                        st.success("Precio agregado al historial.")
                        st.rerun()

                if prices:
                    with st.expander("Eliminar un precio registrado"):
                        st.caption(
                            "Utilícelo únicamente para eliminar registros incorrectos. "
                            "No es necesario borrar precios antiguos."
                        )
                        price_map = {
                            p["id"]: (
                                f"{formato_moneda(float(p['unit_cost']))} — "
                                f"{friendly_source(p.get('source'))} — "
                                f"{p.get('created_at') or ''}"
                            )
                            for p in prices
                        }
                        price_to_delete = st.selectbox(
                            "Registro",
                            options=list(price_map.keys()),
                            format_func=lambda x: price_map[x],
                            key=f"delete_price_select_{price_concept_id}",
                        )
                        confirm_price = st.text_input(
                            "Para eliminar, escriba ELIMINAR PRECIO",
                            key=f"delete_price_confirm_{price_concept_id}",
                        )
                        if st.button(
                            "Eliminar registro",
                            key=f"delete_price_button_{price_concept_id}",
                        ):
                            if confirm_price.strip().upper() != "ELIMINAR PRECIO":
                                st.error("Confirmación incorrecta.")
                            else:
                                db.delete_price(price_to_delete)
                                st.success("Precio eliminado.")
                                st.rerun()

    # =====================================================
    # PROYECTOS
    # =====================================================
    with tab_projects:
        st.subheader("Historial de proyectos")
        st.caption(
            "Consulte los proyectos que han generado presupuestos reales. "
            "Las simulaciones no aparecen aquí porque no se guardan."
        )

        projects = db.list_projects(limit=1000)
        project_search = st.text_input(
            "Buscar proyecto",
            placeholder="Código, cliente, ubicación o tipo de obra",
            key="project_history_search",
        )

        if project_search.strip():
            q = normalizar_texto(project_search)
            projects = [
                p for p in projects
                if q in normalizar_texto(
                    " ".join([
                        str(p.get("code") or ""),
                        str(p.get("name") or ""),
                        str(p.get("location") or ""),
                        str(p.get("project_type") or ""),
                    ])
                )
            ]

        if not projects:
            st.info("No se encontraron proyectos.")
        else:
            project_df = pd.DataFrame([
                {
                    "Código": p.get("code") or "",
                    "Cliente": p.get("name") or "",
                    "Tipo": p.get("project_type") or "",
                    "Ubicación": p.get("location") or "",
                    "Presupuestos": int(p.get("budget_count") or 0),
                    "Último total": p.get("latest_total"),
                    "Fecha": p.get("created_at") or "",
                }
                for p in projects
            ])
            st.dataframe(
                project_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Último total": st.column_config.NumberColumn(format="$ %.2f")
                },
            )

            project_map = {
                p["id"]: f"{p.get('code') or ''} — {p.get('name') or ''}"
                for p in projects
            }
            project_id = st.selectbox(
                "Abrir proyecto",
                options=list(project_map.keys()),
                format_func=lambda x: project_map[x],
                key="open_project_history",
            )
            project = db.get_project(project_id)

            if project:
                project_budgets = db.list_budgets(project_id=project_id)

                with st.container(border=True):
                    st.markdown(f"### {project.get('name') or 'Cliente'}")
                    st.caption(project.get("code") or "")

                    pr1, pr2, pr3 = st.columns(3)
                    pr1.metric("Tipo", project.get("project_type") or "—")
                    pr2.metric("Ubicación", project.get("location") or "—")
                    pr3.metric("Presupuestos", len(project_budgets))

                    if project.get("main_activity"):
                        st.markdown("**Actividad principal**")
                        st.write(project.get("main_activity"))
                    if project.get("dimensions_text"):
                        st.markdown("**Dimensiones / referencias**")
                        st.write(project.get("dimensions_text"))
                    if project.get("description"):
                        st.markdown("**Descripción inicial**")
                        st.write(project.get("description"))

                if project_budgets:
                    st.markdown("#### Presupuestos del proyecto")
                    pb_df = pd.DataFrame([
                        {
                            "Versión": b.get("version"),
                            "Estado": b.get("status"),
                            "Costo directo": b.get("direct_cost"),
                            "Venta sin IVA": b.get("sale_before_tax"),
                            "Total": b.get("total"),
                            "Fecha": b.get("created_at"),
                        }
                        for b in project_budgets
                    ])
                    st.dataframe(
                        pb_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Costo directo": st.column_config.NumberColumn(format="$ %.2f"),
                            "Venta sin IVA": st.column_config.NumberColumn(format="$ %.2f"),
                            "Total": st.column_config.NumberColumn(format="$ %.2f"),
                        },
                    )

                edit_project_key = f"editing_project_{project_id}"
                if edit_project_key not in st.session_state:
                    st.session_state[edit_project_key] = False

                if st.button(
                    "Editar datos del proyecto",
                    key=f"edit_project_button_{project_id}",
                ):
                    st.session_state[edit_project_key] = True
                    st.rerun()

                if st.session_state.get(edit_project_key):
                    with st.container(border=True):
                        st.markdown("#### Editar proyecto")
                        st.caption(
                            "Esta edición corrige los datos descriptivos del proyecto. "
                            "No recalcula presupuestos existentes."
                        )
                        with st.form(f"edit_project_form_{project_id}"):
                            ep1, ep2 = st.columns(2)
                            with ep1:
                                p_name = st.text_input(
                                    "Nombre del cliente",
                                    value=project.get("name") or "",
                                )
                                p_type = st.text_input(
                                    "Tipo de obra",
                                    value=project.get("project_type") or "",
                                )
                                p_location = st.text_input(
                                    "Ubicación",
                                    value=project.get("location") or "",
                                )
                                p_main_activity = st.text_input(
                                    "Actividad principal",
                                    value=project.get("main_activity") or "",
                                )
                            with ep2:
                                p_dimensions = project.get("dimensions_text") or ""
                                p_description = st.text_area(
                                    "Descripción",
                                    value=project.get("description") or "",
                                    height=120,
                                )
                                p_guide = st.text_area(
                                    "Texto guía",
                                    value=project.get("guide_text") or "",
                                    height=90,
                                )
                            save_pc, cancel_pc = st.columns(2)
                            save_project = save_pc.form_submit_button(
                                "Guardar cambios",
                                type="primary",
                                use_container_width=True,
                            )
                            cancel_project = cancel_pc.form_submit_button(
                                "Cancelar",
                                use_container_width=True,
                            )

                        if cancel_project:
                            st.session_state[edit_project_key] = False
                            st.rerun()

                        if save_project:
                            if not p_name.strip():
                                st.error("El nombre del proyecto es obligatorio.")
                            else:
                                db.update_project(
                                    project_id,
                                    p_name,
                                    p_type,
                                    p_location,
                                    p_main_activity,
                                    p_dimensions,
                                    p_description,
                                    p_guide,
                                )
                                st.session_state[edit_project_key] = False
                                st.success("Proyecto actualizado.")
                                st.rerun()

                with st.expander("Opciones avanzadas"):
                    st.warning(
                        "Eliminar el proyecto borra sus presupuestos, partidas y los conceptos/precios "
                        "creados exclusivamente por ese proyecto."
                    )
                    confirm_project = st.checkbox(
                        "Confirmo que deseo eliminar este proyecto completo.",
                        key=f"confirm_delete_project_{project_id}",
                    )
                    if st.button(
                        "Eliminar proyecto completo",
                        key=f"delete_project_button_{project_id}",
                        disabled=not confirm_project,
                    ):
                        db.delete_project(project_id)
                        st.success("Proyecto y su trazabilidad asociada fueron eliminados.")
                        st.rerun()

    # =====================================================
    # PRESUPUESTOS
    # =====================================================
    with tab_budgets:
        st.subheader("Historial de presupuestos")
        st.caption(
            "Consulte presupuestos ya guardados. Esta sección es de consulta; "
            "las correcciones comerciales siguen realizándose en el Excel generado."
        )

        budgets = db.list_budgets(limit=1000)

        budget_search = st.text_input(
            "Buscar presupuesto",
            placeholder="Código de proyecto, nombre o ubicación",
            key="budget_history_search",
        )
        if budget_search.strip():
            q = normalizar_texto(budget_search)
            budgets = [
                b for b in budgets
                if q in normalizar_texto(
                    " ".join([
                        str(b.get("project_code") or ""),
                        str(b.get("project_name") or ""),
                        str(b.get("project_location") or ""),
                    ])
                )
            ]

        if not budgets:
            st.info("No se encontraron presupuestos.")
        else:
            budget_df = pd.DataFrame([
                {
                    "Código": b.get("project_code") or "",
                    "Cliente": b.get("project_name") or "",
                    "Estado": b.get("status") or "",
                    "Costo directo": b.get("direct_cost"),
                    "Indirectos": b.get("indirect_cost"),
                    "Utilidad": b.get("profit"),
                    "Venta sin IVA": b.get("sale_before_tax"),
                    "Total": b.get("total"),
                    "Fecha": b.get("created_at") or "",
                }
                for b in budgets
            ])
            st.dataframe(
                budget_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Costo directo": st.column_config.NumberColumn(format="$ %.2f"),
                    "Indirectos": st.column_config.NumberColumn(format="$ %.2f"),
                    "Utilidad": st.column_config.NumberColumn(format="$ %.2f"),
                    "Venta sin IVA": st.column_config.NumberColumn(format="$ %.2f"),
                    "Total": st.column_config.NumberColumn(format="$ %.2f"),
                },
            )

            budget_map = {
                b["id"]: (
                    f"{b.get('project_code') or ''} — "
                    f"{b.get('project_name') or ''} — "
                    f"{formato_moneda(float(b.get('total') or 0))}"
                )
                for b in budgets
            }
            budget_id = st.selectbox(
                "Abrir presupuesto",
                options=list(budget_map.keys()),
                format_func=lambda x: budget_map[x],
                key="open_budget_history",
            )

            budget = db.get_budget(budget_id)
            items = db.list_budget_items(budget_id)

            if budget:
                with st.container(border=True):
                    st.markdown(
                        f"### {budget.get('project_code') or ''} — "
                        f"{budget.get('project_name') or ''}"
                    )
                    bm1, bm2, bm3, bm4 = st.columns(4)
                    bm1.metric("Costo directo", formato_moneda(float(budget.get("direct_cost") or 0)))
                    bm2.metric("Indirectos", formato_moneda(float(budget.get("indirect_cost") or 0)))
                    bm3.metric("Utilidad", formato_moneda(float(budget.get("profit") or 0)))
                    bm4.metric("Total", formato_moneda(float(budget.get("total") or 0)))

                    st.caption(
                        f"Indirectos: {float(budget.get('indirect_pct') or 0):.2f}% · "
                        f"Utilidad: {float(budget.get('profit_pct') or 0):.2f}% · "
                        f"IVA: {float(budget.get('iva_pct') or 0):.2f}% · "
                        f"Estado: {budget.get('status') or ''}"
                    )

                if items:
                    st.markdown("#### Actividades")
                    item_df = pd.DataFrame([
                        {
                            "Sección": i.get("category") or "",
                            "Concepto": i.get("commercial_title") or i.get("subcategory") or "",
                            "Código": i.get("code") or "",
                            "Descripción": i.get("description") or "",
                            "Unidad": i.get("unit") or "",
                            "Cantidad": i.get("quantity"),
                            "Costo unitario": i.get("unit_cost"),
                            "Venta unitaria": i.get("unit_sale"),
                            "Venta": i.get("sale_amount"),
                            "Fuente": friendly_source(i.get("price_source")),
                        }
                        for i in items
                    ])
                    st.dataframe(
                        item_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Costo unitario": st.column_config.NumberColumn(format="$ %.2f"),
                            "Venta unitaria": st.column_config.NumberColumn(format="$ %.2f"),
                            "Venta": st.column_config.NumberColumn(format="$ %.2f"),
                        },
                    )
                else:
                    st.info("Este presupuesto no contiene actividades registradas.")

                with st.expander("Opciones avanzadas"):
                    st.warning(
                        "Eliminar un presupuesto elimina sus actividades e historial vinculado. "
                        "Si es el único presupuesto del proyecto, también puede eliminarse el proyecto."
                    )
                    confirm_budget = st.checkbox(
                        "Confirmo que deseo eliminar este presupuesto.",
                        key=f"confirm_delete_budget_{budget_id}",
                    )
                    if st.button(
                        "Eliminar presupuesto",
                        key=f"delete_budget_button_{budget_id}",
                        disabled=not confirm_budget,
                    ):
                        db.delete_budget(budget_id)
                        st.success("Presupuesto eliminado.")
                        st.rerun()

    # =====================================================
    # FUENTES EXTERNAS
    # =====================================================
    with tab_external:
        st.subheader("Fuentes externas")
        st.caption(
            "Catálogos de referencia independientes de la base histórica de la empresa. "
            "Por ahora esta versión integra únicamente el Tabulador General de Precios "
            "Unitarios del Gobierno de la Ciudad de México."
        )

        active_cdmx = db.get_active_external_catalog("CDMX")

        with st.container(border=True):
            st.markdown("### Gobierno de la Ciudad de México")
            if active_cdmx:
                ec1, ec2, ec3 = st.columns(3)
                ec1.metric("Edición activa", active_cdmx.get("version_label") or "—")
                ec2.metric(
                    "Conceptos",
                    f"{int(active_cdmx.get('concept_count') or 0):,}",
                )
                ec3.metric(
                    "Importado",
                    str(active_cdmx.get("imported_at") or "")[:10] or "—",
                )
                st.caption(
                    "La app consulta automáticamente esta edición cuando no encuentra "
                    "un precio interno suficientemente confiable."
                )
            else:
                st.warning(
                    "Todavía no existe un catálogo CDMX importado. Hasta que se cargue, "
                    "los presupuestos continuarán utilizando la base interna y Gemini."
                )

            st.info(
                "El P.U. CDMX se trata como referencia integrada antes de IVA. "
                "Cuando la coincidencia es suficientemente alta, la app calcula un "
                "costo base equivalente para conservar los indirectos y utilidad "
                "configurados en el presupuesto sin duplicarlos."
            )

            if st.button(
                "Buscar e importar la edición más reciente",
                type="primary",
                use_container_width=True,
                key="update_cdmx_catalog",
            ):
                with st.spinner(
                    "Consultando la fuente oficial, descargando el PDF y construyendo el catálogo..."
                ):
                    try:
                        update = actualizar_catalogo_cdmx(db)
                        if update["status"] == "already_current":
                            st.success(
                                "El catálogo ya corresponde a la edición oficial más reciente "
                                f"detectada: {update['catalog']['version_label']} "
                                f"({update['concept_count']:,} conceptos)."
                            )
                        else:
                            st.success(
                                f"Catálogo CDMX actualizado a {update['catalog']['version_label']} "
                                f"con {update['concept_count']:,} conceptos."
                            )
                        st.rerun()
                    except Exception as exc:
                        st.error(
                            "No se modificó el catálogo existente porque la actualización falló."
                        )
                        st.exception(exc)

        if active_cdmx:
            st.markdown("#### Consultar catálogo importado")
            ext_search = st.text_input(
                "Buscar por clave o descripción",
                placeholder="Ej. pintura, demolición, impermeabilización",
                key="external_cdmx_search",
            )
            external_rows = db.search_external_concepts(
                "CDMX",
                ext_search,
                limit=300,
            )
            if external_rows:
                ext_df = pd.DataFrame(
                    [
                        {
                            "Clave": r["source_code"],
                            "Concepto": r["description"],
                            "Unidad": r["unit"],
                            "P.U. CDMX": r["unit_price"],
                            "Edición": r["version_label"],
                        }
                        for r in external_rows
                    ]
                )
                st.dataframe(
                    ext_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "P.U. CDMX": st.column_config.NumberColumn(format="$ %.2f")
                    },
                )
                if len(external_rows) >= 300:
                    st.caption(
                        "Se muestran como máximo 300 resultados. Use una búsqueda más específica."
                    )
            else:
                st.info("No se encontraron conceptos con esa búsqueda.")

            with st.expander("Eliminar catálogo CDMX importado"):
                st.warning(
                    "Esto no elimina presupuestos históricos. Solo elimina el catálogo externo "
                    "que se utiliza para nuevas búsquedas."
                )
                confirm_ext_delete = st.checkbox(
                    "Confirmo que deseo eliminar el catálogo CDMX importado.",
                    key="confirm_delete_external_cdmx",
                )
                if st.button(
                    "Eliminar catálogo CDMX",
                    key="delete_external_cdmx",
                    disabled=not confirm_ext_delete,
                ):
                    db.delete_external_source("CDMX")
                    st.success("Catálogo CDMX eliminado.")
                    st.rerun()

    # =====================================================
    # MANTENIMIENTO
    # =====================================================
    with tab_maintenance:
        st.subheader("Mantenimiento")
        st.caption(
            "Herramientas para la etapa de pruebas. Las acciones de esta sección "
            "afectan datos persistentes. Por ahora no se utiliza clave de eliminación."
        )

        stats_now = db.stats()
        with st.container(border=True):
            st.markdown("### Estado actual")
            sm1, sm2, sm3 = st.columns(3)
            sm1.metric("Proyectos", stats_now["projects"])
            sm2.metric("Presupuestos", stats_now["budgets"])
            sm3.metric("Conceptos", stats_now["concepts"])

        try:
            latest = db.get_latest_project_record()
        except Exception as exc:
            latest = None
            st.error(f"No fue posible consultar el último proyecto: {exc}")

        with st.container(border=True):
            st.markdown("### Último proyecto guardado")
            if latest:
                st.write(f"**{latest.get('code') or ''} — {latest.get('name') or ''}**")
                st.caption(
                    f"{latest.get('project_type') or ''} · "
                    f"{latest.get('location') or 'Sin ubicación'} · "
                    f"{latest.get('created_at') or ''}"
                )
                if latest.get("latest_total") is not None:
                    st.write(f"Último total: **{formato_moneda(float(latest['latest_total']))}**")
                st.caption(
                    "Para borrar este proyecto sin ser administrador también existe "
                    "la herramienta de corrección en la sección Generar presupuesto."
                )
            else:
                st.info("La base no contiene proyectos.")

        with st.container(border=True):
            st.markdown("### Eliminar todos los datos de la aplicación")
            st.error(
                "Esta acción borra proyectos, presupuestos, actividades, conceptos e historial "
                "de precios. La estructura de las tablas se conserva para que la aplicación "
                "pueda seguir funcionando."
            )
            wipe_confirm = st.checkbox(
                "Confirmo que deseo vaciar toda la base de datos de la aplicación.",
                key="maintenance_wipe_confirm",
            )
            if st.button(
                "Eliminar todos los datos",
                type="primary",
                use_container_width=True,
                disabled=not wipe_confirm,
                key="maintenance_wipe_database",
            ):
                db.clear_all_data()
                st.session_state.pop("generated", None)
                st.success("Todos los datos de la aplicación fueron eliminados.")
                st.rerun()

    # =====================================================
    # EXPORTAR
    # =====================================================
    with tab_export:
        st.subheader("Exportar información")
        st.caption(
            "Descarga copias CSV para revisión, respaldo o análisis. "
            "Estas descargas no modifican la base de datos."
        )

        export_items = [
            ("concepts", "Catálogo de conceptos", "conceptos.csv"),
            ("price_history", "Historial de precios", "historial_precios.csv"),
            ("projects", "Proyectos", "proyectos.csv"),
            ("budgets", "Presupuestos", "presupuestos.csv"),
            ("budget_items", "Actividades de presupuestos", "actividades_presupuestos.csv"),
            ("external_catalogs", "Catálogos externos", "catalogos_externos.csv"),
            ("external_concepts", "Conceptos externos", "conceptos_externos.csv"),
        ]

        for table_name, label, filename in export_items:
            with st.container(border=True):
                c1, c2 = st.columns([3, 1])
                with c1:
                    st.markdown(f"**{label}**")
                    st.caption(filename)
                rows = db.export_table(table_name)
                csv_bytes = pd.DataFrame(rows).to_csv(index=False).encode("utf-8-sig")
                with c2:
                    st.download_button(
                        "Descargar CSV",
                        data=csv_bytes,
                        file_name=filename,
                        mime="text/csv",
                        key=f"friendly_download_{table_name}",
                        use_container_width=True,
                    )


# =========================================================
# ESTADO DE APLICACIÓN
# =========================================================


st.title("Generador de presupuestos")

database_url = get_secret("DATABASE_URL")
try:
    db, db_error = get_database(database_url, DATABASE_CACHE_VERSION)

    required_database_methods = (
        "get_latest_project_record",
        "clear_all_data",
        "delete_generation",
        "save_generation",
        "save_revision",
        "get_active_external_catalog",
        "external_candidates",
        "replace_external_catalog",
    )
    if any(not hasattr(db, method) for method in required_database_methods):
        st.cache_resource.clear()
        db, db_error = get_database(database_url, DATABASE_CACHE_VERSION)
except Exception as exc:
    st.error("No fue posible conectar con la base de datos.")
    st.exception(exc)
    st.stop()

with st.sidebar:
    st.header("Navegación")
    section = st.radio(
        "Sección",
        ["Generar presupuesto", "Catálogo e historial"],
        key="main_section",
        label_visibility="collapsed",
    )

    if section == "Generar presupuesto":
        st.divider()
        st.header("Parámetros")

        indirect_pct = st.number_input(
            "Indirectos (%)",
            min_value=0.0,
            max_value=100.0,
            value=12.0,
            step=0.5,
            key="indirect_pct",
        )
        profit_pct = st.number_input(
            "Utilidad (%)",
            min_value=0.0,
            max_value=100.0,
            value=10.0,
            step=0.5,
            key="profit_pct",
        )
        iva_pct = st.number_input(
            "IVA (%)",
            min_value=0.0,
            max_value=100.0,
            value=16.0,
            step=1.0,
            key="iva_pct",
        )
        waste_pct = st.number_input(
            "Desperdicio (%)",
            min_value=0.0,
            max_value=50.0,
            value=5.0,
            step=0.5,
            key="waste_pct",
        )

        with st.expander("Configuración"):
            model_name = st.text_input(
                "Modelo Gemini",
                value="gemini-3.8-flash",
                key="model_name",
            )

        st.divider()
        if st.button("Reiniciar página", use_container_width=True):
            st.session_state.clear()
            st.rerun()

# =========================================================
# BASE INTERNA
# =========================================================


if section == "Catálogo e historial":
    render_admin_database(db)
    st.stop()


# =========================================================
# FORMULARIO INICIAL
# =========================================================


DESCRIPTION_EXAMPLE = """Ejemplo:

SEGUNDA PLANTA
Recámara principal de aproximadamente 4.20 x 3.80 m.
- Retiro de piso laminado existente.
- Colocación de piso nuevo.
- Reparación y pintura de muros.

AZOTEA
Área aproximada de 9 x 4 m.
- Limpieza y preparación de superficie.
- Impermeabilización completa.
- Revisión de bajadas pluviales.
"""

DEFAULT_GUIDE_TEXT = """- Considerar protección básica de las áreas de trabajo y zonas de tránsito.
- Incluir limpieza durante los trabajos y limpieza final.
- En pintura, considerar preparación básica, resanes menores, sellador cuando sea necesario y dos manos de pintura.
- En instalaciones y elementos nuevos, considerar suministro, colocación, fijaciones y conexiones, cuando correspondan.
- Mantener materiales y acabados coherentes con el nivel de presupuesto seleccionado.
"""

ADJUSTMENT_EXAMPLE = """Ejemplos:
- Falta considerar limpieza fina al terminar la obra.
- El precio de la pintura me parece muy bajo, revísalo.
- La cantidad de impermeabilización debe ser mayor.
- Cambia la cancelería a aluminio línea pesada con cristal templado.
"""


def volver_a_entrada():
    """
    Regresa al formulario conservando exactamente los datos con los que se
    generó el presupuesto.

    Se utiliza un estado intermedio independiente de los widgets porque
    Streamlit puede retirar del session_state los valores de widgets que dejan
    de renderizarse mientras se muestra la pantalla de resultados.
    """
    generated = st.session_state.get("generated")
    if generated:
        project_data = generated.get("project_data") or {}

        st.session_state["_restore_project_form"] = {
            "client_name": project_data.get("name", ""),
            "project_location": project_data.get("location", ""),
            "project_type": project_data.get(
                "project_type",
                "Remodelación interior general",
            ),
            "budget_level": project_data.get("budget_level", "Medio-alto"),
            "project_description": project_data.get("description", ""),
            "guide_text": (
                project_data.get("guide_text")
                if project_data.get("guide_text") is not None
                else DEFAULT_GUIDE_TEXT
            ),
        }

    st.session_state.pop("generated", None)
    st.rerun()


if "generated" not in st.session_state:
    # Restauración explícita al volver desde un presupuesto ya generado.
    # Debe ocurrir ANTES de crear los widgets del formulario.
    restore_data = st.session_state.pop("_restore_project_form", None)
    if restore_data:
        for field_key, field_value in restore_data.items():
            st.session_state[field_key] = field_value

    if "guide_text" not in st.session_state:
        st.session_state["guide_text"] = DEFAULT_GUIDE_TEXT

    # -----------------------------------------------------
    # RECARGAR PRESUPUESTO EXISTENTE
    # -----------------------------------------------------
    uploaded_budget = st.file_uploader(
        "Cargar presupuesto Excel",
        type=["xlsx"],
        key="reload_budget_excel",
    )

    if uploaded_budget is not None:
        if st.button(
            "Cargar y continuar editando",
            type="primary",
            use_container_width=True,
            key="load_existing_budget",
        ):
            fallback_params = {
                "indirect_pct": float(indirect_pct),
                "profit_pct": float(profit_pct),
                "iva_pct": float(iva_pct),
                "waste_pct": float(waste_pct),
            }

            with st.spinner("Leyendo presupuesto..."):
                try:
                    imported = importar_presupuesto_excel(
                        uploaded_budget.getvalue(),
                        fallback_params=fallback_params,
                        file_name=uploaded_budget.name,
                    )

                    st.session_state["generated"] = {
                        "project_id": None,
                        "budget_id": None,
                        "saved": False,
                        "pending_revision": False,
                        "project_code": imported["project_code"],
                        "version": imported["version"],
                        "project_data": imported["project_data"],
                        "params": imported["params"],
                        "result": imported["result"].model_dump(),
                        "items": imported["items"],
                        "financials": imported["financials"],
                        "excel_bytes": imported["excel_bytes"],
                        "revision_history": [],
                        "pending_revision_notes": [],
                        "imported_from_excel": True,
                    }
                    st.rerun()
                except Exception as exc:
                    st.exception(exc)

    st.divider()

    with st.form("form_proyecto"):
        f1, f2 = st.columns(2)
        with f1:
            client_name = st.text_input(
                "Nombre del cliente",
                placeholder="Ej. Desarrollos de la Vega",
                key="client_name",
            )
        with f2:
            location = st.text_input(
                "Ubicación",
                placeholder="Ej. Coyoacán, CDMX",
                key="project_location",
            )

        project_type = st.selectbox(
            "Tipo de obra",
            [
                "Remodelación interior general",
                "Baño",
                "Cocina",
                "Recámara",
                "Sala / comedor",
                "Local comercial",
                "Oficina",
                "Caseta / acceso",
                "Otro",
            ],
            key="project_type",
        )

        budget_level = st.selectbox(
            "Nivel de presupuesto",
            NIVELES_PRESUPUESTO,
            index=2,
            key="budget_level",
        )

        description = st.text_area(
            "Descripción general de trabajos",
            placeholder=DESCRIPTION_EXAMPLE,
            height=430,
            key="project_description",
        )

        guide_text = st.text_area(
            "Texto guía",
            height=300,
            key="guide_text",
        )

        generate = st.form_submit_button(
            "Generar presupuesto",
            type="primary",
            use_container_width=True,
        )

    if generate:
        api_key = get_api_key_runtime()
        if not api_key:
            st.error("Falta GEMINI_API_KEY en Streamlit Secrets.")
            st.stop()
        if not client_name.strip():
            st.error("Ingrese el nombre del cliente.")
            st.stop()
        if not location.strip():
            st.error("Ingrese la ubicación.")
            st.stop()
        if not description.strip():
            st.error("Ingrese la descripción general de los trabajos.")
            st.stop()

        params = {
            "indirect_pct": float(indirect_pct),
            "profit_pct": float(profit_pct),
            "iva_pct": float(iva_pct),
            "waste_pct": float(waste_pct),
        }
        project_data = {
            "name": client_name.strip(),
            "project_type": project_type,
            "budget_level": budget_level,
            "location": location.strip(),
            "dimension_mode": "Integradas en descripción",
            "dimensions_text": "",
            "description": description.strip(),
            "guide_text": guide_text.strip(),
        }

        with st.spinner("Generando presupuesto..."):
            try:
                project_data["areas_detectadas"] = detectar_areas_proyecto_ia(
                    api_key=api_key, model_name=model_name, project_data=project_data
                )
                result = generar_presupuesto_ia(
                    api_key=api_key, model_name=model_name, project_data=project_data, params=params
                )
                result = auditar_estructura_presupuesto_ia(
                    api_key=api_key, model_name=model_name, project_data=project_data, result=result
                )
                result = verificar_cantidades_python(result)
                items = resolver_items(
                    db, result, project_data, params, api_key=api_key, model_name=model_name
                )
                if not items:
                    raise RuntimeError("La IA no generó actividades utilizables.")

                financials = calcular_financieros(items, params)
                provisional_code = db.next_project_code(
                    project_data["name"],
                    project_data["location"],
                )
                excel_bytes = crear_excel(
                    project_code=provisional_code,
                    project_data=project_data,
                    result=result,
                    items=items,
                    params=params,
                    version=1,
                )

                st.session_state["generated"] = {
                    "project_id": None,
                    "budget_id": None,
                    "saved": False,
                    "pending_revision": False,
                    "project_code": provisional_code,
                    "version": 1,
                    "project_data": project_data,
                    "params": params,
                    "result": result.model_dump(),
                    "items": items,
                    "financials": financials,
                    "excel_bytes": excel_bytes,
                    "revision_history": [],
                    "pending_revision_notes": [],
                }
                st.rerun()
            except Exception as exc:
                st.exception(exc)


# =========================================================
# RESULTADO
# =========================================================


else:
    g = st.session_state["generated"]
    result = PresupuestoIA.model_validate(g["result"])
    items = g["items"]
    financials = g["financials"]
    version = int(g.get("version") or 1)
    saved = bool(g.get("saved"))

    st.subheader(g["project_code"])
    if saved:
        st.caption(f"Guardado · V{version:02d}")
    elif g.get("project_id"):
        st.caption(f"Cambios sin guardar · próxima versión V{version:02d}")
    elif g.get("imported_from_excel"):
        st.caption("Presupuesto recargado desde Excel · cambios sin guardar")
    else:
        st.caption("Borrador sin guardar")

    p1, p2, p3, p4 = st.columns(4)
    with p1:
        st.text_input(
            "Cliente",
            value=g["project_data"]["name"],
            disabled=True,
            key=f"locked_client_{version}_{saved}",
        )
    with p2:
        st.text_input(
            "Ubicación",
            value=g["project_data"]["location"],
            disabled=True,
            key=f"locked_location_{version}_{saved}",
        )
    with p3:
        st.text_input(
            "Tipo de obra",
            value=g["project_data"]["project_type"],
            disabled=True,
            key=f"locked_type_{version}_{saved}",
        )
    with p4:
        st.text_input(
            "Nivel",
            value=g["project_data"].get("budget_level", "Medio-alto"),
            disabled=True,
            key=f"locked_level_{version}_{saved}",
        )

    st.text_area(
        "Descripción general de trabajos",
        value=g["project_data"]["description"],
        height=220,
        disabled=True,
        key=f"locked_description_{version}_{saved}",
    )
    st.text_area(
        "Texto guía",
        value=g["project_data"]["guide_text"] or "",
        height=160,
        disabled=True,
        key=f"locked_guide_{version}_{saved}",
    )

    # Resumen comercial igual al utilizado en 01 Presupuesto:
    # 1) presupuesto interno;
    # 2) presupuesto con 30 % de marca;
    # 3) total final después de aplicar IVA al presupuesto con marca.
    presupuesto_interno = float(financials["sale_before_tax"])
    presupuesto_con_marca = presupuesto_interno * (1.0 + BRAND_MARKUP_PCT / 100.0)
    presupuesto_final_iva = presupuesto_con_marca * (
        1.0 + float(g["params"]["iva_pct"]) / 100.0
    )

    m1, m2, m3 = st.columns(3)
    m1.metric(
        "Presupuesto interno",
        formato_moneda(presupuesto_interno),
    )
    m2.metric(
        f"Presupuesto + {BRAND_MARKUP_PCT:.0f}% marca",
        formato_moneda(presupuesto_con_marca),
    )
    m3.metric(
        "Presupuesto final con IVA",
        formato_moneda(presupuesto_final_iva),
    )

    with st.expander("Detalle interno"):
        i1, i2, i3 = st.columns(3)
        i1.metric("Costo directo", formato_moneda(financials["direct_cost"]))
        i2.metric("Indirectos", formato_moneda(financials["indirect_cost"]))
        i3.metric("Utilidad", formato_moneda(financials["profit"]))

    df = dataframe_resumen(items)
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Cant.": st.column_config.NumberColumn(format="%.2f"),
            "Precio Unitario": st.column_config.NumberColumn(format="$ %.2f"),
            "Importe interno": st.column_config.NumberColumn(format="$ %.2f"),
        },
    )

    if result.consideraciones_generales or result.datos_faltantes:
        with st.expander("Consideraciones"):
            for item in result.consideraciones_generales:
                st.write(f"- {item}")
            if result.datos_faltantes:
                st.markdown("**Datos por confirmar**")
                for item in result.datos_faltantes:
                    st.write(f"- {item}")

    file_status = f"V{version:02d}" if g.get("project_id") else "BORRADOR"
    st.download_button(
        "Descargar Excel",
        data=g["excel_bytes"],
        file_name=f"{g['project_code']}-{file_status}_Presupuesto.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # -----------------------------------------------------
    # Ajuste sencillo con IA
    # -----------------------------------------------------
    st.divider()
    st.subheader("Ajustar con IA")

    adjustment_request = st.text_area(
        "¿Qué quieres revisar, agregar o cambiar?",
        placeholder=ADJUSTMENT_EXAMPLE,
        height=180,
        key=f"adjustment_request_{version}_{saved}",
    )

    if st.button(
        "Aplicar ajuste",
        type="primary",
        use_container_width=True,
        key=f"apply_adjustment_{version}_{saved}",
    ):
        if not adjustment_request.strip():
            st.error("Escriba el cambio que desea realizar.")
        else:
            api_key = get_api_key_runtime()
            if not api_key:
                st.error("Falta GEMINI_API_KEY en Streamlit Secrets.")
            else:
                with st.spinner("Aplicando ajuste..."):
                    try:
                        revision = revisar_presupuesto_ia(
                            api_key=api_key,
                            model_name=model_name,
                            project_data=g["project_data"],
                            params=g["params"],
                            current_result=result,
                            current_items=items,
                            revision_request=adjustment_request.strip(),
                        )

                        revised_result, revised_items, change_log = aplicar_revision_estructural(
                            db=db, current_result=result, current_items=items, revision=revision,
                            project_data=g["project_data"], params=g["params"],
                            api_key=api_key, model_name=model_name,
                        )
                        revised_result = auditar_estructura_presupuesto_ia(
                            api_key=api_key, model_name=model_name,
                            project_data=g["project_data"], result=revised_result,
                        )
                        revised_result = verificar_cantidades_python(revised_result)
                        revised_items = sincronizar_items_con_estructura(
                            revised_result,
                            revised_items,
                        )
                        revised_financials = calcular_financieros(
                            revised_items,
                            g["params"],
                        )

                        # Si el proyecto ya existe en la base, el ajuste queda como
                        # borrador de la siguiente versión hasta que el usuario lo guarde.
                        if g.get("project_id"):
                            target_version = version if g.get("pending_revision") else version + 1
                            pending_revision = True
                        else:
                            target_version = 1
                            pending_revision = False

                        excel_bytes = crear_excel(
                            project_code=g["project_code"],
                            project_data=g["project_data"],
                            result=revised_result,
                            items=revised_items,
                            params=g["params"],
                            version=target_version,
                        )

                        history = list(g.get("revision_history") or [])
                        history.append(
                            {
                                "request": adjustment_request.strip(),
                                "summary": revision.resumen_revision,
                                "changes": change_log,
                            }
                        )
                        pending_notes = list(g.get("pending_revision_notes") or [])
                        if g.get("project_id"):
                            pending_notes.append(adjustment_request.strip())

                        g.update(
                            {
                                "saved": False,
                                "pending_revision": pending_revision,
                                "version": target_version,
                                "result": revised_result.model_dump(),
                                "items": revised_items,
                                "financials": revised_financials,
                                "excel_bytes": excel_bytes,
                                "revision_history": history,
                                "pending_revision_notes": pending_notes,
                            }
                        )
                        st.session_state["generated"] = g
                        st.rerun()
                    except Exception as exc:
                        st.exception(exc)

    history = g.get("revision_history") or []
    if history:
        with st.expander("Ajustes realizados"):
            for i, rev in enumerate(history, start=1):
                st.markdown(f"**Ajuste {i}: {rev['summary']}**")
                st.caption(rev["request"])
                for change in rev.get("changes") or []:
                    st.write(f"- {change}")

    # -----------------------------------------------------
    # Acciones del proyecto
    # -----------------------------------------------------
    st.divider()
    a1, a2, a3 = st.columns(3)

    with a1:
        save_label = "Guardado en base" if saved else (
            "Guardar nueva versión" if g.get("project_id") else "Guardar en base"
        )
        if st.button(
            save_label,
            type="primary" if not saved else "secondary",
            use_container_width=True,
            disabled=saved,
            key=f"save_project_{version}_{saved}",
        ):
            try:
                if not g.get("project_id"):
                    real_code = db.next_project_code(
                        g["project_data"]["name"],
                        g["project_data"]["location"],
                    )
                    project_id, budget_id = db.save_generation(
                        project_code=real_code,
                        project_data=g["project_data"],
                        result=result,
                        items=items,
                        params=g["params"],
                        financials=financials,
                    )
                    real_version = 1
                else:
                    budget_id, real_version = db.save_revision(
                        project_id=g["project_id"],
                        parent_budget_id=g["budget_id"],
                        result=result,
                        items=items,
                        params=g["params"],
                        financials=financials,
                        revision_instruction=(
                            "\n\n".join(g.get("pending_revision_notes") or [])
                            or (history[-1]["request"] if history else "Ajuste del presupuesto")
                        ),
                    )
                    project_id = g["project_id"]
                    real_code = g["project_code"]

                excel_bytes = crear_excel(
                    project_code=real_code,
                    project_data=g["project_data"],
                    result=result,
                    items=items,
                    params=g["params"],
                    version=real_version,
                )

                clean_items = []
                for saved_item in items:
                    cleaned = dict(saved_item)
                    cleaned.pop("record_new_price", None)
                    clean_items.append(cleaned)

                g.update(
                    {
                        "project_id": project_id,
                        "budget_id": budget_id,
                        "project_code": real_code,
                        "version": real_version,
                        "saved": True,
                        "pending_revision": False,
                        "items": clean_items,
                        "excel_bytes": excel_bytes,
                        "pending_revision_notes": [],
                    }
                )
                st.session_state["generated"] = g
                st.rerun()
            except Exception as exc:
                st.exception(exc)

    with a2:
        if st.button(
            "Editar entrada inicial",
            use_container_width=True,
            key=f"edit_initial_{version}_{saved}",
        ):
            volver_a_entrada()

    with a3:
        if st.button(
            "Eliminar proyecto",
            use_container_width=True,
            key=f"delete_project_result_{version}_{saved}",
        ):
            try:
                if g.get("project_id"):
                    db.delete_project(g["project_id"])
                st.session_state.pop("generated", None)
                st.rerun()
            except Exception as exc:
                st.exception(exc)
